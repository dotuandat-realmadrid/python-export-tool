import sqlite3
import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from utils import wrap_text
from database import DB_NAME
from collections import defaultdict
import re
class AddProduct:
    def __init__(self, root, product_id=None, parent=None):
        self.root = root
        self.product_id = product_id
        self.parent = parent
        self.root.title("Chỉnh sửa sản phẩm" if product_id else "Thêm sản phẩm mới")
        self.root.state('zoomed')
       
        self.root.grid_columnconfigure(1, weight=1)
        self.root.grid_rowconfigure(5, weight=1)
       
        tk.Label(root, text="Tên sản phẩm:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.product_name_entry = tk.Entry(root, width=50, font=("Arial", 12))
        self.product_name_entry.grid(row=0, column=1, padx=10, pady=5, sticky="ew")
        tk.Button(root, text="Lưu sản phẩm", command=self.save_product).grid(row=0, column=2, padx=5, pady=5)
       
        self.restore_button = tk.Button(root, text="Khôi phục", command=self.restore_indicators, width=10)
        self.restore_button.grid(row=1, column=2, padx=5, pady=5)
       
        tk.Label(root, text="Chọn loại sản phẩm:").grid(row=1, column=0, padx=5, pady=5, sticky="w")
        self.type_combobox = ttk.Combobox(root, font=("Arial", 12), state="readonly")
        self.type_combobox.grid(row=1, column=1, padx=10, pady=5, sticky="ew")
        self.selected_type = None # SỬA: Chỉ lưu 1 loại sản phẩm (thay vì list selected_types)
        self.type_combobox.bind("<<ComboboxSelected>>", self.on_type_selected)
       
        # Button frame cho sản phẩm tham khảo
        reference_button_frame = tk.Frame(root)
        reference_button_frame.grid(row=2, column=1, padx=5, pady=5, sticky="w")
        # Nút thêm sản phẩm tham khảo
        self.add_reference_button = tk.Button(reference_button_frame,
                                            text="Thêm sản phẩm tham khảo",
                                            command=self.add_new_reference_row,
                                            font=("Arial", 10))
        self.add_reference_button.pack(side=tk.LEFT, padx=(0, 5))
        # Nút xóa sản phẩm tham khảo
        self.delete_reference_button = tk.Button(reference_button_frame,
                                                text="Xóa sản phẩm tham khảo",
                                                command=self.delete_reference_product,
                                                font=("Arial", 10))
        self.delete_reference_button.pack(side=tk.LEFT, padx=(0, 5))
       
        self.add_button = tk.Button(root, text="+", command=self.add_custom_indicator, width=10)
        self.add_button.grid(row=2, column=2, padx=5, pady=5)
        self.add_button.grid_remove()
       
        self.notebook = ttk.Notebook(root)
        self.notebook.grid(row=5, column=0, columnspan=3, sticky="nsew", padx=5, pady=5)
       
        self.three_brands_frame = tk.Frame(self.notebook)
        self.dmkt_frame = tk.Frame(self.notebook)
        self.bom_frame = tk.Frame(self.notebook)
        self.ctkt_bo_frame = tk.Frame(self.notebook)
        self.ctkt_mua_sam_frame = tk.Frame(self.notebook)
       
        self.notebook.add(self.ctkt_mua_sam_frame, text="CTKT mua sắm")
        self.notebook.add(self.three_brands_frame, text="Hãng")
        self.notebook.add(self.bom_frame, text="BOM")
        self.notebook.add(self.dmkt_frame, text="DMKT")
        self.notebook.add(self.ctkt_bo_frame, text="CTKT bộ")
        # Mặc định chọn tab Hãng
        self.notebook.select(self.three_brands_frame)
       
        self.max_row_heights = {
            "three_brands": 80,
            "bom": 80,
            "dmkt": 80,
            "ctkt_bo": 80,
            "ctkt_mua_sam": 25
        }
        self.column_widths = {
            "three_brands": {
                "Mã chỉ tiêu": 200, "Chỉ tiêu": 200, "Đánh giá": 100, "Giá trị": 150,
                "Giá trị tham khảo": 150, "Đơn vị": 100, "Loại chỉ tiêu": 150, "Hành động": 100
            },
            "bom": {
                "Mã chỉ tiêu": 300, "Chỉ tiêu": 300, "Giá trị": 150,
                "Đơn vị": 100, "Loại chỉ tiêu": 150, "Hành động": 100
            },
            "dmkt": {
                "Mã chỉ tiêu": 200, "Yêu cầu kỹ thuật": 200, "Giá trị": 150, "Đơn vị": 100,
                "Hành động": 100
            },
            "ctkt_bo": {
                "Mã chỉ tiêu": 50, "Chỉ tiêu": 200, "Đơn vị": 100, "Giá trị": 300, "Hành động": 100
            },
            "ctkt_mua_sam": {
                "Mã chỉ tiêu": 250, "Chỉ tiêu kỹ thuật chi tiết": 200,
                "Giá trị": 150, "Đơn vị": 100, "Tiêu chí đánh giá": 300,
                "Loại chỉ tiêu": 100, "Hành động": 100
            }
        }
        self.custom_indicators = {}
        self.deleted_indicators = {}
        self.origin_deleted = {}
        self.yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        self.blue_fill = PatternFill(start_color="00A6FF", end_color="00A6FF", fill_type="solid")
       
        self.init_treeview(self.three_brands_frame, "three_brands")
        self.init_treeview(self.dmkt_frame, "dmkt")
        self.init_treeview(self.bom_frame, "bom")
        self.init_treeview(self.ctkt_bo_frame, "ctkt_bo")
        self.init_treeview(self.ctkt_mua_sam_frame, "ctkt_mua_sam")
       
        self.notebook.bind("<<NotebookTabChanged>>", self.on_tab_changed)
       
        self.root.grid_columnconfigure(0, weight=0)
        self.root.grid_columnconfigure(1, weight=1)
        self.root.grid_columnconfigure(2, weight=0)
        self.root.grid_rowconfigure(5, weight=1)
       
        self.reference_products = [] # (man_id, name, product_name)
        self.tab_data = {
            "three_brands": [],
            "bom": [],
            "dmkt": [],
            "ctkt_bo": [],
            "ctkt_mua_sam": []
        }
       
        self.custom_rows_ctkt_ms = []
        self.current_entry = None
        self.current_tree = None
        self.reference_columns = {}
       
        self.load_types()
        if product_id:
            self.load_product_data()
        else:
            # SỬA: Không init default crit_type ở đây, sẽ init khi chọn type
            pass
        self.on_tab_changed(None)
   
    def init_default_crit_type(self):
        """
        GIẢI THÍCH: Khởi tạo crit_type mặc định là "CTCB" cho tất cả indicators của loại sản phẩm đã chọn
        Chạy khi chọn hoặc load type
        """
        if not self.selected_type:
            return
       
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
       
        # SỬA: Lấy type_id từ self.selected_type
        type_id_result = c.execute("SELECT id FROM product_types WHERE name=?", (self.selected_type,)).fetchone()
        if not type_id_result:
            conn.close()
            return
        type_id = type_id_result[0]
       
        # GIẢI THÍCH: Lấy indicators chỉ của type_id này
        indicators = c.execute("SELECT id FROM indicators WHERE type_id=?", (type_id,)).fetchall()
       
        # GIẢI THÍCH: Khởi tạo custom_indicators cho tất cả tab
        for tab_name in ["three_brands", "bom", "dmkt", "ctkt_bo", "ctkt_mua_sam"]:
            self.custom_indicators[tab_name] = {}
            for ind_id in [i[0] for i in indicators]:
                # GIẢI THÍCH: Set crit_type mặc định là "CTCB"
                self.custom_indicators[tab_name][f"crit_type_{ind_id}"] = "CTCB"
       
        conn.close()
   
    def init_treeview(self, frame, tab_name):
        tree_frame = tk.Frame(frame)
        tree_frame.pack(fill=tk.BOTH, expand=True)
        tree_frame.grid_columnconfigure(0, weight=1)
        tree_frame.grid_rowconfigure(0, weight=1)
       
        v_scrollbar = tk.Scrollbar(tree_frame, orient="vertical")
        v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        h_scrollbar = tk.Scrollbar(tree_frame, orient="horizontal")
        h_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)
       
        tree = ttk.Treeview(tree_frame, show="headings",
                            xscrollcommand=h_scrollbar.set,
                            yscrollcommand=v_scrollbar.set)
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
       
        h_scrollbar.config(command=tree.xview)
        v_scrollbar.config(command=tree.yview)
       
        style = ttk.Style()
        style.theme_use('clam')
       
        style.configure("Custom.Treeview",
                        background="#FFFFFF",
                        foreground="#000000",
                        fieldbackground="#FFFFFF",
                        bordercolor="#C0C0C0",
                        lightcolor="#E0E0E0",
                        darkcolor="#A0A0A0",
                        borderwidth=2,
                        rowheight=self.max_row_heights[tab_name])
       
        style.configure("Custom.Treeview.Heading",
                        background="#E0E0E0",
                        foreground="#000000",
                        bordercolor="#A0A0A0",
                        borderwidth=2,
                        relief="solid")
       
        style.configure("DataRow.Treeview",
                        background="#FFFFFF",
                        fieldbackground="#FFFFFF",
                        bordercolor="#C0C0C0",
                        lightcolor="#E0E0E0",
                        darkcolor="#A0A0A0",
                        borderwidth=2,
                        rowheight=self.max_row_heights[tab_name])
       
        tree.configure(style="Custom.Treeview")
       
        tree.tag_configure("blue", background="#00A6FF", foreground="#FFFFFF")
        tree.tag_configure("new_reference", background="#FFFFE0")
       
        tree.bind("<Double-1>", self.on_double_click)
        tree.bind("<Button-1>", self.on_click_action)
        if tab_name == "three_brands":
            tree.bind("<Return>", self.add_reference_from_tree)
       
        setattr(self, f"{tab_name}_tree", tree)
        setattr(self, f"{tab_name}_indicator_map", {})
   
    def on_tab_changed(self, event):
        selected_tab = self.notebook.select()
        if selected_tab:
            tab_text = self.notebook.tab(selected_tab, "text")
            map_dict = {
                "Hãng": "three_brands",
                "BOM": "bom",
                "DMKT": "dmkt",
                "CTKT bộ": "ctkt_bo",
                "CTKT mua sắm": "ctkt_mua_sam"
            }
            if tab_text in map_dict:
                tab_name = map_dict[tab_text]
                self.current_tab = tab_name
                self.current_tree = getattr(self, f"{tab_name}_tree")
                self.current_indicator_map = getattr(self, f"{tab_name}_indicator_map")
               
                for col in self.current_tree["columns"]:
                    self.current_tree.column(col, width=self.column_widths[tab_name].get(col, 200))
               
                if tab_name == "ctkt_mua_sam":
                    row_height = 120
                    self.add_button.grid()
                else:
                    row_height = self.max_row_heights[tab_name]
                    self.add_button.grid_remove()
               
                style = ttk.Style()
                style.configure("Custom.Treeview", rowheight=row_height)
                style.configure("DataRow.Treeview", rowheight=row_height)
            else:
                self.current_tab = None
                self.current_tree = None
                self.current_indicator_map = {}
                self.add_button.grid_remove()
        else:
            self.current_tab = None
            self.current_tree = None
            self.current_indicator_map = {}
            self.add_button.grid_remove()
   
    def save_product(self):
        """
        Lưu thông tin sản phẩm vào database
        - SỬA MỚI: Kiểm tra trùng tên sản phẩm trong cùng product_type
        - Lưu thông tin cơ bản của sản phẩm
        - Lưu mapping với product_types
        - Lưu reference_products (sản phẩm tham khảo)
        - Lưu product_hidden_indicators (chỉ tiêu đã xóa)
        - Lưu product_custom_indicators (giá trị tùy chỉnh, bao gồm yêu cầu khác)
        """
        name = self.product_name_entry.get().strip()
        if not name:
            messagebox.showerror("Lỗi", "Nhập tên sản phẩm")
            return
        if not self.selected_type:
            messagebox.showerror("Lỗi", "Chọn loại sản phẩm")
            return
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        type_id_result = c.execute("SELECT id FROM product_types WHERE name=?", (self.selected_type,)).fetchone()
        if not type_id_result:
            messagebox.showerror("Lỗi", f"Loại sản phẩm '{self.selected_type}' không hợp lệ")
            conn.close()
            return
        type_id = type_id_result[0]
        try:
            # SỬA MỚI: Kiểm tra trùng tên sản phẩm trong cùng product_type
            # Lấy tất cả product_id có cùng type_id và cùng tên
            existing_products = c.execute("""
                SELECT p.id, p.name
                FROM products p
                JOIN product_type_mapping_products ptmp ON p.id = ptmp.product_id
                WHERE ptmp.type_id = ? AND p.name = ?
            """, (type_id, name)).fetchall()
           
            # Nếu đang tạo mới (product_id = None) và đã có sản phẩm trùng tên
            if self.product_id is None and existing_products:
                messagebox.showerror("Lỗi",
                    f"Đã tồn tại sản phẩm '{name}' trong loại '{self.selected_type}'!\n"
                    f"Không thể tạo 2 sản phẩm trùng tên trong cùng loại sản phẩm.")
                conn.close()
                return
           
            # Nếu đang sửa (product_id != None) và có sản phẩm khác trùng tên
            if self.product_id is not None:
                duplicate_products = [p for p in existing_products if p[0] != self.product_id]
                if duplicate_products:
                    messagebox.showerror("Lỗi",
                        f"Đã tồn tại sản phẩm '{name}' trong loại '{self.selected_type}'!\n"
                        f"Không thể đổi tên thành tên trùng với sản phẩm khác trong cùng loại.")
                    conn.close()
                    return
            # Lưu bảng products
            if self.product_id is None:
                c.execute("INSERT INTO products (name, ma_san_pham, note) VALUES (?, ?, ?)", (name, None, None))
                self.product_id = c.lastrowid
            else:
                c.execute("UPDATE products SET name=? WHERE id=?", (name, self.product_id))
            # Lưu mapping loại sản phẩm
            c.execute("DELETE FROM product_type_mapping_products WHERE product_id=?", (self.product_id,))
            c.execute("INSERT INTO product_type_mapping_products (product_id, type_id) VALUES (?, ?)", (self.product_id, type_id))
            # Lưu reference_products
            c.execute("DELETE FROM reference_products WHERE product_id=?", (self.product_id,))
            for sort_order, (man_id, _, _) in enumerate(self.reference_products):
                c.execute("INSERT INTO reference_products (product_id, manufacturer_id, sort_order) VALUES (?, ?, ?)",
                        (self.product_id, man_id, sort_order))
            # Lưu product_hidden_indicators
            c.execute("DELETE FROM product_hidden_indicators WHERE product_id=?", (self.product_id,))
            valid_tabs = ["three_brands", "bom", "dmkt", "ctkt_bo", "ctkt_mua_sam"]
            for tab_name in valid_tabs:
                if tab_name in self.origin_deleted:
                    for ind_id in self.origin_deleted[tab_name]:
                        c.execute("INSERT INTO product_hidden_indicators (product_id, tab_name, indicator_id) VALUES (?, ?, ?)",
                                (self.product_id, tab_name, ind_id))
            # Lưu product_custom_indicators
            c.execute("DELETE FROM product_custom_indicators WHERE product_id=?", (self.product_id,))
            for tab_name, customs in self.custom_indicators.items():
                for key, value in customs.items():
                    if key and value:
                        try:
                            parts = key.split('_')
                            if len(parts) >= 2:
                                prefix = parts[0]
                                ind_id_str = '_'.join(parts[1:])
                                if ind_id_str.lstrip('-').isdigit():
                                    ind_id_save = int(ind_id_str)
                                else:
                                    ind_id_save = ind_id_str
                                custom_value = f"{prefix}_{value}"
                                unique_indicator_id = f"{ind_id_save}_{prefix}"
                                c.execute("""
                                    INSERT INTO product_custom_indicators (product_id, tab_name, indicator_id, custom_value)
                                    VALUES (?, ?, ?, ?)
                                """, (self.product_id, tab_name, unique_indicator_id, custom_value))
                        except Exception as e:
                            print(f"Lỗi khi lưu {key}: {str(e)}")
                            continue
            conn.commit()
            messagebox.showinfo("Thành công", "Đã cập nhật sản phẩm")
            if self.parent:
                self.parent.load_products()
        except Exception as e:
            conn.rollback()
            messagebox.showerror("Lỗi", f"Lỗi khi lưu sản phẩm: {str(e)}")
        finally:
            conn.close()

    def load_product_data(self):
        """
        Load dữ liệu sản phẩm từ database
        - Load thông tin cơ bản
        - Load loại sản phẩm
        - Load reference_products
        - Load product_hidden_indicators
        - Load product_custom_indicators (bao gồm yêu cầu khác)
        """
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        product = c.execute("SELECT name FROM products WHERE id=?", (self.product_id,)).fetchone()
        if not product:
            conn.close()
            return
        self.product_name_entry.insert(0, product[0])
        selected_type = c.execute('''
            SELECT pt.name FROM product_type_mapping_products ptm
            JOIN product_types pt ON ptm.type_id = pt.id
            WHERE ptm.product_id=?
        ''', (self.product_id,)).fetchone()
        if selected_type:
            self.selected_type = selected_type[0]
            self.update_type_combobox_selection()
        references = c.execute('''
            SELECT m.id, m.name, m.product_name
            FROM reference_products rp
            JOIN manufacturers m ON rp.manufacturer_id = m.id
            WHERE rp.product_id=? ORDER BY rp.sort_order
        ''', (self.product_id,)).fetchall()
        self.reference_products = [(r[0], r[1], r[2]) for r in references]
        # Hidden indicators
        self.deleted_indicators = {}
        self.origin_deleted = {}
        hidden_records = c.execute("SELECT tab_name, indicator_id FROM product_hidden_indicators WHERE product_id=?",
                                (self.product_id,)).fetchall()
        cascade_map = {
            "three_brands": ["bom", "dmkt", "ctkt_bo", "ctkt_mua_sam"],
            "bom": ["dmkt", "ctkt_bo"],
            "dmkt": ["ctkt_bo"],
            "ctkt_bo": [],
            "ctkt_mua_sam": []
        }
        for tab_name, ind_id in hidden_records:
            if tab_name in cascade_map:
                self.origin_deleted.setdefault(tab_name, set()).add(ind_id)
                self.deleted_indicators.setdefault(tab_name, set()).add(ind_id)
                for dep in cascade_map[tab_name]:
                    self.deleted_indicators.setdefault(dep, set()).add(ind_id)
        # Custom indicators
        all_custom = c.execute(
            "SELECT tab_name, indicator_id, custom_value FROM product_custom_indicators WHERE product_id=?",
            (self.product_id,)
        ).fetchall()
        self.custom_indicators = {t: {} for t in ["three_brands", "bom", "dmkt", "ctkt_bo", "ctkt_mua_sam"]}
        for tab_name, indicator_id_composite, custom_value in all_custom:
            try:
                if isinstance(indicator_id_composite, str) and '_' in indicator_id_composite:
                    parts = indicator_id_composite.rsplit('_', 1)
                    if len(parts) == 2:
                        ind_id_str, prefix = parts
                        if custom_value.startswith(f"{prefix}_"):
                            actual_value = custom_value[len(f"{prefix}_"):]
                            key = f"{prefix}_{ind_id_str}"
                            self.custom_indicators[tab_name][key] = actual_value
            except Exception as e:
                print(f"Lỗi khi xử lý {indicator_id_composite}: {str(e)}")
        # Custom rows cho CTKT mua sắm
        self.custom_rows_ctkt_ms = []
        for key in self.custom_indicators.get("ctkt_mua_sam", {}):
            if key.startswith("chi_tieu_") and key.split("_")[-1].startswith("custom_"):
                custom_id = "_".join(key.split("_")[2:])
                row_data = {
                    "id": custom_id,
                    "chi_tieu": self.custom_indicators["ctkt_mua_sam"].get(f"chi_tieu_{custom_id}", ""),
                    "yeu_cau": self.custom_indicators["ctkt_mua_sam"].get(f"yeu_cau_{custom_id}", ""),
                    "so_sanh": self.custom_indicators["ctkt_mua_sam"].get(f"so_sanh_{custom_id}", ""),
                    "don_vi": self.custom_indicators["ctkt_mua_sam"].get(f"don_vi_{custom_id}", ""),
                    "tieu_chi": self.custom_indicators["ctkt_mua_sam"].get(f"tieu_chi_{custom_id}", ""),
                    "crit_type": self.custom_indicators["ctkt_mua_sam"].get(f"crit_type_{custom_id}", "CTCB")
                }
                self.custom_rows_ctkt_ms.append(row_data)
        self.load_all_tabs()
        conn.close()
   
    def load_types(self):
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        types = c.execute("SELECT name FROM product_types ORDER BY name").fetchall()
        self.type_combobox['values'] = [t[0] for t in types]
        conn.close()
   
    def update_type_combobox_selection(self):
        if self.selected_type: # SỬA: Set cho single type
            self.type_combobox.set(self.selected_type)
   
    def on_type_selected(self, event=None):
        selected_type = self.type_combobox.get()
        if selected_type and selected_type != self.selected_type: # SỬA: Chỉ set nếu khác cũ
            if self.selected_type:
                # GIẢI THÍCH: Xác nhận trước khi đổi type, vì sẽ clear data
                if not messagebox.askyesno("Xác nhận", "Đổi loại sản phẩm sẽ xóa dữ liệu hiện tại. Tiếp tục?"):
                    self.type_combobox.set(self.selected_type) # Giữ nguyên cũ
                    return
               
                # GIẢI THÍCH: Clear data khi đổi type
                self.custom_indicators = {}
                self.deleted_indicators = {}
                self.origin_deleted = {}
                self.reference_products = []
                self.custom_rows_ctkt_ms = []
                self.reference_columns = {}
                # Clear tất cả trees
                for tab_name in ["three_brands", "bom", "dmkt", "ctkt_bo", "ctkt_mua_sam"]:
                    tree = getattr(self, f"{tab_name}_tree")
                    tree.delete(*tree.get_children())
                    indicator_map = getattr(self, f"{tab_name}_indicator_map")
                    indicator_map.clear()
           
            self.selected_type = selected_type
            self.init_default_crit_type() # Init crit_type cho type mới
            self.load_all_tabs() # Load tabs với type mới
            self.on_tab_changed(None)
   
    def add_new_reference_row(self):
        """
        GIẢI THÍCH: Thêm sản phẩm tham khảo mới (manufacturer)
        - Chỉ cho phép thêm ở tab Hãng
        - Tạo manufacturer mới với tên và product_name rỗng
        - Tạo mapping với product_types đã chọn
        - Tạo product_specifications rỗng cho tất cả indicators
        - Tạo reference_products nếu đã có product_id
        """
        if self.current_tab != "three_brands":
            messagebox.showerror("Lỗi", "Chỉ có thể thêm sản phẩm tham khảo trong tab Hãng")
            return
       
        if not self.selected_type: # SỬA: Kiểm tra self.selected_type thay vì list
            messagebox.showerror("Lỗi", "Chọn loại sản phẩm trước khi thêm sản phẩm tham khảo")
            return
       
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
       
        # SỬA: Lấy chỉ 1 type_id từ self.selected_type
        type_id_result = c.execute("SELECT id FROM product_types WHERE name=?", (self.selected_type,)).fetchone()
        if not type_id_result:
            conn.close()
            return
        type_id = type_id_result[0]
       
        try:
            # GIẢI THÍCH: Tạo manufacturer mới với tên và product_name rỗng
            c.execute("INSERT INTO manufacturers (name, product_name) VALUES (?, ?)", ("", ""))
            man_id = c.lastrowid
           
            # SỬA: Tạo mapping chỉ với 1 type_id
            c.execute("INSERT INTO product_type_mapping (manufacturer_id, type_id) VALUES (?, ?)", (man_id, type_id))
           
            # SỬA: Tạo product_specifications rỗng chỉ cho indicators của 1 type_id
            indicator_ids = c.execute("SELECT id FROM indicators WHERE type_id=?", (type_id,)).fetchall()
            for ind_id_tuple in indicator_ids:
                c.execute("INSERT INTO product_specifications (manufacturer_id, indicator_id, specification_value) VALUES (?, ?, ?)",
                            (man_id, ind_id_tuple[0], ""))
           
            # GIẢI THÍCH: Tạo reference_products nếu đã có product_id
            if self.product_id:
                sort_order = len(self.reference_products)
                c.execute("INSERT INTO reference_products (product_id, manufacturer_id, sort_order) VALUES (?, ?, ?)",
                        (self.product_id, man_id, sort_order))
           
            conn.commit()
           
            # GIẢI THÍCH: Thêm vào danh sách reference_products
            self.reference_products.append((man_id, "", ""))
           
            # GIẢI THÍCH: Reload tab Hãng và BOM
            self.load_three_brands_tab()
            self.load_bom_tab()
           
            messagebox.showinfo("Thành công", "Đã thêm sản phẩm tham khảo mới")
           
        except Exception as e:
            conn.rollback()
            messagebox.showerror("Lỗi", f"Lỗi khi thêm sản phẩm tham khảo: {str(e)}")
        finally:
            conn.close()

    def delete_reference_product(self):
        """
        GIẢI THÍCH: Xóa sản phẩm tham khảo
        - Chỉ cho phép xóa ở tab Hãng
        - Cho phép chọn nhiều sản phẩm để xóa cùng lúc
        - SỬA MỚI: Tối ưu tốc độ xóa bằng cách giảm reload không cần thiết
        """
        if self.current_tab != "three_brands":
            messagebox.showerror("Lỗi", "Chỉ có thể xóa sản phẩm tham khảo trong tab Hãng")
            return
       
        if not self.reference_products:
            messagebox.showinfo("Thông báo", "Không có sản phẩm tham khảo nào để xóa")
            return
        # Tạo cửa sổ chọn sản phẩm để xóa
        delete_win = tk.Toplevel(self.root)
        delete_win.title("Xóa sản phẩm tham khảo")
        delete_win.geometry("600x400")
        delete_win.resizable(False, False)
        delete_win.transient(self.root)
        delete_win.grab_set()
        main_frame = tk.Frame(delete_win)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        label = tk.Label(main_frame, text="Chọn sản phẩm tham khảo để xóa:",
                        font=("Arial", 12, "bold"))
        label.pack(pady=(0, 10))
        # Treeview
        tree_frame = tk.Frame(main_frame)
        tree_frame.pack(fill=tk.BOTH, expand=True)
        v_scrollbar = tk.Scrollbar(tree_frame, orient="vertical")
        v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        delete_tree = ttk.Treeview(tree_frame, columns=("STT", "Tên sản phẩm", "Tên hãng"),
                                show="headings", yscrollcommand=v_scrollbar.set, selectmode="extended")
        v_scrollbar.config(command=delete_tree.yview)
        delete_tree.heading("STT", text="STT")
        delete_tree.heading("Tên sản phẩm", text="Tên sản phẩm")
        delete_tree.heading("Tên hãng", text="Tên hãng")
        delete_tree.column("STT", width=50, anchor="center")
        delete_tree.column("Tên sản phẩm", width=250, anchor="w")
        delete_tree.column("Tên hãng", width=250, anchor="w")
        delete_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        # Thêm dữ liệu vào Treeview
        for index, (man_id, name, prod_name) in enumerate(self.reference_products, 1):
            delete_tree.insert("", "end",
                            values=(index,
                                    prod_name or "Chưa đặt tên sản phẩm",
                                    name or "Chưa đặt tên hãng"),
                            tags=(str(man_id),))
        # Frame chứa các nút
        button_frame = tk.Frame(main_frame)
        button_frame.pack(pady=10)
        def do_delete_selected():
            selected = delete_tree.selection()
            if not selected:
                messagebox.showerror("Lỗi", "Vui lòng chọn ít nhất một sản phẩm tham khảo để xóa")
                return
            # Hỏi xác nhận ngắn gọn
            if not messagebox.askyesno("Xác nhận", "Bạn có chắc muốn xóa các sản phẩm tham khảo đã chọn không?"):
                return
            conn = sqlite3.connect(DB_NAME)
            c = conn.cursor()
            deleted_count = 0
            try:
                # BƯỚC 1: Xóa các sản phẩm tham khảo đã chọn - BATCH DELETE
                # GIẢI THÍCH: Thu thập tất cả man_id cần xóa trước
                man_ids_to_delete = []
                for item in selected:
                    tags = delete_tree.item(item)['tags']
                    if tags:
                        man_id = int(tags[0])
                        man_ids_to_delete.append(man_id)
               
                if not man_ids_to_delete:
                    conn.close()
                    return
               
                # GIẢI THÍCH: Xóa hàng loạt thay vì từng cái một - NHANH HƠN
                placeholders = ','.join('?' * len(man_ids_to_delete))
                c.execute(f"DELETE FROM product_specifications WHERE manufacturer_id IN ({placeholders})", man_ids_to_delete)
                c.execute(f"DELETE FROM product_type_mapping WHERE manufacturer_id IN ({placeholders})", man_ids_to_delete)
                c.execute(f"DELETE FROM reference_products WHERE manufacturer_id IN ({placeholders})", man_ids_to_delete)
                c.execute(f"DELETE FROM manufacturers WHERE id IN ({placeholders})", man_ids_to_delete)
               
                deleted_count = len(man_ids_to_delete)
               
                # GIẢI THÍCH: Cập nhật danh sách reference_products
                self.reference_products = [(m_id, n, p) for m_id, n, p in self.reference_products
                                        if m_id not in man_ids_to_delete]
                conn.commit()
               
                # BƯỚC 2: Cập nhật giá trị cho các chỉ tiêu - CHỈ CẬP NHẬT CUSTOM_INDICATORS
                # GIẢI THÍCH: Không cần query lại database, chỉ cập nhật custom_indicators
                if self.selected_type:
                    type_id_result = c.execute("SELECT id FROM product_types WHERE name=?", (self.selected_type,)).fetchone()
                    if type_id_result:
                        type_id = type_id_result[0]
                       
                        # GIẢI THÍCH: Query một lần duy nhất để lấy tất cả indicators
                        hidden_ids = self.deleted_indicators.get("three_brands", set())
                        if hidden_ids:
                            placeholders = ','.join('?' * len(hidden_ids))
                            indicators = c.execute(
                                f"SELECT id FROM indicators WHERE type_id = ? AND id NOT IN ({placeholders})",
                                (type_id,) + tuple(hidden_ids)
                            ).fetchall()
                        else:
                            indicators = c.execute("SELECT id FROM indicators WHERE type_id = ?", (type_id,)).fetchall()
                       
                        # GIẢI THÍCH: Cập nhật custom_indicators mà không reload tab
                        for ind_tuple in indicators:
                            ind_id = ind_tuple[0]
                            danh_gia = self.custom_indicators.get("three_brands", {}).get(f"danh_gia_{ind_id}", "not")
                           
                            # GIẢI THÍCH: Tính giá trị mới nhưng chưa load tab
                            # Sẽ tính dựa trên reference_products đã được cập nhật
                            new_extreme_value = self.calculate_extreme_value(ind_id, self.three_brands_tree, self.reference_columns, danh_gia)
                            self.custom_indicators.setdefault("three_brands", {})[f"so_sanh_{ind_id}"] = new_extreme_value
                           
                            # GIẢI THÍCH: Cập nhật cho các tab phụ thuộc (chỉ cập nhật custom_indicators)
                            # BOM
                            self.custom_indicators.setdefault("bom", {})[f"so_sanh_{ind_id}"] = new_extreme_value
                           
                            # DMKT
                            self.custom_indicators.setdefault("dmkt", {})[f"so_sanh_{ind_id}"] = new_extreme_value
                           
                            # CTKT bộ
                            self.custom_indicators.setdefault("ctkt_bo", {})[f"gia_tri_{ind_id}"] = new_extreme_value
                           
                            # CTKT mua sắm - cập nhật cả tiêu chí
                            self.custom_indicators.setdefault("ctkt_mua_sam", {})[f"so_sanh_{ind_id}"] = new_extreme_value
               
                conn.close()
               
                # BƯỚC 3: Reload tất cả các tab MỘT LẦN DUY NHẤT
                # GIẢI THÍCH: Load lại tab để hiển thị thay đổi
                # Sử dụng update_idletasks() để giảm lag
                delete_win.destroy()
               
                self.root.update_idletasks()
                self.load_three_brands_tab()
               
                self.root.update_idletasks()
                self.load_bom_tab()
               
                self.root.update_idletasks()
                self.load_dmkt_tab()
               
                self.root.update_idletasks()
                self.load_ctkt_bo_tab()
               
                self.root.update_idletasks()
                self.load_ctkt_mua_sam_tab()
               
                # Thông báo ngắn gọn
                messagebox.showinfo("Thành công",
                    f"Đã xóa {deleted_count} sản phẩm tham khảo và cập nhật giá trị")
            except Exception as e:
                conn.rollback()
                messagebox.showerror("Lỗi", f"Lỗi khi xóa: {str(e)}")
            finally:
                conn.close()
        def cancel_delete():
            delete_win.destroy()
        # Các nút
        delete_button = ttk.Button(button_frame,
                                text=f"XÓA ({len(delete_tree.get_children())} sản phẩm)",
                                command=do_delete_selected,
                                width=20)
        delete_button.pack(side=tk.LEFT, padx=5)
        cancel_button = ttk.Button(button_frame,
                                text="HỦY",
                                command=cancel_delete,
                                width=20)
        cancel_button.pack(side=tk.LEFT, padx=5)
        # Phím tắt
        delete_win.bind('<Return>', lambda e: do_delete_selected())
        delete_win.bind('<Escape>', lambda e: cancel_delete())
        delete_tree.focus_set()
        # Cập nhật text của nút Xóa
        def update_delete_button_text():
            selected_count = len(delete_tree.selection())
            if selected_count > 0:
                delete_button.config(text=f"XÓA ({selected_count} sản phẩm)")
            else:
                delete_button.config(text=f"XÓA ({len(delete_tree.get_children())} sản phẩm)")
        delete_tree.bind('<<TreeviewSelect>>', lambda e: update_delete_button_text())

    def add_reference_from_tree(self, event):
        if self.current_tab != "three_brands" or not self.current_entry:
            return
        text = self.current_entry.get().strip()
        self.current_entry.destroy()
        self.current_entry = None
       
        if not text:
            return
       
        selected_item = self.three_brands_tree.selection()
        if not selected_item:
            return
        item = selected_item[0]
        ind_id = self.three_brands_indicator_map.get(item)
       
        if not ind_id or not ind_id.startswith("name_") and not ind_id.startswith("product_name_"):
            return
       
        man_id = int(ind_id.split("_")[-1])
        field = "name" if ind_id.startswith("name_") else "product_name"
       
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        try:
            c.execute(f"UPDATE manufacturers SET {field}=? WHERE id=?", (text, man_id))
            conn.commit()
            self.reference_products = [(m_id, text if field == "name" and m_id == man_id else n,
                                      text if field == "product_name" and m_id == man_id else p)
                                     for m_id, n, p in self.reference_products]
            self.load_three_brands_tab()
            self.load_bom_tab()
        except Exception as e:
            conn.rollback()
            messagebox.showerror("Lỗi", f"Lỗi khi cập nhật sản phẩm tham khảo: {str(e)}")
        finally:
            conn.close()
   
    def load_all_tabs(self, event=None):
        if not self.selected_type:
            return
        for tab_name in self.tab_data.keys():
            self.load_single_tab(tab_name)
        self.on_tab_changed(None)
   
    def load_single_tab(self, tab_name):
        if not self.selected_type:
            return
        if tab_name == "three_brands":
            self.load_three_brands_tab()
        elif tab_name == "bom":
            self.load_bom_tab()
        elif tab_name == "dmkt":
            self.load_dmkt_tab()
        elif tab_name == "ctkt_bo":
            self.load_ctkt_bo_tab()
        elif tab_name == "ctkt_mua_sam":
            self.load_ctkt_mua_sam_tab()
       
        if tab_name == "three_brands":
            self.load_single_tab("bom")
            self.load_single_tab("dmkt")
            self.load_single_tab("ctkt_bo")
            self.load_single_tab("ctkt_mua_sam")
        elif tab_name == "bom":
            self.load_single_tab("dmkt")
            self.load_single_tab("ctkt_bo")
        elif tab_name == "dmkt":
            self.load_single_tab("ctkt_bo")
   
    def load_three_brands_tab(self):
        """
        GIẢI THÍCH: Load dữ liệu cho tab Hãng
        - SỬA: Thêm cột "Đánh giá" và cột "Tham chiếu" sau mỗi sản phẩm tham khảo
        - Đồng bộ với logic từ detail_product.py
        """
        if not self.selected_type:
            return
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
      
        # Lấy type_id từ self.selected_type
        type_id_result = c.execute("SELECT id FROM product_types WHERE name=?", (self.selected_type,)).fetchone()
        if not type_id_result:
            conn.close()
            return
        type_id = type_id_result[0]
      
        hidden_ids = self.deleted_indicators.get("three_brands", set())
        placeholders = ','.join('?' * len(hidden_ids)) if hidden_ids else '0'
        indicators = c.execute(f"SELECT id, indicator_code, indicator, unit, value FROM indicators WHERE type_id = ? AND id NOT IN ({placeholders})",
                            (type_id,) + tuple(hidden_ids)).fetchall()
      
        self.three_brands_tree.delete(*self.three_brands_tree.get_children())
        self.three_brands_tree["columns"] = []
        self.three_brands_indicator_map.clear()
      
        # GIẢI THÍCH: Tạo display_columns với cột tham chiếu sau mỗi sản phẩm tham khảo
        display_columns = ["Mã chỉ tiêu", "Chỉ tiêu", "Đánh giá", "Giá trị", "Giá trị tham khảo", "Đơn vị", "Loại chỉ tiêu"]
        self.reference_columns = {}
        self.reference_value_columns = {} # Dict để lưu cột tham chiếu
      
        # GIẢI THÍCH: Tạo header cho cột sản phẩm tham khảo và cột tham chiếu
        for i, (man_id, _, _) in enumerate(self.reference_products, 1):
            manufacturer_info = c.execute("SELECT name, product_name FROM manufacturers WHERE id=?", (man_id,)).fetchone()
            if manufacturer_info:
                name, prod_name = manufacturer_info
                col_name = f"{prod_name or 'Sản phẩm'} - {name or 'Hãng'}" if name or prod_name else f"Sản phẩm tham khảo {i}"
            else:
                col_name = f"Sản phẩm tham khảo {i}"
          
            display_columns.append(col_name)
            self.reference_columns[col_name] = man_id
          
            # Thêm cột tham chiếu sau mỗi sản phẩm tham khảo
            ref_col_name = f"Tham chiếu {i}"
            display_columns.append(ref_col_name)
            self.reference_value_columns[ref_col_name] = man_id # Lưu man_id để biết thuộc sản phẩm nào
      
        display_columns.append("Hành động")
      
        self.three_brands_tree["columns"] = display_columns
        for col in display_columns:
            self.three_brands_tree.heading(col, text=col)
            # GIẢI THÍCH: Điều chỉnh width cho các cột
            if col.startswith("Tham chiếu"):
                width = 250 # SỬA: Tăng width cho cột tham chiếu để hiển thị hết chữ dài
            elif col in ["Mã chỉ tiêu", "Chỉ tiêu"]:
                width = 300 # SỬA: Tăng width cho cột dài để tránh cắt chữ, cho phép kéo ngang nếu vượt màn hình
            else:
                width = self.column_widths["three_brands"].get(col, 200)
          
            self.three_brands_tree.column(col, width=width, minwidth=width, stretch=False,
                                        anchor="center" if col in ["Mã chỉ tiêu", "Giá trị", "Giá trị tham khảo", "Đơn vị", "Loại chỉ tiêu", "Hành động"] or col in self.reference_columns or col in self.reference_value_columns else "w")
            # GIẢI THÍCH: Thêm minwidth=width và stretch=False để các cột giữ kích thước cố định, không bị nén khi thêm nhiều cột, buộc thanh cuộn ngang xuất hiện khi tổng width vượt quá khung treeview, giúp hiển thị hết nội dung mà không mất tên cột.
      
        # GIẢI THÍCH SỬA: Nếu không có sản phẩm tham khảo, đặt stretch=True cho các cột chính để giới hạn độ rộng, làm thanh cuộn ngang nhỏ lại (vừa khung)
        if not self.reference_products:
            for col in display_columns:
                if col not in ["Hành động"]:
                    self.three_brands_tree.column(col, stretch=True)
      
        max_height = self.max_row_heights.get("three_brands", 30)
      
        # Hàng 1: Tên sản phẩm tham khảo - SỬA: Cho phép chỉnh sửa cột tham chiếu ở hàng 1
        values = ["Tên sản phẩm tham khảo", "", "", "", "", "", ""]
        for col_name in self.reference_columns.keys():
            man_id = self.reference_columns[col_name]
            manufacturer_info = c.execute("SELECT product_name FROM manufacturers WHERE id=?", (man_id,)).fetchone()
            prod_name = manufacturer_info[0] if manufacturer_info else ""
            values.append(prod_name or "Nhập tên sản phẩm")
          
            # SỬA MỚI: Hàng 1 - Cho phép chỉnh sửa cột tham chiếu
            ref_key = f"ref_value_{man_id}_product_name" # Sử dụng key đặc biệt cho tên sản phẩm
            ref_value = self.custom_indicators.get("three_brands", {}).get(ref_key, "")
            values.append(wrap_text(ref_value, 20)) # SỬA: Tăng wrap_text lên 20 để hiển thị tốt hơn
        values.append("") # Không có nút Xóa
        item = self.three_brands_tree.insert("", "end", values=values, tags=("new_reference",))
        self.three_brands_indicator_map[item] = "product_name"
      
        # Hàng 2: Tên hãng - SỬA: Không cho chỉnh sửa cột tham chiếu ở hàng 2
        values = ["Tên hãng", "", "", "", "", "", ""]
        for col_name in self.reference_columns.keys():
            man_id = self.reference_columns[col_name]
            manufacturer_info = c.execute("SELECT name FROM manufacturers WHERE id=?", (man_id,)).fetchone()
            name = manufacturer_info[0] if manufacturer_info else ""
            values.append(name or "Nhập tên hãng")
            values.append("") # Cột tham chiếu để trống ở hàng 2
        values.append("") # Không có nút Xóa
        item = self.three_brands_tree.insert("", "end", values=values, tags=("new_reference",))
        self.three_brands_indicator_map[item] = "name"
      
        had_indicators = False
        # SỬA MỚI: Thu thập tất cả indicator_code để kiểm tra parent
        all_codes = [ind[1] for ind in indicators]
        def is_parent(code, all_codes):
            return any(c.startswith(code + '.') for c in all_codes if c != code)
      
        # Các hàng thông số kỹ thuật - SỬA: Không cho chỉnh sửa cột tham chiếu ở các hàng này
        for index, ind in enumerate(indicators, 1):
            had_indicators = True
            ind_id, ind_code, indc, unit, value = ind
            values = [wrap_text(ind_code, 50), wrap_text(indc, 50)] # SỬA: Tăng wrap_text lên 50 để hiển thị hết chữ dài
          
            # SỬA MỚI: Kiểm tra nếu là parent thì để trống "Đánh giá", "Giá trị", "Loại chỉ tiêu"
            parent_tag = ()
            if is_parent(ind_code, all_codes):
                danh_gia = ""
                so_sanh = ""
                crit_type = ""
                unit = ""  # Để trống Đơn vị theo yêu cầu không edit, nhưng hiển thị rỗng
                parent_tag = ("parent",)
            else:
                # Cột "Đánh giá"
                danh_gia = self.custom_indicators.get("three_brands", {}).get(f"danh_gia_{ind_id}", "not")
          
                # Cột "Giá trị"
                so_sanh = self.custom_indicators.get("three_brands", {}).get(f"so_sanh_{ind_id}", "")
                if not so_sanh:
                    so_sanh = self.calculate_extreme_value(ind_id, self.three_brands_tree, self.reference_columns, danh_gia)
                    if so_sanh:
                        self.custom_indicators.setdefault("three_brands", {})[f"so_sanh_{ind_id}"] = so_sanh
          
                # Cột "Loại chỉ tiêu"
                crit_type = self.custom_indicators.get("three_brands", {}).get(f"crit_type_{ind_id}", "CTCB")
                if crit_type not in ["CTCB", "KCB"]:
                    crit_type = "CTCB"
                    self.custom_indicators["three_brands"][f"crit_type_{ind_id}"] = crit_type
          
            values.append(danh_gia)
            values.append(wrap_text(so_sanh, 20))
          
            # Cột "Giá trị tham khảo"
            gia_tri_tham_khao = value if value is not None else ""
            values.append(wrap_text(gia_tri_tham_khao, 20))
          
            values.append(wrap_text(unit or "", 20))
            values.append(crit_type)
          
            # GIẢI THÍCH: Thêm giá trị cho cột sản phẩm tham khảo và cột tham chiếu
            for col_name in self.reference_columns.keys():
                man_id = self.reference_columns[col_name]
                val = c.execute("SELECT specification_value FROM product_specifications WHERE manufacturer_id=? AND indicator_id=?",
                                (man_id, ind_id)).fetchone()
                specific_value = wrap_text(val[0] if val else "", 50)
                values.append(specific_value)
              
                # SỬA MỚI: Hàng thông số kỹ thuật - Không cho chỉnh sửa cột tham chiếu
                ref_key = f"ref_value_{man_id}_{ind_id}"
                ref_value = self.custom_indicators.get("three_brands", {}).get(ref_key, "")
                values.append(wrap_text(ref_value, 20))
          
            values.append("Xóa")
          
            line_count = max([v.count('\n') + 1 for v in values if isinstance(v, str)])
            item = self.three_brands_tree.insert("", "end", values=values, tags=("data_row",) + parent_tag)
            self.three_brands_indicator_map[item] = ind_id
          
            self.check_and_mark_row(self.three_brands_tree, item, ind_id, so_sanh, None, "three_brands")
          
            height = line_count * 30
            max_height = max(max_height, height)
      
        if had_indicators:
            max_height = min(max_height, 120)
            self.max_row_heights["three_brands"] = max_height
            style = ttk.Style()
            style.configure("Custom.Treeview", rowheight=max_height)
            style.configure("DataRow.Treeview", rowheight=max_height)
      
        conn.close()

    def load_bom_tab(self):
        """
        GIẢI THÍCH: Load dữ liệu cho tab BOM - SAO CHÉP 100% logic từ load_three_brands_tab
        - Không được sửa đổi bất kỳ logic nào so với load_three_brands_tab
        - Chỉ thay đổi tên biến three_brands → bom
        """
        if not self.selected_type:
            return
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
      
        # Lấy type_id từ self.selected_type
        type_id_result = c.execute("SELECT id FROM product_types WHERE name=?", (self.selected_type,)).fetchone()
        if not type_id_result:
            conn.close()
            return
        type_id = type_id_result[0]
      
        hidden_ids = self.deleted_indicators.get("bom", set())
        placeholders = ','.join('?' * len(hidden_ids)) if hidden_ids else '0'
        indicators = c.execute(f"SELECT id, indicator_code, indicator, unit FROM indicators WHERE type_id = ? AND id NOT IN ({placeholders})",
                            (type_id,) + tuple(hidden_ids)).fetchall()
      
        # Xóa toàn bộ
        self.bom_tree.delete(*self.bom_tree.get_children())
        self.bom_tree["columns"] = []
        self.bom_indicator_map.clear()
      
        # SAO CHÉP 100% từ load_three_brands_tab dòng 108-109
        self.reference_columns = {}
        self.reference_value_columns = {} # Dict để lưu cột tham chiếu
      
        # SAO CHÉP 100% từ load_three_brands_tab dòng 111-125
        # GIẢI THÍCH: Tạo display_columns với cột tham chiếu sau mỗi sản phẩm tham khảo
        display_columns = ["Mã chỉ tiêu", "Chỉ tiêu", "Giá trị", "Đơn vị", "Loại chỉ tiêu"]
      
        # GIẢI THÍCH: Tạo header cho cột sản phẩm tham khảo và cột tham chiếu
        for i, (man_id, _, _) in enumerate(self.reference_products, 1):
            manufacturer_info = c.execute("SELECT name, product_name FROM manufacturers WHERE id=?", (man_id,)).fetchone()
            if manufacturer_info:
                name, prod_name = manufacturer_info
                col_name = f"{prod_name or 'Sản phẩm'} - {name or 'Hãng'}" if name or prod_name else f"Sản phẩm tham khảo {i}"
            else:
                col_name = f"Sản phẩm tham khảo {i}"
          
            display_columns.append(col_name)
            self.reference_columns[col_name] = man_id
          
            # Thêm cột tham chiếu sau mỗi sản phẩm tham khảo
            ref_col_name = f"Tham chiếu {i}"
            display_columns.append(ref_col_name)
            self.reference_value_columns[ref_col_name] = man_id # Lưu man_id để biết thuộc sản phẩm nào
      
        display_columns.append("Hành động")
      
        # SAO CHÉP 100% từ load_three_brands_tab dòng 127-138
        self.bom_tree["columns"] = display_columns
        for col in display_columns:
            self.bom_tree.heading(col, text=col)
            # GIẢI THÍCH: Điều chỉnh width cho các cột
            if col.startswith("Tham chiếu"):
                width = 250 # SỬA: Tăng width cho cột tham chiếu để hiển thị hết chữ dài
            elif col in ["Mã chỉ tiêu", "Chỉ tiêu"]:
                width = 300 # SỬA: Tăng width cho cột dài để tránh cắt chữ, cho phép kéo ngang nếu vượt màn hình
            else:
                width = self.column_widths["bom"].get(col, 200)
          
            self.bom_tree.column(col, width=width, minwidth=width, stretch=False,
                                anchor="center" if col in ["Mã chỉ tiêu", "Giá trị", "Đơn vị", "Loại chỉ tiêu", "Hành động"] or col in self.reference_columns or col in self.reference_value_columns else "w")
            # GIẢI THÍCH: Thêm minwidth=width và stretch=False để các cột giữ kích thước cố định, không bị nén khi thêm nhiều cột, buộc thanh cuộn ngang xuất hiện khi tổng width vượt quá khung treeview, giúp hiển thị hết nội dung mà không mất tên cột.
      
        # GIẢI THÍCH SỬA: Nếu không có sản phẩm tham khảo, đặt stretch=True cho các cột chính để giới hạn độ rộng, làm thanh cuộn ngang nhỏ lại (vừa khung)
        if not self.reference_products:
            for col in display_columns:
                if col not in ["Hành động"]:
                    self.bom_tree.column(col, stretch=True)
      
        max_height = self.max_row_heights.get("bom", 80)
      
        # Hàng 1: Tên sản phẩm tham khảo
        values = ["Tên sản phẩm tham khảo", "", "", "", ""]
        for col_name in self.reference_columns.keys():
            man_id = self.reference_columns[col_name]
            manufacturer_info = c.execute("SELECT product_name FROM manufacturers WHERE id=?", (man_id,)).fetchone()
            prod_name = manufacturer_info[0] if manufacturer_info else ""
            values.append(prod_name or "Nhập tên sản phẩm")
          
            # Hàng 1 - Cho phép chỉnh sửa cột tham chiếu
            ref_key = f"ref_value_{man_id}_product_name"
            ref_value = self.custom_indicators.get("three_brands", {}).get(ref_key, "") # SỬA: Lấy ref_value từ "three_brands" vì BOM không có custom riêng
            values.append(wrap_text(ref_value, 20))
        values.append("") # Không có nút Xóa
        item = self.bom_tree.insert("", "end", values=values, tags=("new_reference",))
        self.bom_indicator_map[item] = "product_name_bom" # SỬA: Thay "product_name" thành "product_name_bom" để khớp filter_ids trong export
      
        # Hàng 2: Tên hãng
        values = ["Tên hãng", "", "", "", ""]
        for col_name in self.reference_columns.keys():
            man_id = self.reference_columns[col_name]
            manufacturer_info = c.execute("SELECT name FROM manufacturers WHERE id=?", (man_id,)).fetchone()
            name = manufacturer_info[0] if manufacturer_info else ""
            values.append(name or "Nhập tên hãng")
            values.append("") # Cột tham chiếu để trống ở hàng 2
        values.append("") # Không có nút Xóa
        item = self.bom_tree.insert("", "end", values=values, tags=("new_reference",))
        self.bom_indicator_map[item] = "name_bom" # SỬA: Thay "name" thành "name_bom" để khớp filter_ids trong export
      
        had_indicators = False
        # SỬA MỚI: Thu thập tất cả indicator_code để kiểm tra parent
        all_codes = [ind[1] for ind in indicators]
        def is_parent(code, all_codes):
            return any(c.startswith(code + '.') for c in all_codes if c != code)
      
        # Các hàng thông số kỹ thuật
        for index, ind in enumerate(indicators, 1):
            had_indicators = True
            ind_id, ind_code, indc, unit = ind
            values = [wrap_text(ind_code, 50), wrap_text(indc, 50)]
          
            # SỬA MỚI: Kiểm tra nếu là parent thì để trống "Giá trị", "Loại chỉ tiêu"
            parent_tag = ()
            if is_parent(ind_code, all_codes):
                so_sanh = ""
                crit_type = ""
                unit = ""  # Để trống Đơn vị theo yêu cầu không edit, nhưng hiển thị rỗng
                parent_tag = ("parent",)
            else:
                # Cột "Giá trị"
                so_sanh = self.custom_indicators.get("bom", {}).get(f"so_sanh_{ind_id}", "")
                if not so_sanh:
                    so_sanh = self.custom_indicators.get("three_brands", {}).get(f"so_sanh_{ind_id}", "")
                    if not so_sanh:
                        danh_gia = self.custom_indicators.get("three_brands", {}).get(f"danh_gia_{ind_id}", "not")
                        so_sanh = self.calculate_extreme_value(ind_id, self.bom_tree, self.reference_columns, danh_gia)
                    if so_sanh:
                        self.custom_indicators.setdefault("bom", {})[f"so_sanh_{ind_id}"] = so_sanh
          
                # Cột "Loại chỉ tiêu"
                crit_type = self.custom_indicators.get("bom", {}).get(f"crit_type_{ind_id}", "CTCB")
                if crit_type not in ["CTCB", "KCB"]:
                    crit_type = "CTCB"
                    self.custom_indicators["bom"][f"crit_type_{ind_id}"] = crit_type
          
            values.append(wrap_text(so_sanh, 20))
          
            values.append(wrap_text(unit or "", 20))
            values.append(crit_type)
          
            # GIẢI THÍCH: Thêm giá trị cho cột sản phẩm tham khảo và cột tham chiếu
            for col_name in self.reference_columns.keys():
                man_id = self.reference_columns[col_name]
                val = c.execute("SELECT specification_value FROM product_specifications WHERE manufacturer_id=? AND indicator_id=?",
                                (man_id, ind_id)).fetchone()
                specific_value = wrap_text(val[0] if val else "", 50)
                values.append(specific_value)
              
                # Hàng thông số kỹ thuật - Không cho chỉnh sửa cột tham chiếu
                ref_key = f"ref_value_{man_id}_{ind_id}"
                ref_value = self.custom_indicators.get("three_brands", {}).get(ref_key, "") # SỬA: Lấy ref_value từ "three_brands" vì BOM không có custom riêng
                values.append(wrap_text(ref_value, 20))
          
            values.append("Xóa")
          
            line_count = max([v.count('\n') + 1 for v in values if isinstance(v, str)])
            item = self.bom_tree.insert("", "end", values=values, tags=("data_row",) + parent_tag)
            self.bom_indicator_map[item] = ind_id
          
            self.check_and_mark_row(self.bom_tree, item, ind_id, so_sanh, None, "bom")
          
            height = line_count * 30
            max_height = max(max_height, height)
      
        if had_indicators:
            max_height = min(max_height, 120)
            self.max_row_heights["bom"] = max_height
            style = ttk.Style()
            style.configure("Custom.Treeview", rowheight=max_height)
            style.configure("DataRow.Treeview", rowheight=max_height)
      
        conn.close()

    def load_dmkt_tab(self):
        """
        GIẢI THÍCH: Load tab DMKT - ĐỒNG BỘ với project_manager.py
        - Hiển thị header với "- "
        - Hiển thị item với "+ "
        - Cascade từ bom
        - SỬA: Khi yêu cầu kỹ thuật không có chỉ tiêu (tức là indicator rỗng hoặc không có sub-indicator), thì không cần thêm hàng "+ hiển thị chỉ tiêu" nữa, mà hiển thị các thông số giá trị, đơn vị, hành động ngay tại hàng yêu cầu kỹ thuật (header). Nếu có chỉ tiêu (indicator không rỗng), thì giữ nguyên logic cũ: header rỗng, sub rows với "+ indicator".
        - SỬA: Sửa STT: header là 1,2,3,...; sub là 1.1,1.2,...
        """
        if not self.selected_type:
            return
      
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
      
        # SỬA: Lấy type_id từ self.selected_type
        type_id_result = c.execute("SELECT id FROM product_types WHERE name=?", (self.selected_type,)).fetchone()
        if not type_id_result:
            conn.close()
            return
        type_id = type_id_result[0]
      
        hidden_ids = self.deleted_indicators.get("dmkt", set())
        placeholders = ','.join('?' * len(hidden_ids)) if hidden_ids else '0'
      
        indicators = c.execute(f"""
            SELECT id, indicator_code, indicator, unit
            FROM indicators
            WHERE type_id = ? AND id NOT IN ({placeholders})
        """, (type_id,) + tuple(hidden_ids)).fetchall()
      
        self.dmkt_tree.delete(*self.dmkt_tree.get_children())
        self.dmkt_tree["columns"] = ("Mã chỉ tiêu", "Yêu cầu kỹ thuật", "Giá trị", "Đơn vị", "Hành động")
        for col in self.dmkt_tree["columns"]:
            self.dmkt_tree.heading(col, text=col)
            self.dmkt_tree.column(col, width=self.column_widths["dmkt"].get(col, 200),
                                anchor="center" if col in ["Mã chỉ tiêu", "Giá trị", "Đơn vị", "Hành động"] else "w")
      
        self.dmkt_indicator_map.clear()
        max_height = self.max_row_heights["dmkt"]
      
        # SỬA MỚI: Thu thập tất cả indicator_code để kiểm tra parent
        all_codes = [ind[1] for ind in indicators]
        def is_parent(code, all_codes):
            return any(c.startswith(code + '.') for c in all_codes if c != code)
      
        # SỬA: Bỏ phân cấp, loop trực tiếp qua từng indicator, mỗi hàng độc lập
        for ind in indicators:
            ind_id, ind_code, indc, unit = ind
          
            # SỬA MỚI: Kiểm tra nếu là parent thì để trống "Giá trị", "Đơn vị"
            parent_tag = ()
            if is_parent(ind_code, all_codes):
                so_sanh = ""
                unit = ""  # Để trống Đơn vị theo yêu cầu không edit, nhưng hiển thị rỗng
                parent_tag = ("parent",)
            else:
                # GIẢI THÍCH: Cascade từ bom
                so_sanh = self.custom_indicators.get("dmkt", {}).get(f"so_sanh_{ind_id}", "")
                if not so_sanh:
                    bom_value = self.custom_indicators.get("bom", {}).get(f"so_sanh_{ind_id}", "")
                    if not bom_value:
                        hang_value = self.custom_indicators.get("three_brands", {}).get(f"so_sanh_{ind_id}", "")
                        if not hang_value:
                            danh_gia = self.custom_indicators.get("three_brands", {}).get(f"danh_gia_{ind_id}", "not") # GIẢI THÍCH: Thêm lấy danh_gia để tính extreme cho dmkt, giống add_product.py.
                            hang_value = self.calculate_extreme_value(ind_id, self.dmkt_tree, self.reference_columns, danh_gia) # GIẢI THÍCH: Thay calculate_min_value bằng calculate_extreme_value để tính giá trị extreme dựa trên danh_gia, giống add_product.py.
                        bom_value = hang_value
                    so_sanh = bom_value
          
            # SỬA: Display với mã chỉ tiêu = ind_code, yêu cầu kỹ thuật = indc (indicator)
            values = [wrap_text(ind_code, 50), wrap_text(indc, 50), wrap_text(so_sanh or "", 20), wrap_text(unit or "", 20), "Xóa"]
            line_count = max([v.count('\n') + 1 for v in values if isinstance(v, str)])
            item = self.dmkt_tree.insert("", "end", values=values, tags=("data_row",) + parent_tag)
            self.dmkt_indicator_map[item] = ind_id
          
            # GIẢI THÍCH: Tham chiếu từ bom
            bom_reference = self.custom_indicators.get("bom", {}).get(f"so_sanh_{ind_id}", "")
            if not bom_reference:
                hang_reference = self.custom_indicators.get("three_brands", {}).get(f"so_sanh_{ind_id}", "")
                if not hang_reference:
                    danh_gia = self.custom_indicators.get("three_brands", {}).get(f"danh_gia_{ind_id}", "not") # GIẢI THÍCH: Thêm lấy danh_gia để tính extreme cho dmkt tham chiếu, giống add_product.py.
                    hang_reference = self.calculate_extreme_value(ind_id, self.dmkt_tree, self.reference_columns, danh_gia) # GIẢI THÍCH: Thay calculate_min_value bằng calculate_extreme_value để tính giá trị extreme dựa trên danh_gia, giống add_product.py.
                bom_reference = hang_reference
            self.check_and_mark_row(self.dmkt_tree, item, ind_id, so_sanh, bom_reference, "dmkt")
          
            height = line_count * 30
            max_height = max(max_height, height)
      
        max_height = min(max_height, 120)
        self.max_row_heights["dmkt"] = max_height
        style = ttk.Style()
        style.configure("Custom.Treeview", rowheight=max_height)
        style.configure("DataRow.Treeview", rowheight=max_height)
        conn.close()

    def load_ctkt_bo_tab(self):
        if not self.selected_type:
            return
      
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
      
        type_id_result = c.execute("SELECT id FROM product_types WHERE name=?", (self.selected_type,)).fetchone()
        if not type_id_result:
            conn.close()
            return
        type_id = type_id_result[0]
      
        hidden_ids = self.deleted_indicators.get("ctkt_bo", set())
        placeholders = ','.join('?' * len(hidden_ids)) if hidden_ids else '0'
      
        indicators = c.execute(f"""
            SELECT id, indicator_code, indicator, unit
            FROM indicators
            WHERE type_id = ? AND id NOT IN ({placeholders})
        """, (type_id,) + tuple(hidden_ids)).fetchall()
      
        self.ctkt_bo_tree.delete(*self.ctkt_bo_tree.get_children())
        self.ctkt_bo_tree["columns"] = ("Mã chỉ tiêu", "Chỉ tiêu", "Đơn vị", "Giá trị", "Hành động")
        for col in self.ctkt_bo_tree["columns"]:
            self.ctkt_bo_tree.heading(col, text=col)
            self.ctkt_bo_tree.column(col, width=self.column_widths["ctkt_bo"].get(col, 300),
                                    anchor="center" if col in ["Mã chỉ tiêu", "Đơn vị", "Giá trị", "Hành động"] else "w")
      
        self.ctkt_bo_indicator_map.clear()
        max_height = self.max_row_heights["ctkt_bo"]
      
        # SỬA MỚI: Thu thập tất cả indicator_code để kiểm tra parent
        all_codes = [ind[1] for ind in indicators]
        def is_parent(code, all_codes):
            return any(c.startswith(code + '.') for c in all_codes if c != code)
      
        # SỬA: Bỏ phân cấp, loop trực tiếp qua từng indicator, mỗi hàng độc lập
        for ind in indicators:
            ind_id, ind_code, indc, unit = ind
              
            # SỬA MỚI: Kiểm tra nếu là parent thì để trống "Giá trị", "Đơn vị"
            parent_tag = ()
            if is_parent(ind_code, all_codes):
                gia_tri = ""
                unit = ""  # Để trống Đơn vị theo yêu cầu không edit, nhưng hiển thị rỗng
                parent_tag = ("parent",)
            else:
                gia_tri = self.custom_indicators.get("ctkt_bo", {}).get(f"gia_tri_{ind_id}", "")
                if not gia_tri:
                    dmkt_value = self.custom_indicators.get("dmkt", {}).get(f"so_sanh_{ind_id}", "")
                    if not dmkt_value:
                        ctkt_ms_value = self.custom_indicators.get("ctkt_mua_sam", {}).get(f"so_sanh_{ind_id}", "")
                        if not ctkt_ms_value:
                            bom_value = self.custom_indicators.get("bom", {}).get(f"so_sanh_{ind_id}", "")
                            if not bom_value:
                                hang_value = self.custom_indicators.get("three_brands", {}).get(f"so_sanh_{ind_id}", "")
                                if not hang_value:
                                    danh_gia = self.custom_indicators.get("three_brands", {}).get(f"danh_gia_{ind_id}", "not")
                                    hang_value = self.calculate_extreme_value(ind_id, self.ctkt_bo_tree, self.reference_columns, danh_gia)
                                bom_value = hang_value
                            ctkt_ms_value = bom_value
                        dmkt_value = ctkt_ms_value
                    gia_tri = dmkt_value
              
            # SỬA: Luôn thêm mô tả dựa trên danh_gia, kể cả khi gia_tri rỗng hoặc không phải số nguyên
            danh_gia = self.custom_indicators.get("three_brands", {}).get(f"danh_gia_{ind_id}", "not")
            display_indc = indc
            if danh_gia == ">=":
                display_indc += ", không nhỏ hơn"
            elif danh_gia == "<=":
                display_indc += ", không lớn hơn"
            elif danh_gia == "=":
                display_indc += ", bằng"
            elif danh_gia == ">":
                display_indc += ", lớn hơn"
            elif danh_gia == "<":
                display_indc += ", nhỏ hơn"
            # Không thêm gì nếu "not"
              
            values = [wrap_text(ind_code, 50), wrap_text(display_indc, 50), wrap_text(unit or "-", 20), wrap_text(gia_tri, 50), "Xóa"]
            line_count = max([v.count('\n') + 1 for v in values if isinstance(v, str)])
            item = self.ctkt_bo_tree.insert("", "end", values=values, tags=("data_row",) + parent_tag)
            self.ctkt_bo_indicator_map[item] = ind_id
              
            dmkt_reference = self.custom_indicators.get("dmkt", {}).get(f"so_sanh_{ind_id}", "")
            if not dmkt_reference:
                ctkt_ms_reference = self.custom_indicators.get("ctkt_mua_sam", {}).get(f"so_sanh_{ind_id}", "")
                if not ctkt_ms_reference:
                    bom_reference = self.custom_indicators.get("bom", {}).get(f"so_sanh_{ind_id}", "")
                    if not bom_reference:
                        hang_reference = self.custom_indicators.get("three_brands", {}).get(f"so_sanh_{ind_id}", "")
                        if not hang_reference:
                            danh_gia = self.custom_indicators.get("three_brands", {}).get(f"danh_gia_{ind_id}", "not")
                            hang_reference = self.calculate_extreme_value(ind_id, self.ctkt_bo_tree, self.reference_columns, danh_gia)
                        bom_reference = hang_reference
                    ctkt_ms_reference = bom_reference
                dmkt_reference = ctkt_ms_reference
            self.check_and_mark_row(self.ctkt_bo_tree, item, ind_id, gia_tri, dmkt_reference, "ctkt_bo")
              
            height = line_count * 30
            max_height = max(max_height, height)
      
        max_height = min(max_height, 120)
        self.max_row_heights["ctkt_bo"] = max_height
        style = ttk.Style()
        style.configure("Custom.Treeview", rowheight=max_height)
        style.configure("DataRow.Treeview", rowheight=max_height)
        conn.close()

    def load_ctkt_mua_sam_tab(self):
        if not self.selected_type:
            return
     
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
     
        type_id = c.execute("SELECT id FROM product_types WHERE name=?", (self.selected_type,)).fetchone()
        if type_id:
            type_id = type_id[0]
     
        hidden_ids = self.deleted_indicators.get("ctkt_mua_sam", set())
        placeholders = ','.join('?' * len(hidden_ids)) if hidden_ids else '0'
     
        indicators = c.execute(f"""
            SELECT id, indicator_code, indicator, unit
            FROM indicators
            WHERE type_id = ? AND id NOT IN ({placeholders})
        """, (type_id,) + tuple(hidden_ids)).fetchall()
     
        self.ctkt_mua_sam_tree.delete(*self.ctkt_mua_sam_tree.get_children())
        self.ctkt_mua_sam_tree["columns"] = ("Mã chỉ tiêu", "Chỉ tiêu kỹ thuật chi tiết",
                                            "Giá trị", "Đơn vị", "Tiêu chí đánh giá", "Loại chỉ tiêu", "Hành động")
     
        for col in self.ctkt_mua_sam_tree["columns"]:
            self.ctkt_mua_sam_tree.heading(col, text=col)
            if col == "Mã chỉ tiêu":
                width = 250
            elif col == "Chỉ tiêu kỹ thuật chi tiết":
                width = 200
            elif col == "Tiêu chí đánh giá":
                width = 450
            else:
                width = self.column_widths["ctkt_mua_sam"].get(col, 300)
         
            self.ctkt_mua_sam_tree.column(col, width=width,
                                        anchor="center" if col in ["Mã chỉ tiêu", "Giá trị", "Đơn vị", "Loại chỉ tiêu", "Hành động"] else "w")
     
        self.ctkt_mua_sam_indicator_map.clear()
        max_height = 120
        req_number = 1
     
        groups = defaultdict(list)
        for ind in indicators:
            groups[ind[1]].append(ind)
     
        # SỬA MỚI: Thu thập tất cả indicator_code để kiểm tra parent (req là indicator_code)
        all_codes = list(groups.keys())
        def is_parent(code, all_codes):
            return any(c.startswith(code + '.') for c in all_codes if c != code)
     
        for req, ind_list in groups.items():
            if not ind_list:
                continue
         
            sub_stt = 1
         
            first_ind = ind_list[0]
            ind_id = first_ind[0]
            unit = first_ind[3] or ""
         
            # SỬA MỚI: Di chuyển gán yeu_cau_str ra ngoài if is_parent để tránh lỗi UnboundLocalError
            yeu_cau_str = first_ind[2] or ""
         
            # SỬA MỚI: Kiểm tra nếu là parent thì để trống "Giá trị", "Đơn vị", "Tiêu chí đánh giá", "Loại chỉ tiêu"
            parent_tag = ()
            if is_parent(req, all_codes):
                so_sanh = ""
                unit = ""
                crit_type = ""
                tieu_chi_display = ""
                tieu_chi_raw = ""
                parent_tag = ("parent",)
            else:
                so_sanh = self.custom_indicators.get("ctkt_mua_sam", {}).get(f"so_sanh_{ind_id}", "")
                if not so_sanh:
                    hang_value = self.custom_indicators.get("three_brands", {}).get(f"so_sanh_{ind_id}", "")
                    if not hang_value:
                        danh_gia = self.custom_indicators.get("three_brands", {}).get(f"danh_gia_{ind_id}", "not")
                        hang_value = self.calculate_extreme_value(ind_id, self.three_brands_tree, self.reference_columns, danh_gia)
                    so_sanh = hang_value
                    self.custom_indicators.setdefault("ctkt_mua_sam", {})[f"so_sanh_{ind_id}"] = so_sanh
         
                # GIẢI THÍCH SỬA: Nếu yeu_cau_str (indicator) rỗng (tức yêu cầu kỹ thuật không có chỉ tiêu), thì dùng req (requirement) thay thế để xây dựng tiêu chí đánh giá.
                if not yeu_cau_str:
                    yeu_cau_str = req # Sử dụng Yêu cầu kỹ thuật (req) nếu không có chỉ tiêu
         
                is_numeric = re.match(r'^-?\d+(\.\d+)?$', str(so_sanh).strip())
                unit_str = f" {unit}" if unit else ""
         
                danh_gia = self.custom_indicators.get("three_brands", {}).get(f"danh_gia_{ind_id}", "not")
         
                if is_numeric and yeu_cau_str:
                    if danh_gia == "<=":
                        dat_line = f"- Đạt: {yeu_cau_str} ≤ {so_sanh}{unit_str}"
                        khong_dat_line = f"- Không đạt: {yeu_cau_str} > {so_sanh}{unit_str}"
                    elif danh_gia == ">=":
                        dat_line = f"- Đạt: {yeu_cau_str} ≥ {so_sanh}{unit_str}"
                        khong_dat_line = f"- Không đạt: {yeu_cau_str} < {so_sanh}{unit_str}"
                    elif danh_gia == "=":
                        dat_line = f"- Đạt: {yeu_cau_str} = {so_sanh}{unit_str}"
                        khong_dat_line = f"- Không đạt: {yeu_cau_str} ≠ {so_sanh}{unit_str}"
                    elif danh_gia == "<":
                        dat_line = f"- Đạt: {yeu_cau_str} < {so_sanh}{unit_str}"
                        khong_dat_line = f"- Không đạt: {yeu_cau_str} ≥ {so_sanh}{unit_str}"
                    elif danh_gia == ">":
                        dat_line = f"- Đạt: {yeu_cau_str} > {so_sanh}{unit_str}"
                        khong_dat_line = f"- Không đạt: {yeu_cau_str} ≤ {so_sanh}{unit_str}"
                    else:
                        dat_line = f"- Đạt: hàng hóa nhà thầu chào đáp ứng yêu cầu kỹ thuật"
                        khong_dat_line = f"- Không đạt: hàng hóa nhà thầu chào không đáp ứng yêu cầu kỹ thuật"
                else:
                    dat_line = f"- Đạt: hàng hóa nhà thầu chào đáp ứng yêu cầu kỹ thuật"
                    khong_dat_line = f"- Không đạt: hàng hóa nhà thầu chào không đáp ứng yêu cầu kỹ thuật"
         
                tieu_chi_display = f"{dat_line}\r\n{khong_dat_line}"
         
                crit_type = self.custom_indicators.get("ctkt_mua_sam", {}).get(f"crit_type_{ind_id}", "CTCB")
                if crit_type not in ["CTCB", "KCB"]:
                    crit_type = "CTCB"
                    self.custom_indicators["ctkt_mua_sam"][f"crit_type_{ind_id}"] = crit_type
         
                tieu_chi_raw = f"{dat_line}\n{khong_dat_line}"
                self.custom_indicators["ctkt_mua_sam"][f"tieu_chi_{ind_id}"] = tieu_chi_raw
         
            values = [
                    wrap_text(req, 40),
                    wrap_text(yeu_cau_str, 30),
                    wrap_text(so_sanh, 20),
                    wrap_text(unit, 15),
                    tieu_chi_display,
                    crit_type,
                    "Xóa"]
         
            item = self.ctkt_mua_sam_tree.insert("", "end", values=values, tags=("data_row", "group_header") + parent_tag)
            self.ctkt_mua_sam_indicator_map[item] = ind_id
         
            hang_reference = self.custom_indicators.get("three_brands", {}).get(f"so_sanh_{ind_id}", "")
            if not hang_reference:
                danh_gia = self.custom_indicators.get("three_brands", {}).get(f"danh_gia_{ind_id}", "not")
                hang_reference = self.calculate_extreme_value(ind_id, self.three_brands_tree, self.reference_columns, danh_gia)
            self.check_and_mark_row(self.ctkt_mua_sam_tree, item, ind_id, so_sanh, tieu_chi_raw, "ctkt_mua_sam")
         
            height = 120
            max_height = max(max_height, height)
            sub_stt += 1
         
            for sub_ind in ind_list[1:]:
                ind_id = sub_ind[0]
                unit = sub_ind[3] or ""
             
                # SỬA MỚI: Di chuyển gán yeu_cau_str ra ngoài if is_parent để tránh lỗi UnboundLocalError
                yeu_cau_str = sub_ind[2] or ""
             
                parent_tag = ()
                if is_parent(sub_ind[1], all_codes):
                    so_sanh = ""
                    unit = ""
                    crit_type = ""
                    tieu_chi_display = ""
                    tieu_chi_raw = ""
                    parent_tag = ("parent",)
                else:
                    so_sanh = self.custom_indicators.get("ctkt_mua_sam", {}).get(f"so_sanh_{ind_id}", "")
                    if not so_sanh:
                        hang_value = self.custom_indicators.get("three_brands", {}).get(f"so_sanh_{ind_id}", "")
                        if not hang_value:
                            danh_gia = self.custom_indicators.get("three_brands", {}).get(f"danh_gia_{ind_id}", "not")
                            hang_value = self.calculate_extreme_value(ind_id, self.ctkt_mua_sam_tree, self.reference_columns, danh_gia)
                        so_sanh = hang_value
                        self.custom_indicators.setdefault("ctkt_mua_sam", {})[f"so_sanh_{ind_id}"] = so_sanh
             
                    # GIẢI THÍCH SỬA: Tương tự, nếu yeu_cau_str rỗng (dù ít xảy ra ở sub_ind), dùng req thay thế.
                    if not yeu_cau_str:
                        yeu_cau_str = req
             
                    is_numeric = re.match(r'^-?\d+(\.\d+)?$', str(so_sanh).strip())
                    unit_str = f" {unit}" if unit else ""
             
                    danh_gia = self.custom_indicators.get("three_brands", {}).get(f"danh_gia_{ind_id}", "not")
             
                    if is_numeric and yeu_cau_str:
                        if danh_gia == "<=":
                            dat_line = f"- Đạt: {yeu_cau_str} ≤ {so_sanh}{unit_str}"
                            khong_dat_line = f"- Không đạt: {yeu_cau_str} > {so_sanh}{unit_str}"
                        elif danh_gia == ">=":
                            dat_line = f"- Đạt: {yeu_cau_str} ≥ {so_sanh}{unit_str}"
                            khong_dat_line = f"- Không đạt: {yeu_cau_str} < {so_sanh}{unit_str}"
                        elif danh_gia == "=":
                            dat_line = f"- Đạt: {yeu_cau_str} = {so_sanh}{unit_str}"
                            khong_dat_line = f"- Không đạt: {yeu_cau_str} ≠ {so_sanh}{unit_str}"
                        elif danh_gia == "<":
                            dat_line = f"- Đạt: {yeu_cau_str} < {so_sanh}{unit_str}"
                            khong_dat_line = f"- Không đạt: {yeu_cau_str} ≥ {so_sanh}{unit_str}"
                        elif danh_gia == ">":
                            dat_line = f"- Đạt: {yeu_cau_str} > {so_sanh}{unit_str}"
                            khong_dat_line = f"- Không đạt: {yeu_cau_str} ≤ {so_sanh}{unit_str}"
                        else:
                            dat_line = f"- Đạt: hàng hóa nhà thầu chào đáp ứng yêu cầu kỹ thuật"
                            khong_dat_line = f"- Không đạt: hàng hóa nhà thầu chào không đáp ứng yêu cầu kỹ thuật"
                    else:
                        dat_line = f"- Đạt: hàng hóa nhà thầu chào đáp ứng yêu cầu kỹ thuật"
                        khong_dat_line = f"- Không đạt: hàng hóa nhà thầu chào không đáp ứng yêu cầu kỹ thuật"
             
                    tieu_chi_display = f"{dat_line}\r\n{khong_dat_line}"
             
                    crit_type = self.custom_indicators.get("ctkt_mua_sam", {}).get(f"crit_type_{ind_id}", "CTCB")
                    if crit_type not in ["CTCB", "KCB"]:
                        crit_type = "CTCB"
                        self.custom_indicators["ctkt_mua_sam"][f"crit_type_{ind_id}"] = crit_type
             
                    tieu_chi_raw = f"{dat_line}\n{khong_dat_line}"
                    self.custom_indicators["ctkt_mua_sam"][f"tieu_chi_{ind_id}"] = tieu_chi_raw
             
                values = [
                        "",
                        wrap_text(yeu_cau_str, 30),
                        wrap_text(so_sanh, 20),
                        wrap_text(unit, 15),
                        tieu_chi_display,
                        crit_type,
                        "Xóa"]
             
                item = self.ctkt_mua_sam_tree.insert("", "end", values=values, tags=("data_row",) + parent_tag)
                self.ctkt_mua_sam_indicator_map[item] = ind_id
             
                hang_reference = self.custom_indicators.get("three_brands", {}).get(f"so_sanh_{ind_id}", "")
                if not hang_reference:
                    danh_gia = self.custom_indicators.get("three_brands", {}).get(f"danh_gia_{ind_id}", "not")
                    hang_reference = self.calculate_extreme_value(ind_id, self.ctkt_mua_sam_tree, self.reference_columns, danh_gia)
                self.check_and_mark_row(self.ctkt_mua_sam_tree, item, ind_id, so_sanh, tieu_chi_raw, "ctkt_mua_sam")
             
                height = 120
                max_height = max(max_height, height)
                sub_stt += 1
         
            req_number += 1
     
        # ===== PHẦN SỬA MỚI: Xử lý "Yêu cầu khác" với Mã chỉ tiêu tự động =====
        yeu_cau_khac_item = None
        for child in self.ctkt_mua_sam_tree.get_children():
            if "yeu_cau_khac" in self.ctkt_mua_sam_tree.item(child).get('tags', ()):
                yeu_cau_khac_item = child
                break
     
        if not yeu_cau_khac_item:
            # BƯỚC 1: Thu thập tất cả các Mã chỉ tiêu hiện có trong tree
            all_codes = []
            for child in self.ctkt_mua_sam_tree.get_children():
                # Lấy giá trị cột đầu tiên (Mã chỉ tiêu)
                code_value = self.ctkt_mua_sam_tree.item(child)['values'][0]
                # Chỉ xử lý nếu mã không rỗng
                if code_value and str(code_value).strip():
                    all_codes.append(str(code_value).strip())
           
            # BƯỚC 2: Tìm số nguyên lớn nhất từ chỉ số đầu tiên của các mã phân cấp
            max_first_index = 0
           
            for code in all_codes:
                # Tách lấy phần đầu tiên trước dấu chấm đầu tiên (nếu có)
                # Ví dụ: "2.1.1" -> "2", "3" -> "3", "2.1" -> "2"
                first_part = code.split('.')[0] if '.' in code else code
               
                # Kiểm tra nếu phần đầu là số nguyên hợp lệ
                if first_part.isdigit():
                    current_index = int(first_part)
                    # Lưu lại số lớn nhất
                    max_first_index = max(max_first_index, current_index)
           
            # BƯỚC 3: Tính Mã chỉ tiêu mới = số lớn nhất + 1
            # Ví dụ:
            # - Nếu có mã "1", "2", "2.1", "2.1.1" -> max_first_index = 2 -> next_code = "3"
            # - Nếu có mã "1", "3.2", "3.2.1" -> max_first_index = 3 -> next_code = "4"
            next_code = str(max_first_index + 1)
           
            # BƯỚC 4: Tạo hàng "Yêu cầu khác" với Mã chỉ tiêu mới
            values = [next_code, "Yêu cầu khác", "", "", "", "", ""]
            yeu_cau_khac_item = self.ctkt_mua_sam_tree.insert("", "end", values=values, tags=("special_header", "yeu_cau_khac"))
            self.ctkt_mua_sam_indicator_map[yeu_cau_khac_item] = "yeu_cau_khac"
            height = 30
            max_height = max(max_height, height)
        # ===== KẾT THÚC PHẦN SỬA MỚI =====
     
        # Xử lý custom items dưới "Yêu cầu khác" với STT dạng {yeu_cau_khac_stt}.{sub_stt}
        custom_ids = set()
        for key in self.custom_indicators.get("ctkt_mua_sam", {}).keys():
            if key.startswith("chi_tieu_") and key.split("_")[-1].startswith("-"):
                custom_ids.add(key.split("_")[-1])
     
        sorted_custom_ids = sorted(custom_ids, key=lambda x: int(x) if x.lstrip('-').isdigit() else 0, reverse=True)
     
        for custom_id_str in sorted_custom_ids:
            tieu_chi_raw = self.custom_indicators["ctkt_mua_sam"].get(f"tieu_chi_{custom_id_str}", "")
            crit_type = self.custom_indicators["ctkt_mua_sam"].get(f"crit_type_{custom_id_str}", "CTCB")
            chi_tieu = self.custom_indicators["ctkt_mua_sam"].get(f"chi_tieu_{custom_id_str}", "")
            yeu_cau = self.custom_indicators["ctkt_mua_sam"].get(f"yeu_cau_{custom_id_str}", "")
         
            if crit_type not in ["CTCB", "KCB"]:
                crit_type = "CTCB"
                self.custom_indicators["ctkt_mua_sam"][f"crit_type_{custom_id_str}"] = crit_type
         
            tieu_chi_display = self.format_tieu_chi_for_display(tieu_chi_raw)
         
            values = [
                    wrap_text(chi_tieu, 40),
                    wrap_text(yeu_cau, 30),
                    "",
                    "",
                    tieu_chi_display,
                    crit_type,
                    "Xóa"]
         
            item = self.ctkt_mua_sam_tree.insert("", "end", values=values, tags=("data_row", "custom_single"))
            self.ctkt_mua_sam_indicator_map[item] = custom_id_str
            self.custom_rows_ctkt_ms.append(item)
         
            height = 120
            max_height = max(max_height, height)
     
        max_height = min(max_height, 120)
        self.max_row_heights["ctkt_mua_sam"] = max_height
        style = ttk.Style()
        style.configure("Custom.Treeview", rowheight=max_height)
        style.configure("DataRow.Treeview", rowheight=max_height)
        conn.close()

    def format_tieu_chi_for_display(self, tieu_chi_raw):
        if not tieu_chi_raw:
            return ""
       
        tieu_chi_display = tieu_chi_raw.replace('\n', '\r\n')
        return tieu_chi_display
   
    def on_double_click(self, event):
        """
        GIẢI THÍCH: Xử lý sự kiện double click để chỉnh sửa các ô trong tree
        - SỬA MỚI: Chỉ cho phép chỉnh sửa cột tham chiếu ở HÀNG ĐẦU TIÊN (tên sản phẩm) trong tab Hãng
        - Tab BOM không cho chỉnh sửa cột tham chiếu
        - GIẢI THÍCH SỬA: Khi thay đổi "Đánh giá" (danh_gia) ở tab three_brands, hàm save_danh_gia sẽ cập nhật custom_indicators, tính lại so_sanh, và gọi update_cascade_marking để cascade giá trị đến các tab phụ thuộc bao gồm ctkt_bo. Trong load_ctkt_bo_tab (gọi qua load_single_tab trong update_cascade_marking), nếu yêu cầu kỹ thuật không có chỉ tiêu (is_single_no_indicator), sẽ tự động thêm mô tả như ", không nhỏ hơn" hoặc ", không lớn hơn" vào display_indc (header) dựa trên danh_gia mới từ custom_indicators["three_brands"], sửa lỗi hiển thị tab ctkt bộ không cập nhật mô tả khi thay đổi cột đánh giá ở tab Hãng (tự động thay đổi theo luôn nhờ reload tab với dữ liệu danh_gia mới).
        """
        if not self.current_tree or not self.current_tab:
            return

        item = self.current_tree.identify_row(event.y)
        if not item:
            return

        column = self.current_tree.identify_column(event.x)
        col_idx = int(column.replace("#", "")) - 1
        columns = self.current_tree["columns"]
        if col_idx < 0 or col_idx >= len(columns):
            return

        col_name = columns[col_idx]

        ind_id = self.current_indicator_map.get(item)
        if not ind_id:
            return

        tags = self.current_tree.item(item)['tags']

        # Vô hiệu hóa double click cho hàng "Yêu cầu khác"
        if self.current_tab == "ctkt_mua_sam" and "yeu_cau_khac" in tags:
            return

        # SỬA MỚI: Nếu là hàng parent, không cho edit bất kỳ cột nào trừ "Hành động"
        if "parent" in tags and col_name != "Hành động":
            messagebox.showinfo("Thông báo", f"Không thể chỉnh sửa cột '{col_name}' ở hàng mục cha (chỉ xóa được qua Hành động)")
            return

        # XỬ LÝ ĐẶC BIỆT CHO 2 HÀNG ĐẦU TIÊN (tên sản phẩm, tên hãng)
        if self.current_tab == "three_brands" and (ind_id == "name" or ind_id == "product_name"):
            # Chỉ cho phép chỉnh sửa cột sản phẩm tham khảo và cột tham chiếu (chỉ ở hàng 1)
            if col_name not in self.reference_columns and col_name not in self.reference_value_columns:
                return
            
            man_id = None
            if col_name in self.reference_columns:
                man_id = self.reference_columns[col_name]
            elif col_name in self.reference_value_columns:
                man_id = self.reference_value_columns[col_name]
            
            if not man_id:
                return
            
            # SỬA MỚI: Chỉ cho phép chỉnh sửa cột tham chiếu ở HÀNG 1 (tên sản phẩm)
            if col_name in self.reference_value_columns and ind_id != "product_name":
                messagebox.showinfo("Thông báo", "Chỉ có thể chỉnh sửa cột tham chiếu ở hàng 'Tên sản phẩm tham khảo'")
                return
            
            if self.current_entry:
                self.current_entry.destroy()
            
            bbox = self.current_tree.bbox(item, column)
            if not bbox:
                return
            
            entry = tk.Entry(self.current_tree, width=20)
            entry.place(x=bbox[0], y=bbox[1], width=bbox[2], height=bbox[3])
            entry.insert(0, self.current_tree.item(item, "values")[col_idx])
            entry.focus_set()
            self.current_entry = entry
            
            def save_name_entry(event):
                new_value = entry.get().strip()
                entry.destroy()
                self.current_entry = None
                
                values = list(self.current_tree.item(item, "values"))
                values[col_idx] = new_value
                self.current_tree.item(item, values=tuple(values))
                
                # Xử lý lưu dữ liệu
                if col_name in self.reference_columns:
                    # Lưu tên sản phẩm hoặc tên hãng
                    field = "product_name" if ind_id == "product_name" else "name"
                    conn = sqlite3.connect(DB_NAME)
                    c = conn.cursor()
                    try:
                        c.execute(f"UPDATE manufacturers SET {field}=? WHERE id=?", (new_value, man_id))
                        conn.commit()
                        
                        self.reference_products = [(m_id, new_value if field == "name" and m_id == man_id else n,
                                                new_value if field == "product_name" and m_id == man_id else p)
                                                for m_id, n, p in self.reference_products]
                        
                        self.load_three_brands_tab()
                        self.load_bom_tab()
                        
                    except Exception as e:
                        conn.rollback()
                        messagebox.showerror("Lỗi", f"Lỗi khi cập nhật: {str(e)}")
                    finally:
                        conn.close()
                
                elif col_name in self.reference_value_columns:
                    # SỬA MỚI: Lưu giá trị tham chiếu cho hàng tên sản phẩm
                    ref_key = f"ref_value_{man_id}_product_name"
                    self.custom_indicators.setdefault("three_brands", {})[ref_key] = new_value
                    
                    # Reload tab BOM để cập nhật giá trị tham chiếu
                    self.load_bom_tab()
            
            entry.bind("<Return>", save_name_entry)
            entry.bind("<FocusOut>", save_name_entry)
            return

        # KIỂM TRA CỘT THAM CHIẾU
        is_reference_value_column = col_name.startswith("Tham chiếu")

        # SỬA MỚI: Xử lý cột tham chiếu - Chỉ cho phép ở hàng 1 tab Hãng
        if is_reference_value_column:
            if self.current_tab == "three_brands":
                # GIẢI THÍCH: Chỉ cho phép chỉnh sửa cột tham chiếu ở HÀNG 1 (tên sản phẩm)
                if ind_id != "product_name":
                    messagebox.showinfo("Thông báo", "Chỉ có thể chỉnh sửa cột tham chiếu ở hàng 'Tên sản phẩm tham khảo'")
                    return
                
                # GIẢI THÍCH: Lấy man_id từ reference_value_columns
                if not hasattr(self, 'reference_value_columns'):
                    return
                
                man_id = self.reference_value_columns.get(col_name)
                if not man_id:
                    return
                
                # GIẢI THÍCH: Tạo Entry để chỉnh sửa
                if self.current_entry:
                    self.current_entry.destroy()
                
                bbox = self.current_tree.bbox(item, column)
                if not bbox:
                    return
                
                entry = tk.Entry(self.current_tree, width=20)
                entry.place(x=bbox[0], y=bbox[1], width=bbox[2], height=bbox[3])
                entry.insert(0, self.current_tree.item(item, "values")[col_idx])
                entry.focus_set()
                self.current_entry = entry
                
                def save_ref_value_entry(event):
                    """GIẢI THÍCH: Lưu giá trị tham chiếu"""
                    new_value = entry.get().strip()
                    entry.destroy()
                    self.current_entry = None
                    
                    # Cập nhật giá trị trong tree
                    values = list(self.current_tree.item(item, "values"))
                    values[col_idx] = new_value
                    self.current_tree.item(item, values=tuple(values))
                    
                    # Lưu vào custom_indicators với key đặc biệt cho tên sản phẩm
                    ref_key = f"ref_value_{man_id}_product_name"
                    self.custom_indicators.setdefault("three_brands", {})[ref_key] = new_value
                    
                    # Reload tab BOM để cập nhật giá trị tham chiếu
                    self.load_bom_tab()
                
                entry.bind("<Return>", save_ref_value_entry)
                entry.bind("<FocusOut>", save_ref_value_entry)
                return
            
            elif self.current_tab == "bom":
                # GIẢI THÍCH: Tab BOM không cho chỉnh sửa cột tham chiếu
                messagebox.showinfo("Thông báo", "Cột tham chiếu chỉ có thể chỉnh sửa ở tab Hãng (hàng 'Tên sản phẩm tham khảo')")
                return

        # KIỂM TRA CHO TAB ctkt_mua_sam TỪ HÀNG "Yêu cầu khác" TRỞ ĐI
        is_after_yeu_cau_khac = False
        if self.current_tab == "ctkt_mua_sam":
            children = self.current_tree.get_children()
            yeu_cau_khac_index = -1
            for i, child in enumerate(children):
                child_tags = self.current_tree.item(child).get('tags', ())
                if "yeu_cau_khac" in child_tags:
                    yeu_cau_khac_index = i
                    break
            if yeu_cau_khac_index != -1:
                item_index = list(children).index(item)
                if item_index > yeu_cau_khac_index:
                    is_after_yeu_cau_khac = True

        # Định nghĩa các cột có thể chỉnh sửa cho từng tab
        editable_columns = {
            "three_brands": ["Đánh giá", "Giá trị", "Loại chỉ tiêu"] + list(self.reference_columns.keys()),
            "bom": ["Giá trị", "Loại chỉ tiêu"],
            "dmkt": ["Giá trị"],
            "ctkt_bo": ["Giá trị"],
            "ctkt_mua_sam": ["Giá trị", "Đơn vị", "Tiêu chí đánh giá", "Loại chỉ tiêu"]
        }

        # Cho phép chỉnh sửa "Mã chỉ tiêu" và "Chỉ tiêu kỹ thuật chi tiết" chỉ cho các hàng sau "Yêu cầu khác"
        if self.current_tab == "ctkt_mua_sam" and is_after_yeu_cau_khac:
            editable_columns["ctkt_mua_sam"] = ["Mã chỉ tiêu", "Chỉ tiêu kỹ thuật chi tiết", "Giá trị", "Đơn vị", "Tiêu chí đánh giá", "Loại chỉ tiêu"]

        # Kiểm tra cột có được phép chỉnh sửa không
        if col_name not in editable_columns.get(self.current_tab, []):
            return

        # Kiểm tra cột sản phẩm tham khảo chỉ được chỉnh sửa ở tab Hãng
        if col_name in self.reference_columns and self.current_tab != "three_brands":
            messagebox.showinfo("Thông báo", "Cột sản phẩm tham khảo chỉ có thể chỉnh sửa ở tab Hãng")
            return

        # Nếu là cột "Tiêu chí đánh giá" trong tab ctkt_mua_sam, mở giao diện chỉnh sửa
        if self.current_tab == "ctkt_mua_sam" and col_name == "Tiêu chí đánh giá":
            current_tieu_chi = self.current_tree.item(item, "values")[col_idx]
            self.open_tieu_chi_editor(self.current_tree, item, column, ind_id, current_tieu_chi)
            return

        # Xóa entry hoặc combobox cũ nếu đang có
        if self.current_entry:
            self.current_entry.destroy()

        bbox = self.current_tree.bbox(item, column)
        if not bbox:
            return

        # Tạo widget chỉnh sửa tùy thuộc vào cột
        if col_name == "Loại chỉ tiêu":
            # Sử dụng Combobox cho cột Loại chỉ tiêu
            combobox = ttk.Combobox(self.current_tree, values=["CTCB", "KCB"], state="readonly", font=("Arial", 12))
            combobox.place(x=bbox[0], y=bbox[1], width=bbox[2], height=bbox[3])
            
            current_value = self.current_tree.item(item, "values")[col_idx]
            combobox.set(current_value if current_value in ["CTCB", "KCB"] else "CTCB")
            combobox.focus_set()
            self.current_entry = combobox
            
            def save_combobox(event):
                new_value = combobox.get()
                combobox.destroy()
                self.current_entry = None
                
                values = list(self.current_tree.item(item, "values"))
                values[col_idx] = new_value
                self.current_tree.item(item, values=tuple(values))
                
                key = f"crit_type_{ind_id}"
                self.custom_indicators.setdefault(self.current_tab, {})[key] = new_value
            
            combobox.bind("<<ComboboxSelected>>", save_combobox)
            combobox.bind("<Return>", save_combobox)
            combobox.bind("<FocusOut>", save_combobox)

        elif col_name == "Đánh giá" and self.current_tab == "three_brands":
            # Sử dụng Combobox cho cột Đánh giá
            combobox = ttk.Combobox(self.current_tree, values=["not", ">", "<", "=", "<=", ">="], state="readonly", font=("Arial", 12))
            combobox.place(x=bbox[0], y=bbox[1], width=bbox[2], height=bbox[3])
            
            current_value = self.current_tree.item(item, "values")[col_idx]
            combobox.set(current_value if current_value in ["not", ">", "<", "=", "<=", ">="] else "not")
            combobox.focus_set()
            self.current_entry = combobox
            
            def save_danh_gia(event):
                new_danh_gia = combobox.get()
                combobox.destroy()
                self.current_entry = None
                values = list(self.current_tree.item(item, "values"))
                values[col_idx] = new_danh_gia
                self.current_tree.item(item, values=tuple(values))
                
                # Lưu vào custom_indicators
                self.custom_indicators.setdefault(self.current_tab, {})[f"danh_gia_{ind_id}"] = new_danh_gia
                
                # Tính lại Giá trị dựa trên đánh giá mới
                new_so_sanh = self.calculate_extreme_value(ind_id, self.current_tree, self.reference_columns, new_danh_gia)
                self.custom_indicators[self.current_tab][f"so_sanh_{ind_id}"] = new_so_sanh
                
                # Cập nhật cột Giá trị trong tree
                gia_tri_idx = columns.index("Giá trị")
                values[gia_tri_idx] = wrap_text(new_so_sanh, 20)
                self.current_tree.item(item, values=tuple(values))
                
                # Cập nhật cascade cho các tab phụ thuộc
                self.update_cascade_marking(str(ind_id), self.current_tab)
                
                # SỬA THÊM: Reload explicit tab ctkt_bo để đảm bảo cập nhật mô tả (dù update_cascade_marking đã load)
                self.load_single_tab("ctkt_bo")
                
                # Thêm lời gọi check_and_mark_row để bỏ bôi xanh nếu giá trị không rỗng
                self.check_and_mark_row(self.current_tree, item, ind_id, new_so_sanh, None, self.current_tab)
            
            combobox.bind("<<ComboboxSelected>>", save_danh_gia)
            combobox.bind("<Return>", save_danh_gia)
            combobox.bind("<FocusOut>", save_danh_gia)

        else:
            # Tạo Entry cho các cột khác
            entry = tk.Entry(self.current_tree, width=20)
            entry.place(x=bbox[0], y=bbox[1], width=bbox[2], height=bbox[3])
            entry.insert(0, self.current_tree.item(item, "values")[col_idx])
            entry.focus_set()
            self.current_entry = entry
            
            def save_entry(event):
                """Lưu giá trị đã chỉnh sửa"""
                new_value = entry.get().strip()
                entry.destroy()
                self.current_entry = None
                
                # Thêm ràng buộc cột Giá trị theo yêu cầu mới
                if col_name == "Giá trị" and not str(ind_id).startswith('-'):
                    try:
                        new_num = float(new_value) if new_value else float('inf')
                        is_numeric = True
                    except ValueError:
                        is_numeric = False
                    
                    if is_numeric:
                        danh_gia = self.custom_indicators.get("three_brands", {}).get(f"danh_gia_{ind_id}", "not")
                        if danh_gia != "not":
                            # Tính min và max từ các sản phẩm tham khảo
                            min_val, max_val = self.get_min_max_from_references(ind_id)
                            if min_val is not None and max_val is not None:
                                if danh_gia == "<=" and new_num > max_val:
                                    messagebox.showerror("Lỗi", f"Giá trị phải <= max của sản phẩm ({max_val})")
                                    new_value = str(max_val)
                                elif danh_gia == ">=" and new_num < min_val:
                                    messagebox.showerror("Lỗi", f"Giá trị phải >= min của sản phẩm ({min_val})")
                                    new_value = str(min_val)
                                elif danh_gia == ">" and new_num < min_val:
                                    messagebox.showerror("Lỗi", f"Giá trị phải > min của sản phẩm ({min_val - 1 if min_val.is_integer() else min_val})")
                                    new_value = str(min_val)
                                elif danh_gia == "<" and new_num > max_val:
                                    messagebox.showerror("Lỗi", f"Giá trị phải < max của sản phẩm ({max_val + 1 if max_val.is_integer() else max_val})")
                                    new_value = str(max_val)
                                elif danh_gia == "=" and new_num > max_val:
                                    messagebox.showerror("Lỗi", f"Giá trị phải <= max của sản phẩm ({max_val})")
                                    new_value = str(max_val)
                        
                        # Ràng buộc với tab phụ thuộc (prev_value)
                        prev_value = self.get_prev_tab_value(ind_id, self.current_tab)
                        if prev_value:
                            try:
                                prev_num = float(prev_value)
                                if danh_gia == "=":
                                    if new_num > prev_num:
                                        messagebox.showerror("Lỗi", f"Giá trị phải <= {prev_num} (theo tab trước)")
                                        new_value = str(prev_num)
                                else:
                                    if danh_gia == "<=" and new_num > prev_num:
                                        messagebox.showerror("Lỗi", f"Giá trị phải <= {prev_num}")
                                        new_value = prev_value
                                    elif danh_gia == ">=" and new_num < prev_num:
                                        messagebox.showerror("Lỗi", f"Giá trị phải >= {prev_num}")
                                        new_value = prev_value
                                    elif danh_gia == "<" and new_num >= prev_num:
                                        messagebox.showerror("Lỗi", f"Giá trị phải < {prev_num}")
                                        new_value = str(prev_num - 1) if prev_num.is_integer() else prev_value
                                    elif danh_gia == ">" and new_num <= prev_num:
                                        messagebox.showerror("Lỗi", f"Giá trị phải > {prev_num}")
                                        new_value = str(prev_num + 1) if prev_num.is_integer() else prev_value
                            except ValueError:
                                pass
                
                values = list(self.current_tree.item(item, "values"))
                values[col_idx] = new_value
                self.current_tree.item(item, values=tuple(values))
                
                # Xử lý chỉnh sửa cột sản phẩm tham khảo ở tab Hãng
                if self.current_tab == "three_brands" and col_name in self.reference_columns:
                    man_id = self.reference_columns[col_name]
                    conn = sqlite3.connect(DB_NAME)
                    c = conn.cursor()
                    try:
                        # Lưu giá trị mới vào database cho manufacturer cụ thể
                        c.execute("INSERT OR REPLACE INTO product_specifications (manufacturer_id, indicator_id, specification_value) VALUES (?, ?, ?)",
                                (man_id, ind_id, new_value))
                        conn.commit()
                        
                        # Tính lại extreme cho cột "Giá trị" (so_sanh)
                        danh_gia = self.custom_indicators.get("three_brands", {}).get(f"danh_gia_{ind_id}", "not")
                        extreme_value = self.calculate_extreme_value(ind_id, self.current_tree, self.reference_columns, danh_gia)
                        
                        # Cập nhật cột "Giá trị" thành extreme mới
                        gia_tri_col_idx = columns.index("Giá trị")
                        values[gia_tri_col_idx] = wrap_text(extreme_value, 20)
                        
                        self.current_tree.item(item, values=tuple(values))
                        
                        # Lưu giá trị extreme vào custom_indicators
                        self.custom_indicators.setdefault("three_brands", {})[f"so_sanh_{ind_id}"] = extreme_value
                        
                        # Cập nhật cascade cho các tab phụ thuộc
                        self.update_cascade_marking(str(ind_id), "three_brands")
                        
                        # Reload các tab liên quan để đồng bộ
                        self.load_single_tab("bom")
                        self.load_single_tab("dmkt")
                        self.load_single_tab("ctkt_bo")
                        self.load_single_tab("ctkt_mua_sam")
                        
                    except Exception as e:
                        conn.rollback()
                        messagebox.showerror("Lỗi", f"Lỗi khi cập nhật: {str(e)}")
                    finally:
                        conn.close()
                
                # Xử lý chỉnh sửa tab CTKT mua sắm
                elif self.current_tab == "ctkt_mua_sam":
                    # Xử lý chỉnh sửa chỉ tiêu tùy chỉnh
                    if isinstance(ind_id, str) and ind_id.startswith("-"):
                        key_map = {
                            "Mã chỉ tiêu": f"chi_tieu_{ind_id}",
                            "Chỉ tiêu kỹ thuật chi tiết": f"yeu_cau_{ind_id}",
                            "Giá trị": f"so_sanh_{ind_id}",
                            "Đơn vị": f"don_vi_{ind_id}",
                            "Loại chỉ tiêu": f"crit_type_{ind_id}"
                        }
                        key = key_map.get(col_name)
                        if key:
                            self.custom_indicators.setdefault(self.current_tab, {})[key] = new_value
                    else:
                        # Xử lý chỉ tiêu thông thường
                        key_map = {
                            "Mã chỉ tiêu": f"chi_tieu_{ind_id}",
                            "Chỉ tiêu kỹ thuật chi tiết": f"yeu_cau_{ind_id}",
                            "Giá trị": f"so_sanh_{ind_id}",
                            "Đơn vị": f"don_vi_{ind_id}",
                            "Loại chỉ tiêu": f"crit_type_{ind_id}"
                        }
                        key = key_map.get(col_name)
                        if key:
                            self.custom_indicators.setdefault(self.current_tab, {})[key] = new_value
                            
                            # Tự động tạo lại tiêu chí đánh giá khi thay đổi giá trị hoặc yêu cầu
                            if col_name in ["Giá trị", "Chỉ tiêu kỹ thuật chi tiết"]:
                                self.auto_update_tieu_chi(item, ind_id)
                
                # Xử lý chỉnh sửa các tab khác (three_brands, bom, dmkt, ctkt_bo)
                else:
                    key_map = {
                        "Giá trị": f"so_sanh_{ind_id}",
                        "Loại chỉ tiêu": f"crit_type_{ind_id}"
                    }
                    key = key_map.get(col_name)
                    if key:
                        self.custom_indicators.setdefault(self.current_tab, {})[key] = new_value
                        
                        # Nếu thay đổi giá trị, cập nhật cascade cho các tab phụ thuộc
                        if col_name == "Giá trị":
                            self.update_cascade_marking(str(ind_id), self.current_tab)
                
                # TỰ ĐỘNG CẬP NHẬT CÁC TAB PHỤ THUỘC KHI CỘT "Giá trị" THAY ĐỔI
                if col_name == "Giá trị":
                    # Lưu giá trị mới vào custom_indicators của tab hiện tại
                    value_key = "so_sanh_{ind_id}" if self.current_tab != "ctkt_bo" else "gia_tri_{ind_id}"
                    self.custom_indicators.setdefault(self.current_tab, {})[value_key] = new_value
                    
                    # Định nghĩa cascade map cho tab hiện tại
                    cascade_maps = {
                        "three_brands": ["bom", "dmkt", "ctkt_bo", "ctkt_mua_sam"],
                        "bom": ["dmkt", "ctkt_bo"],
                        "dmkt": ["ctkt_bo"],
                        "ctkt_bo": [],
                        "ctkt_mua_sam": []
                    }
                    dependent_tabs = cascade_maps.get(self.current_tab, [])
                    
                    # Cập nhật cascade và reload các tab phụ thuộc
                    self.update_cascade_marking(str(ind_id), self.current_tab)
                    for dep_tab in dependent_tabs:
                        self.load_single_tab(dep_tab)
                
                # Kiểm tra và đánh dấu row nếu cần
                self.check_and_mark_row(self.current_tree, item, ind_id,
                                    self.custom_indicators.get(self.current_tab, {}).get(f"so_sanh_{ind_id}", ""),
                                    self.custom_indicators.get(self.current_tab, {}).get(f"tieu_chi_{ind_id}", "") if self.current_tab == "ctkt_mua_sam" else None,
                                    self.current_tab)
            
            entry.bind("<Return>", save_entry)
            entry.bind("<FocusOut>", save_entry)

    def is_numeric_value(self, value):
        """
        GIẢI THÍCH: Kiểm tra xem giá trị có phải là số không
        """
        try:
            float(value.strip())
            return True
        except (ValueError, AttributeError):
            return False
        
    def auto_update_tieu_chi(self, item, ind_id):
        """
        GIẢI THÍCH: Tự động cập nhật tiêu chí đánh giá khi thay đổi giá trị hoặc yêu cầu kỹ thuật
        Đồng bộ với logic trong open_tieu_chi_editor
        """
        try:
            # GIẢI THÍCH: Lấy các cột cần thiết
            columns = self.current_tree["columns"]
            yeu_cau_col_idx = columns.index("Chỉ tiêu kỹ thuật chi tiết")
            gia_tri_col_idx = columns.index("Giá trị")
            unit_col_idx = columns.index("Đơn vị")
            tieu_chi_col_idx = columns.index("Tiêu chí đánh giá")
           
            # GIẢI THÍCH: Lấy giá trị từ tree
            values = self.current_tree.item(item, "values")
            yeu_cau = values[yeu_cau_col_idx].replace('\n', ' ')
            gia_tri = values[gia_tri_col_idx].replace('\n', ' ')
            unit = values[unit_col_idx].replace('\n', ' ')
           
            # Lấy danh_gia từ tab Hãng
            danh_gia = self.custom_indicators.get("three_brands", {}).get(f"danh_gia_{ind_id}", "not")
           
            # GIẢI THÍCH: Tạo tiêu chí đánh giá mới theo format chuẩn
            is_numeric = re.match(r'^-?\d+(\.\d+)?$', str(gia_tri).strip())
            unit_str = f" {unit}" if unit else ""
           
            if is_numeric and yeu_cau:
                if danh_gia == "<=":
                    dat_line = f"- Đạt: {yeu_cau} ≤ {gia_tri}{unit_str}"
                    khong_dat_line = f"- Không đạt: {yeu_cau} > {gia_tri}{unit_str}"
                elif danh_gia == ">=":
                    dat_line = f"- Đạt: {yeu_cau} ≥ {gia_tri}{unit_str}"
                    khong_dat_line = f"- Không đạt: {yeu_cau} < {gia_tri}{unit_str}"
                elif danh_gia == "=":
                    dat_line = f"- Đạt: {yeu_cau} = {gia_tri}{unit_str}"
                    khong_dat_line = f"- Không đạt: {yeu_cau} ≠ {gia_tri}{unit_str}"
                elif danh_gia == "<":
                    dat_line = f"- Đạt: {yeu_cau} < {gia_tri}{unit_str}"
                    khong_dat_line = f"- Không đạt: {yeu_cau} ≥ {gia_tri}{unit_str}"
                elif danh_gia == ">":
                    dat_line = f"- Đạt: {yeu_cau} > {gia_tri}{unit_str}"
                    khong_dat_line = f"- Không đạt: {yeu_cau} ≤ {gia_tri}{unit_str}"
                else:
                    dat_line = f"- Đạt: hàng hóa nhà thầu chào đáp ứng yêu cầu kỹ thuật"
                    khong_dat_line = f"- Không đạt: hàng hóa nhà thầu chào không đáp ứng yêu cầu kỹ thuật"
            else:
                dat_line = f"- Đạt: hàng hóa nhà thầu chào đáp ứng yêu cầu kỹ thuật"
                khong_dat_line = f"- Không đạt: hàng hóa nhà thầu chào không đáp ứng yêu cầu kỹ thuật"
           
            # GIẢI THÍCH: Tạo format hiển thị với \r\n
            tieu_chi_display = f"{dat_line}\r\n{khong_dat_line}"
           
            # GIẢI THÍCH: Lưu format raw với \n vào custom_indicators
            tieu_chi_raw = f"{dat_line}\n{khong_dat_line}"
            self.custom_indicators.setdefault(self.current_tab, {})[f"tieu_chi_{ind_id}"] = tieu_chi_raw
           
            # GIẢI THÍCH: Cập nhật trực tiếp vào tree
            new_values = list(values)
            new_values[tieu_chi_col_idx] = tieu_chi_display
            self.current_tree.item(item, values=tuple(new_values))
           
        except Exception as e:
            print(f"DEBUG AUTO UPDATE TIEU CHI ERROR: {str(e)}")

    def open_tieu_chi_editor(self, tree, item, column, ind_id_str, current_tieu_chi):
        """
        GIẢI THÍCH: Mở cửa sổ chỉnh sửa tiêu chí đánh giá
        Parse chính xác với format "- Đạt: " và "- Không đạt: "
        SỬA: Sử dụng grid nhất quán để tránh xung đột geometry manager
        """
        values = tree.item(item)['values']
       
        try:
            chi_tieu_idx = self.current_tree["columns"].index("Mã chỉ tiêu")
            yeu_cau_idx = self.current_tree["columns"].index("Chỉ tiêu kỹ thuật chi tiết")
            gia_tri_idx = self.current_tree["columns"].index("Giá trị")
            don_vi_idx = self.current_tree["columns"].index("Đơn vị")
            tieu_chi_idx = self.current_tree["columns"].index("Tiêu chí đánh giá")
        except ValueError:
            messagebox.showerror("Lỗi", "Không tìm thấy cột cần thiết")
            return
       
        chi_tieu = values[chi_tieu_idx].replace('\n', ' ') if chi_tieu_idx < len(values) else ""
        yeu_cau = values[yeu_cau_idx].replace('\n', ' ') if yeu_cau_idx < len(values) else ""
        gia_tri = values[gia_tri_idx].replace('\n', ' ') if gia_tri_idx < len(values) else ""
        don_vi = values[don_vi_idx].replace('\n', ' ') if don_vi_idx < len(values) else ""
       
        # GIẢI THÍCH: Chuẩn hóa input - chuyển từ \r\n về \n
        current_tieu_chi = str(current_tieu_chi).replace('\r\n', '\n').replace('\r', '\n').strip()
       
        dat_text = ""
        khong_dat_text = ""
       
        if current_tieu_chi:
            lines = current_tieu_chi.split('\n')
            for line in lines:
                line = line.strip()
                if line.startswith("- Đạt:"):
                    dat_text = line[7:].strip() # Bỏ "- Đạt: "
                elif line.startswith("- Không đạt:"):
                    khong_dat_text = line[13:].strip() # Bỏ "- Không đạt: "
       
        # GIẢI THÍCH: Tạo cửa sổ chỉnh sửa
        edit_win = tk.Toplevel(self.root)
        edit_win.title("Chỉnh sửa tiêu chí đánh giá")
        edit_win.geometry("1100x650")
        edit_win.resizable(True, True)
        edit_win.grid_rowconfigure(0, weight=1)
        edit_win.grid_columnconfigure(0, weight=1)
       
        main_frame = tk.Frame(edit_win, bg="white")
        main_frame.grid(row=0, column=0, sticky="nsew", padx=20, pady=20)
        main_frame.grid_rowconfigure(2, weight=1)
        main_frame.grid_columnconfigure(0, weight=1)
       
        title_label = tk.Label(main_frame, text="Chỉnh sửa tiêu chí đánh giá",
                            font=("Arial", 20, "bold"), bg="white")
        title_label.grid(row=0, column=0, sticky="ew", pady=(0, 25))
       
        info_frame = tk.LabelFrame(main_frame, text="Thông tin chỉ tiêu",
                                font=("Arial", 12, "bold"), bg="white")
        info_frame.grid(row=1, column=0, sticky="ew", pady=(0, 15))
        info_frame.grid_columnconfigure(1, weight=1)
       
        tk.Label(info_frame, text="Mã chỉ tiêu:",
                font=("Arial", 11, "bold"), bg="white").grid(row=0, column=0, sticky="w", padx=15, pady=12)
        chi_tieu_label = tk.Label(info_frame, text=chi_tieu, relief="solid", borderwidth=1, bg="white",
                                font=("Arial", 10), anchor="w", justify="left", wraplength=700)
        chi_tieu_label.grid(row=0, column=1, sticky="ew", padx=15, pady=12)
       
        tk.Label(info_frame, text="Chỉ tiêu kỹ thuật chi tiết:",
                font=("Arial", 11, "bold"), bg="white").grid(row=1, column=0, sticky="w", padx=15, pady=8)
        yeu_cau_label = tk.Label(info_frame, text=yeu_cau, relief="solid", borderwidth=1, bg="white",
                                font=("Arial", 10), anchor="w", justify="left", wraplength=700)
        yeu_cau_label.grid(row=1, column=1, sticky="ew", padx=15, pady=8)
       
        tk.Label(info_frame, text="Giá trị:", font=("Arial", 11, "bold"), bg="white").grid(row=2, column=0, sticky="w", padx=15, pady=8)
        gia_tri_frame = tk.Frame(info_frame, bg="white")
        gia_tri_frame.grid(row=2, column=1, sticky="ew", padx=15, pady=8)
        gia_tri_entry = tk.Entry(gia_tri_frame, width=35, font=("Arial", 11), relief="solid", borderwidth=1)
        gia_tri_entry.grid(row=0, column=0, sticky="ew")
        gia_tri_frame.grid_columnconfigure(0, weight=1)
        gia_tri_entry.insert(0, gia_tri)
       
        tk.Label(info_frame, text="Đơn vị:", font=("Arial", 11, "bold"), bg="white").grid(row=3, column=0, sticky="w", padx=15, pady=8)
        don_vi_label = tk.Label(info_frame, text=don_vi, relief="solid", borderwidth=1, bg="white",
                            font=("Arial", 11), width=20, anchor="center")
        don_vi_label.grid(row=3, column=1, sticky="w", padx=15, pady=8)
       
        # GIẢI THÍCH: Frame tiêu chí đánh giá
        tieu_chi_frame = tk.LabelFrame(main_frame, text="Tiêu chí đánh giá",
                                    font=("Arial", 12, "bold"), bg="white")
        tieu_chi_frame.grid(row=2, column=0, sticky="nsew", pady=(0, 20))
        tieu_chi_frame.grid_columnconfigure(0, weight=1)
        tieu_chi_frame.grid_columnconfigure(1, weight=1)
        tieu_chi_frame.grid_rowconfigure(1, weight=1)
       
        # GIẢI THÍCH: Cột Đạt
        dat_label_frame = tk.Frame(tieu_chi_frame, bg="white")
        dat_label_frame.grid(row=0, column=0, sticky="w", padx=15, pady=12)
        tk.Label(dat_label_frame, text="Đạt:", font=("Arial", 13, "bold"), bg="white").grid(row=0, column=0)
        tk.Label(dat_label_frame, text=" (Điều kiện để đạt yêu cầu)",
                font=("Arial", 10), bg="white").grid(row=0, column=1, padx=(5, 0))
       
        dat_entry = tk.Text(tieu_chi_frame, height=10, width=55, wrap=tk.WORD,
                        font=("Arial", 11), relief="solid", borderwidth=1)
        dat_entry.grid(row=1, column=0, padx=15, pady=8, sticky="nsew")
        dat_entry.insert("1.0", dat_text)
        dat_entry.config(state=tk.NORMAL)
       
        # GIẢI THÍCH: Cột Không đạt
        khong_dat_label_frame = tk.Frame(tieu_chi_frame, bg="white")
        khong_dat_label_frame.grid(row=0, column=1, sticky="w", padx=15, pady=12)
        tk.Label(khong_dat_label_frame, text="Không đạt:", font=("Arial", 13, "bold"), bg="white").grid(row=0, column=0)
        tk.Label(khong_dat_label_frame, text=" (Điều kiện không đạt yêu cầu)",
                font=("Arial", 10), bg="white").grid(row=0, column=1, padx=(5, 0))
       
        khong_dat_entry = tk.Text(tieu_chi_frame, height=10, width=55, wrap=tk.WORD,
                                font=("Arial", 11), relief="solid", borderwidth=1)
        khong_dat_entry.grid(row=1, column=1, padx=15, pady=8, sticky="nsew")
        khong_dat_entry.insert("1.0", khong_dat_text)
        khong_dat_entry.config(state=tk.NORMAL)
       
        # GIẢI THÍCH: Button frame
        button_frame = tk.Frame(main_frame, bg="white")
        button_frame.grid(row=3, column=0, sticky="ew", pady=25)
        button_frame.grid_columnconfigure(0, weight=1)
       
        def save_tieu_chi():
            """GIẢI THÍCH: Lưu tiêu chí đánh giá"""
            new_gia_tri = gia_tri_entry.get().strip()
            new_dat = dat_entry.get("1.0", tk.END).strip()
            new_khong_dat = khong_dat_entry.get("1.0", tk.END).strip()
           
            if not new_dat and not new_khong_dat:
                messagebox.showerror("Lỗi", "Cần ít nhất một điều kiện!")
                return
           
            # GIẢI THÍCH: Tạo format tiêu chí với dấu "- "
            parts = []
            if new_dat:
                parts.append(f"- Đạt: {new_dat}")
            if new_khong_dat:
                parts.append(f"- Không đạt: {new_khong_dat}")
           
            # GIẢI THÍCH: Format raw với \n cho lưu trữ
            new_tieu_chi_raw = "\n".join(parts)
           
            # GIẢI THÍCH: Format hiển thị với \r\n cho treeview
            new_tieu_chi_display = new_tieu_chi_raw.replace('\n', '\r\n')
           
            try:
                if item in tree.get_children():
                    # GIẢI THÍCH: Cập nhật trực tiếp vào tree
                    tieu_chi_col_num = f"#{tieu_chi_idx + 1}"
                    tree.set(item, tieu_chi_col_num, new_tieu_chi_display)
                   
                    # GIẢI THÍCH: Cập nhật giá trị nếu có thay đổi
                    if new_gia_tri and new_gia_tri != gia_tri:
                        gia_tri_col_num = f"#{gia_tri_idx + 1}"
                        tree.set(item, gia_tri_col_num, wrap_text(new_gia_tri, 20))
                       
                        # GIẢI THÍCH: Cập nhật custom_indicators
                        key = f"so_sanh_{ind_id_str}"
                        self.custom_indicators.setdefault(self.current_tab, {})[key] = new_gia_tri
                       
                        # GIẢI THÍCH: Kiểm tra và đánh dấu row
                        self.check_and_mark_row(tree, item, ind_id_str, new_gia_tri, new_tieu_chi_raw, self.current_tab)
                   
                    # GIẢI THÍCH: Lưu format raw vào custom_indicators
                    tieu_chi_key = f"tieu_chi_{ind_id_str}"
                    self.custom_indicators.setdefault(self.current_tab, {})[tieu_chi_key] = new_tieu_chi_raw
                   
                    # GIẢI THÍCH: Cập nhật giao diện
                    tree.update_idletasks()
                   
                    edit_win.destroy()
                    messagebox.showinfo("Thành công", "Đã cập nhật tiêu chí đánh giá!")
                   
            except Exception as e:
                messagebox.showerror("Lỗi", f"Lỗi khi lưu: {str(e)}")
       
        def cancel_edit():
            """GIẢI THÍCH: Hủy chỉnh sửa"""
            edit_win.destroy()
       
        button_inner_frame = tk.Frame(button_frame, bg="white")
        button_inner_frame.grid(row=0, column=0, sticky="e")
       
        save_button = tk.Button(button_inner_frame, text="Lưu thay đổi", command=save_tieu_chi,
                            fg="black", font=("Arial", 12, "bold"),
                            width=18, height=2, relief="flat", bd=0, cursor="hand2")
        save_button.grid(row=0, column=0, padx=(0, 15))
       
        cancel_button = tk.Button(button_inner_frame, text="Hủy bỏ", command=cancel_edit,
                                fg="black", font=("Arial", 12, "bold"),
                                width=15, height=2, relief="flat", bd=0, cursor="hand2")
        cancel_button.grid(row=0, column=1, padx=(0, 15))
       
        # GIẢI THÍCH: Focus vào ô đầu tiên
        if dat_text:
            dat_entry.focus_set()
        elif khong_dat_text:
            khong_dat_entry.focus_set()
        else:
            dat_entry.focus_set()
       
        edit_win.bind('<Return>', lambda e: save_tieu_chi())
        edit_win.bind('<Escape>', lambda e: cancel_edit())
       
        # GIẢI THÍCH: Hover effects
        def on_enter(button):
            button.config(relief="raised", bd=1)
        def on_leave(button):
            button.config(relief="flat", bd=0)
       
        save_button.bind('<Enter>', lambda e: on_enter(save_button))
        save_button.bind('<Leave>', lambda e: on_leave(save_button))
        cancel_button.bind('<Enter>', lambda e: on_enter(cancel_button))
        cancel_button.bind('<Leave>', lambda e: on_leave(cancel_button))

    def restore_indicators(self):
        """
        GIẢI THÍCH: Khôi phục các chỉ tiêu đã xóa
        - Chỉ khôi phục chỉ tiêu từ origin_deleted của tab hiện tại
        - Cascade khôi phục cho các tab phụ thuộc
        """
        if not self.current_tab:
            messagebox.showerror("Lỗi", "Chọn tab để khôi phục")
            return
        
        # GIẢI THÍCH: Lấy các chỉ tiêu đã xóa từ origin_deleted của tab hiện tại
        hidden_ids = self.origin_deleted.get(self.current_tab, set())
        if not hidden_ids:
            messagebox.showinfo("Thông báo", f"Tab '{self.current_tab}' không có chỉ tiêu nào bị xóa để khôi phục")
            return
        
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        
        # GIẢI THÍCH: Lấy thông tin chi tiết của các chỉ tiêu đã xóa
        hidden_indicators = []
        for ind_id in hidden_ids:
            indicator_info = c.execute("SELECT indicator_code, indicator FROM indicators WHERE id=?", (ind_id,)).fetchone()
            if indicator_info:
                hidden_indicators.append((ind_id, indicator_info[0], indicator_info[1]))
        
        conn.close()
        
        if not hidden_indicators:
            messagebox.showinfo("Thông báo", f"Tab '{self.current_tab}' không có chỉ tiêu nào bị xóa để khôi phục")
            return
        
        # GIẢI THÍCH: Tạo cửa sổ khôi phục
        restore_win = tk.Toplevel(self.root)
        restore_win.title(f"Khôi phục chỉ tiêu - Tab {self.current_tab}")
        restore_win.geometry("600x400")
        
        tree_frame = tk.Frame(restore_win)
        tree_frame.pack(fill=tk.BOTH, expand=True)
        
        v_scrollbar = tk.Scrollbar(tree_frame, orient="vertical")
        v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        restore_tree = ttk.Treeview(tree_frame, columns=("STT", "Mã chỉ tiêu", "Chỉ tiêu"),
                                    show="headings", yscrollcommand=v_scrollbar.set, selectmode="extended")
        v_scrollbar.config(command=restore_tree.yview)
        
        restore_tree.heading("STT", text="STT")
        restore_tree.heading("Mã chỉ tiêu", text="Mã chỉ tiêu")
        restore_tree.heading("Chỉ tiêu", text="Chỉ tiêu")
        restore_tree.column("STT", width=50, anchor="center")
        restore_tree.column("Mã chỉ tiêu", width=250, anchor="w")
        restore_tree.column("Chỉ tiêu", width=250, anchor="w")
        restore_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # GIẢI THÍCH: Thêm các chỉ tiêu đã xóa vào tree
        for index, (ind_id, req, indc) in enumerate(hidden_indicators, 1):
            item = restore_tree.insert("", "end", values=(index, wrap_text(req, 40), wrap_text(indc, 40)),
                                    tags=(str(ind_id),)) # Lưu ind_id trong tags
        
        button_frame = tk.Frame(restore_win)
        button_frame.pack(pady=5)
        
        def do_restore_selected():
            """GIẢI THÍCH: Khôi phục các chỉ tiêu đã chọn"""
            selected = restore_tree.selection()
            if not selected:
                messagebox.showerror("Lỗi", "Chọn chỉ tiêu để khôi phục")
                return
            
            restored_count = 0
            
            # GIẢI THÍCH: Định nghĩa các tab phụ thuộc
            cascade_map = {
                "three_brands": ["bom", "dmkt", "ctkt_bo", "ctkt_mua_sam"],
                "bom": ["dmkt", "ctkt_bo"],
                "dmkt": ["ctkt_bo"],
                "ctkt_mua_sam": [],
                "ctkt_bo": []
            }
            
            dependent_tabs = cascade_map.get(self.current_tab, [])
            print(f"DEBUG KHÔI PHỤC: Tab hiện tại '{self.current_tab}' - cascade khôi phục cho {dependent_tabs}")
            
            # SỬA MỚI: Set để lưu tất cả id cần khôi phục (tránh trùng)
            all_ids_to_restore = set()
            
            conn = sqlite3.connect(DB_NAME)
            c = conn.cursor()
            
            for item in selected:
                tags = restore_tree.item(item)['tags']
                if tags:
                    ind_id_str = tags[0]
                    ind_id = int(ind_id_str)
                else:
                    print(f"DEBUG LỖI: Không tìm thấy ind_id trong tags của item {item}")
                    continue
                
                # Thêm chính ind_id
                all_ids_to_restore.add(ind_id)
                
                # SỬA MỚI: Thêm tất cả con đệ quy nếu là cha
                descendants = self.get_all_descendants(ind_id, c)
                all_ids_to_restore.update(descendants)
                
                # SỬA MỚI: Thêm tất cả tổ tiên (cha, ông,...) nếu bị xóa
                ancestors = self.get_all_ancestors(ind_id, c)
                all_ids_to_restore.update(ancestors)
            
            conn.close()
            
            # GIẢI THÍCH: Khôi phục tất cả id trong set
            for ind_id in all_ids_to_restore:
                # GIẢI THÍCH: Khôi phục gốc - xóa khỏi origin_deleted và deleted_indicators của tab hiện tại
                if self.current_tab in self.origin_deleted and ind_id in self.origin_deleted[self.current_tab]:
                    self.origin_deleted[self.current_tab].discard(ind_id)
                    restored_count += 1
                    print(f"DEBUG KHÔI PHỤC: Xóa {ind_id} khỏi origin_deleted của tab '{self.current_tab}'")
                
                if self.current_tab in self.deleted_indicators and ind_id in self.deleted_indicators[self.current_tab]:
                    self.deleted_indicators[self.current_tab].discard(ind_id)
                    print(f"DEBUG KHÔI PHỤC: Xóa {ind_id} khỏi deleted_indicators của tab '{self.current_tab}'")
                
                # GIẢI THÍCH: Cascade khôi phục - xóa khỏi deleted_indicators của các tab phụ thuộc
                for dep_tab in dependent_tabs:
                    if dep_tab in self.deleted_indicators and ind_id in self.deleted_indicators[dep_tab]:
                        self.deleted_indicators[dep_tab].discard(ind_id)
                        print(f"DEBUG CASCADE KHÔI PHỤC: Xóa {ind_id} khỏi deleted_indicators của tab phụ thuộc '{dep_tab}'")
            
            if restored_count > 0:
                # GIẢI THÍCH: Reload tab hiện tại và các tab phụ thuộc
                print(f"DEBUG KHÔI PHỤC: Bắt đầu reload {restored_count} chỉ tiêu đã khôi phục")
                self.load_single_tab(self.current_tab)
                
                for dep_tab in dependent_tabs:
                    print(f"DEBUG CASCADE RELOAD: Reload tab phụ thuộc '{dep_tab}'")
                    self.load_single_tab(dep_tab)
                
                restore_win.destroy()
                messagebox.showinfo("Thành công", f"Đã khôi phục {restored_count} chỉ tiêu từ tab '{self.current_tab}' và cascade cho {len(dependent_tabs)} tab phụ thuộc")
            else:
                messagebox.showwarning("Thông báo", "Không có chỉ tiêu nào được khôi phục")
        
        tk.Button(button_frame, text="Khôi phục", command=do_restore_selected).pack(side=tk.LEFT, padx=5)
        tk.Button(button_frame, text="Hủy", command=restore_win.destroy).pack(side=tk.LEFT, padx=5)

    def get_all_descendants(self, ind_id, cursor):
        """
        GIẢI THÍCH: Helper function để lấy tất cả id con đệ quy của một chỉ tiêu (dựa trên indicator_code LIKE code + '.%')
        - Trả về set các id con (không bao gồm chính ind_id)
        """
        descendants = set()
        
        # Lấy code của ind_id
        code = cursor.execute("SELECT indicator_code FROM indicators WHERE id=?", (ind_id,)).fetchone()
        if not code:
            return descendants
        code = code[0]
        
        # Query tất cả id có code LIKE code + '.%'
        con_rows = cursor.execute(
            "SELECT id FROM indicators WHERE indicator_code LIKE ? AND type_id = (SELECT type_id FROM indicators WHERE id = ?) AND id != ?",
            (code + '.%', ind_id, ind_id)
        ).fetchall()
        
        for row in con_rows:
            con_id = row[0]
            descendants.add(con_id)
            # Đệ quy lấy con của con
            descendants.update(self.get_all_descendants(con_id, cursor))
        
        return descendants

    def get_all_ancestors(self, ind_id, cursor):
        """
        GIẢI THÍCH: Helper function để lấy tất cả id tổ tiên (cha, ông,...) của một chỉ tiêu nếu bị xóa
        - Dựa trên cắt dần indicator_code (ví dụ 1.2.1 -> 1.2 -> 1)
        - Chỉ thêm nếu tổ tiên bị xóa (trong origin_deleted của tab hiện tại)
        - Trả về set các id tổ tiên
        """
        ancestors = set()
        
        # Lấy code của ind_id
        code = cursor.execute("SELECT indicator_code FROM indicators WHERE id=?", (ind_id,)).fetchone()
        if not code:
            return ancestors
        code = code[0]
        
        # Cắt dần code để tìm cha
        while '.' in code:
            # Cắt đến dấu chấm cuối cùng (cha trực tiếp)
            parent_code = code.rsplit('.', 1)[0]
            
            # Query id của parent_code
            parent_id_row = cursor.execute(
                "SELECT id FROM indicators WHERE indicator_code = ? AND type_id = (SELECT type_id FROM indicators WHERE id = ?)",
                (parent_code, ind_id)
            ).fetchone()
            
            if parent_id_row:
                parent_id = parent_id_row[0]
                # Chỉ thêm nếu parent bị xóa (trong origin_deleted)
                if self.current_tab in self.origin_deleted and parent_id in self.origin_deleted[self.current_tab]:
                    ancestors.add(parent_id)
                    # Tiếp tục đệ quy lên ông
                    ancestors.update(self.get_all_ancestors(parent_id, cursor))
            
            # Cập nhật code cho lần cắt tiếp
            code = parent_code
        
        return ancestors

    def update_cascade_marking(self, ind_id_str, changed_tab):
        """
        GIẢI THÍCH: Cập nhật giá trị cascade sau khi một tab thay đổi giá trị
        - Cập nhật giá trị cho các tab phụ thuộc
        - Đặc biệt cập nhật tiêu chí đánh giá cho tab CTKT mua sắm
        - SỬA: Reset giá trị cascade theo yêu cầu mới
        """
        # GIẢI THÍCH: Skip cascade nếu ind_id âm (custom)
        if ind_id_str.startswith('-'):
            return
       
        # GIẢI THÍCH: Định nghĩa thứ tự phụ thuộc
        tab_hierarchy = ["three_brands", "bom", "dmkt", "ctkt_bo"]
       
        try:
            changed_index = tab_hierarchy.index(changed_tab)
        except ValueError:
            # GIẢI THÍCH: Nếu tab không nằm trong hierarchy, chỉ xử lý ctkt_mua_sam nếu changed_tab là three_brands
            if changed_tab == "three_brands":
                dependent_tab = "ctkt_mua_sam"
                tree = getattr(self, f"{dependent_tab}_tree", None)
                indicator_map = getattr(self, f"{dependent_tab}_indicator_map", {})
               
                if tree:
                    # GIẢI THÍCH: Lấy giá trị từ three_brands
                    current_value = self.custom_indicators.get(changed_tab, {}).get(f"so_sanh_{ind_id_str}", "")
                    if not current_value:
                        danh_gia = self.custom_indicators.get("three_brands", {}).get(f"danh_gia_{ind_id_str}", "not")
                        current_value = self.calculate_extreme_value(int(ind_id_str), getattr(self, f"{changed_tab}_tree", None), self.reference_columns, danh_gia)
                   
                    for item, mapped_ind_id in indicator_map.items():
                        if str(mapped_ind_id) == ind_id_str and not str(mapped_ind_id).startswith('-'):
                            # GIẢI THÍCH: Cập nhật giá trị so_sanh trong ctkt_mua_sam
                            self.custom_indicators.setdefault(dependent_tab, {})[f"so_sanh_{ind_id_str}"] = current_value
                           
                            try:
                                # GIẢI THÍCH: Lấy các cột cần thiết để tạo lại tiêu chí
                                yeu_cau_col_idx = tree["columns"].index("Chỉ tiêu kỹ thuật chi tiết")
                                unit_col_idx = tree["columns"].index("Đơn vị")
                                tieu_chi_col_idx = tree["columns"].index("Tiêu chí đánh giá")
                                gia_tri_col_idx = tree["columns"].index("Giá trị")
                               
                                yeu_cau = tree.set(item, f"#{yeu_cau_col_idx + 1}").replace('\n', ' ')
                                unit = tree.set(item, f"#{unit_col_idx + 1}").replace('\n', ' ')
                               
                                # GIẢI THÍCH: Cập nhật giá trị trong tree
                                tree.set(item, f"#{gia_tri_col_idx + 1}", wrap_text(current_value, 20))
                               
                                # GIẢI THÍCH: Tạo tiêu chí mới dựa trên danh_gia
                                danh_gia = self.custom_indicators.get("three_brands", {}).get(f"danh_gia_{ind_id_str}", "not")
                                is_numeric = re.match(r'^-?\d+(\.\d+)?$', str(current_value).strip())
                                unit_str = f" {unit}" if unit else ""
                               
                                if is_numeric and yeu_cau:
                                    if danh_gia == "<=":
                                        new_tieu_chi = f"- Đạt: {yeu_cau} ≤ {current_value}{unit_str}\n- Không đạt: {yeu_cau} > {current_value}{unit_str}"
                                    elif danh_gia == ">=":
                                        new_tieu_chi = f"- Đạt: {yeu_cau} ≥ {current_value}{unit_str}\n- Không đạt: {yeu_cau} < {current_value}{unit_str}"
                                    elif danh_gia == "=":
                                        new_tieu_chi = f"- Đạt: {yeu_cau} = {current_value}{unit_str}\n- Không đạt: {yeu_cau} ≠ {current_value}{unit_str}"
                                    elif danh_gia == "<":
                                        new_tieu_chi = f"- Đạt: {yeu_cau} < {current_value}{unit_str}\n- Không đạt: {yeu_cau} ≥ {current_value}{unit_str}"
                                    elif danh_gia == ">":
                                        new_tieu_chi = f"- Đạt: {yeu_cau} > {current_value}{unit_str}\n- Không đạt: {yeu_cau} ≤ {current_value}{unit_str}"
                                    else:
                                        new_tieu_chi = f"- Đạt: hàng hóa nhà thầu chào đáp ứng yêu cầu kỹ thuật\n- Không đạt: hàng hóa nhà thầu chào không đáp ứng yêu cầu kỹ thuật"
                                else:
                                    new_tieu_chi = f"- Đạt: hàng hóa nhà thầu chào đáp ứng yêu cầu kỹ thuật\n- Không đạt: hàng hóa nhà thầu chào không đáp ứng yêu cầu kỹ thuật"
                               
                                # GIẢI THÍCH: Lưu tiêu chí mới
                                tieu_chi_key = f"tieu_chi_{ind_id_str}"
                                self.custom_indicators[dependent_tab][tieu_chi_key] = new_tieu_chi
                               
                                # GIẢI THÍCH: Cập nhật trực tiếp trong tree
                                tieu_chi_col_num = f"#{tieu_chi_col_idx + 1}"
                                tree.set(item, tieu_chi_col_num, new_tieu_chi.replace('\n', '\r\n'))
                               
                                # GIẢI THÍCH: Kiểm tra và đánh dấu row
                                self.check_and_mark_row(tree, item, ind_id_str, current_value, new_tieu_chi, dependent_tab)
                               
                            except Exception as e:
                                print(f"DEBUG CASCADE TIEU CHI ERROR: {str(e)}")
                            break
                return
            else:
                return
       
        # GIẢI THÍCH: Lấy giá trị mới từ tab đã thay đổi
        current_value = self.custom_indicators.get(changed_tab, {}).get(
            f"so_sanh_{ind_id_str}" if changed_tab != "ctkt_bo" else f"gia_tri_{ind_id_str}", "")
        if not current_value:
            danh_gia = self.custom_indicators.get("three_brands", {}).get(f"danh_gia_{ind_id_str}", "not")
            current_value = self.calculate_extreme_value(int(ind_id_str), getattr(self, f"{changed_tab}_tree", None), self.reference_columns, danh_gia)
       
        # GIẢI THÍCH: Cập nhật tất cả các tab sau tab đã thay đổi
        for i in range(changed_index + 1, len(tab_hierarchy)):
            dependent_tab = tab_hierarchy[i]
            tree = getattr(self, f"{dependent_tab}_tree", None)
            indicator_map = getattr(self, f"{dependent_tab}_indicator_map", {})
           
            if not tree:
                continue
           
            # GIẢI THÍCH: Tìm item tương ứng với ind_id_str
            for item, mapped_ind_id in indicator_map.items():
                if str(mapped_ind_id) == ind_id_str:
                    # GIẢI THÍCH: Cập nhật giá trị trong custom_indicators
                    if dependent_tab == "bom":
                        self.custom_indicators.setdefault(dependent_tab, {})[f"so_sanh_{ind_id_str}"] = current_value
                    elif dependent_tab == "dmkt":
                        self.custom_indicators.setdefault(dependent_tab, {})[f"so_sanh_{ind_id_str}"] = current_value
                    elif dependent_tab == "ctkt_bo":
                        self.custom_indicators.setdefault(dependent_tab, {})[f"gia_tri_{ind_id_str}"] = current_value
                   
                    # GIẢI THÍCH: Reload tab để cập nhật giao diện
                    self.load_single_tab(dependent_tab)
                    break
       
        # GIẢI THÍCH: Xử lý cascade cho ctkt_mua_sam
        dependent_tab = "ctkt_mua_sam"
        tree = getattr(self, f"{dependent_tab}_tree", None)
        indicator_map = getattr(self, f"{dependent_tab}_indicator_map", {})
       
        if tree and changed_tab == "three_brands":
            for item, mapped_ind_id in indicator_map.items():
                if str(mapped_ind_id) == ind_id_str and not str(mapped_ind_id).startswith('-'):
                    # GIẢI THÍCH: Cập nhật giá trị so_sanh
                    self.custom_indicators.setdefault(dependent_tab, {})[f"so_sanh_{ind_id_str}"] = current_value
                   
                    # GIẢI THÍCH: Tạo lại tiêu chí đánh giá
                    try:
                        yeu_cau_col_idx = tree["columns"].index("Chỉ tiêu kỹ thuật chi tiết")
                        unit_col_idx = tree["columns"].index("Đơn vị")
                        tieu_chi_col_idx = tree["columns"].index("Tiêu chí đánh giá")
                        gia_tri_col_idx = tree["columns"].index("Giá trị")
                       
                        yeu_cau = tree.set(item, f"#{yeu_cau_col_idx + 1}").replace('\n', ' ')
                        unit = tree.set(item, f"#{unit_col_idx + 1}").replace('\n', ' ')
                       
                        # GIẢI THÍCH: Cập nhật giá trị trong tree
                        tree.set(item, f"#{gia_tri_col_idx + 1}", wrap_text(current_value, 20))
                       
                        # GIẢI THÍCH: Tạo tiêu chí mới dựa trên danh_gia
                        danh_gia = self.custom_indicators.get("three_brands", {}).get(f"danh_gia_{ind_id_str}", "not")
                        is_numeric = re.match(r'^-?\d+(\.\d+)?$', str(current_value).strip())
                        unit_str = f" {unit}" if unit else ""
                       
                        if is_numeric and yeu_cau:
                            if danh_gia == "<=":
                                new_tieu_chi = f"- Đạt: {yeu_cau} ≤ {current_value}{unit_str}\n- Không đạt: {yeu_cau} > {current_value}{unit_str}"
                            elif danh_gia == ">=":
                                new_tieu_chi = f"- Đạt: {yeu_cau} ≥ {current_value}{unit_str}\n- Không đạt: {yeu_cau} < {current_value}{unit_str}"
                            elif danh_gia == "=":
                                new_tieu_chi = f"- Đạt: {yeu_cau} = {current_value}{unit_str}\n- Không đạt: {yeu_cau} ≠ {current_value}{unit_str}"
                            elif danh_gia == "<":
                                new_tieu_chi = f"- Đạt: {yeu_cau} < {current_value}{unit_str}\n- Không đạt: {yeu_cau} ≥ {current_value}{unit_str}"
                            elif danh_gia == ">":
                                new_tieu_chi = f"- Đạt: {yeu_cau} > {current_value}{unit_str}\n- Không đạt: {yeu_cau} ≤ {current_value}{unit_str}"
                            else:
                                new_tieu_chi = f"- Đạt: hàng hóa nhà thầu chào đáp ứng yêu cầu kỹ thuật\n- Không đạt: hàng hóa nhà thầu chào không đáp ứng yêu cầu kỹ thuật"
                        else:
                            new_tieu_chi = f"- Đạt: hàng hóa nhà thầu chào đáp ứng yêu cầu kỹ thuật\n- Không đạt: hàng hóa nhà thầu chào không đáp ứng yêu cầu kỹ thuật"
                       
                        # GIẢI THÍCH: Lưu tiêu chí mới
                        tieu_chi_key = f"tieu_chi_{ind_id_str}"
                        self.custom_indicators[dependent_tab][tieu_chi_key] = new_tieu_chi
                       
                        # GIẢI THÍCH: Cập nhật trực tiếp trong tree
                        tieu_chi_col_num = f"#{tieu_chi_col_idx + 1}"
                        tree.set(item, tieu_chi_col_num, new_tieu_chi.replace('\n', '\r\n'))
                       
                        # GIẢI THÍCH: Kiểm tra và đánh dấu row
                        self.check_and_mark_row(tree, item, ind_id_str, current_value, new_tieu_chi, dependent_tab)
                       
                    except Exception as e:
                        print(f"DEBUG CASCADE TIEU CHI ERROR: {str(e)}")
                    break

    # Sửa hàm add_custom_indicator
    def add_custom_indicator(self):
        """
        GIẢI THÍCH: Thêm chỉ tiêu tùy chỉnh cho tab CTKT mua sắm - ĐỒNG BỘ với project_manager.py
        """
        if self.current_tab != "ctkt_mua_sam":
            return
       
        # Tạo cửa sổ thêm chỉ tiêu
        add_win = tk.Toplevel(self.root)
        add_win.title("Thêm chỉ tiêu tùy chỉnh cho Yêu cầu khác")
        add_win.geometry("800x600")
        add_win.resizable(True, True)
        add_win.grid_rowconfigure(0, weight=1)
        add_win.grid_columnconfigure(0, weight=1)
       
        main_frame = tk.Frame(add_win, padx=20, pady=20)
        main_frame.grid(row=0, column=0, sticky="nsew")
        main_frame.grid_rowconfigure(1, weight=1)
        main_frame.grid_columnconfigure(0, weight=1)
       
        title_label = tk.Label(main_frame, text="Thêm chỉ tiêu tùy chỉnh cho Yêu cầu khác",
                            font=("Arial", 14, "bold"))
        title_label.grid(row=0, column=0, sticky="ew", pady=(0, 20))
       
        input_frame = tk.LabelFrame(main_frame, text="Thông tin chỉ tiêu",
                                    font=("Arial", 12, "bold"), padx=10, pady=10)
        input_frame.grid(row=1, column=0, sticky="nsew", pady=(0, 20))
        input_frame.grid_rowconfigure(5, weight=1)
        input_frame.grid_columnconfigure(1, weight=1)
       
        # Chỉ tiêu kỹ thuật chi tiết
        tk.Label(input_frame, text="Mã chỉ tiêu:",
                font=("Arial", 11, "bold")).grid(row=0, column=0, sticky="w", padx=(0, 10), pady=(0, 5))
        chi_tieu_entry = tk.Text(input_frame, height=3, width=50, wrap=tk.WORD,
                                font=("Arial", 10), relief="solid", borderwidth=1)
        chi_tieu_entry.grid(row=0, column=1, sticky="ew", pady=(0, 5), padx=(0, 10))
       
        # Yêu cầu kỹ thuật
        tk.Label(input_frame, text="Chỉ tiêu kỹ thuật chi tiết:",
                font=("Arial", 11, "bold")).grid(row=1, column=0, sticky="w", padx=(0, 10), pady=(0, 5))
        yeu_cau_entry = tk.Text(input_frame, height=3, width=50, wrap=tk.WORD,
                            font=("Arial", 10), relief="solid", borderwidth=1)
        yeu_cau_entry.grid(row=1, column=1, sticky="ew", pady=(0, 5), padx=(0, 10))
       
        # Đạt
        tk.Label(input_frame, text="Đạt:", font=("Arial", 11, "bold")).grid(row=2, column=0, sticky="w", padx=(0, 10), pady=(0, 5))
        dat_entry = tk.Text(input_frame, height=4, width=50, wrap=tk.WORD,
                        font=("Arial", 10), relief="solid", borderwidth=1)
        dat_entry.grid(row=2, column=1, sticky="ew", pady=(0, 5), padx=(0, 10))
       
        # Không đạt
        tk.Label(input_frame, text="Không đạt:", font=("Arial", 11, "bold")).grid(row=3, column=0, sticky="w", padx=(0, 10), pady=(0, 5))
        khong_dat_entry = tk.Text(input_frame, height=4, width=50, wrap=tk.WORD,
                                font=("Arial", 10), relief="solid", borderwidth=1)
        khong_dat_entry.grid(row=3, column=1, sticky="ew", pady=(0, 5), padx=(0, 10))
       
        # Loại chỉ tiêu
        tk.Label(input_frame, text="Loại chỉ tiêu:", font=("Arial", 12, "bold")).grid(row=4, column=0, sticky="w", padx=(0, 10), pady=(0, 5))
        crit_type_combo = ttk.Combobox(input_frame, values=["CTCB", "KCB"], width=47, state="readonly", font=("Arial", 12))
        crit_type_combo.set("CTCB")
        crit_type_combo.grid(row=4, column=1, sticky="ew", pady=(0, 5), padx=(0, 10))
   
        # Button frame
        button_frame = tk.Frame(main_frame)
        button_frame.grid(row=2, column=0, pady=20)
        button_frame.grid_columnconfigure(0, weight=1)
       
        def save_single_custom():
            """Lưu 1 chỉ tiêu tùy chỉnh - Tạo format đúng và validate"""
            chi_tieu = chi_tieu_entry.get("1.0", tk.END).strip()
            yeu_cau = yeu_cau_entry.get("1.0", tk.END).strip()
            dat = dat_entry.get("1.0", tk.END).strip()
            khong_dat = khong_dat_entry.get("1.0", tk.END).strip()
            crit_type = crit_type_combo.get()
           
            # Validate bắt buộc - Yêu cầu kỹ thuật phải có
            if not yeu_cau:
                messagebox.showerror("Lỗi", "Vui lòng điền 'Chỉ tiêu kỹ thuật chi tiết'")
                yeu_cau_entry.focus_set()
                return
           
            # Validate - ít nhất một trong hai điều kiện Đạt/Không đạt
            if not dat and not khong_dat:
                messagebox.showerror("Lỗi", "Cần ít nhất một trong hai điều kiện 'Đạt' hoặc 'Không đạt'!")
                if not dat:
                    dat_entry.focus_set()
                else:
                    khong_dat_entry.focus_set()
                return
           
            # Tạo tieu_chi_raw theo format chuẩn
            parts = []
            if dat:
                parts.append(f"- Đạt: {dat}")
            if khong_dat:
                parts.append(f"- Không đạt: {khong_dat}")
           
            tieu_chi_raw = "\n".join(parts)
           
            # Tạo row_data để truyền vào add_single_custom_row_to_tab
            row_data = {
                'chi_tieu': chi_tieu,
                'yeu_cau': yeu_cau,
                'dat': dat,
                'khong_dat': khong_dat,
                'tieu_chi_raw': tieu_chi_raw,
                'crit_type': crit_type
            }
           
            try:
                self.add_single_custom_row_to_tab(row_data)
                add_win.destroy()
                messagebox.showinfo("Thành công", f"Đã thêm chỉ tiêu tùy chỉnh '{chi_tieu[:30]}...' thành công!")
            except Exception as e:
                messagebox.showerror("Lỗi", f"Lỗi khi thêm chỉ tiêu: {str(e)}")
       
        def cancel_add():
            """Hủy thêm chỉ tiêu"""
            add_win.destroy()
       
        # Buttons
        tk.Button(button_frame, text="Thêm", command=save_single_custom,
                fg="black", font=("Arial", 12, "bold"),
                width=12, height=2, relief="raised").pack(side=tk.RIGHT, padx=(0, 10))
        tk.Button(button_frame, text="Hủy", command=cancel_add,
                fg="black", font=("Arial", 12, "bold"),
                width=12, height=2, relief="raised").pack(side=tk.RIGHT, padx=(0, 10))
       
        # Focus vào ô đầu tiên
        chi_tieu_entry.focus_set()
        add_win.bind('<Return>', lambda e: save_single_custom())
        add_win.bind('<Escape>', lambda e: cancel_add())

    def add_single_custom_row_to_tab(self, row_data):
        yeu_cau_khac_item = None
        for child in self.ctkt_mua_sam_tree.get_children():
            if "yeu_cau_khac" in self.ctkt_mua_sam_tree.item(child).get('tags', ()):
                yeu_cau_khac_item = child
                break
       
        # Tính sub_stt cho custom item
        custom_items = [item for item in self.ctkt_mua_sam_tree.get_children() if "custom_single" in self.ctkt_mua_sam_tree.item(item).get('tags', ())]
       
        # Tạo custom_id mới
        custom_ids = set()
        for key in self.custom_indicators.get("ctkt_mua_sam", {}).keys():
            if key.startswith("chi_tieu_") and key.split("_")[-1].startswith("-"):
                custom_ids.add(key.split("_")[-1])
       
        valid_custom_ids = [int(cid) for cid in custom_ids if cid.lstrip('-').isdigit()]
        if valid_custom_ids:
            next_custom_id = min(valid_custom_ids) - 1
        else:
            next_custom_id = -1
        ind_id_str = str(next_custom_id)
       
        dat = row_data.get('dat', '').strip()
        khong_dat = row_data.get('khong_dat', '').strip()
       
        parts = []
        if dat:
            parts.append(f"- Đạt: {dat}")
        if khong_dat:
            parts.append(f"- Không đạt: {khong_dat}")
       
        tieu_chi_raw = "\n".join(parts)
        tieu_chi_display = self.format_tieu_chi_for_display(tieu_chi_raw)
       
        self.custom_indicators.setdefault("ctkt_mua_sam", {})
        self.custom_indicators["ctkt_mua_sam"][f"chi_tieu_{ind_id_str}"] = row_data['chi_tieu']
        self.custom_indicators["ctkt_mua_sam"][f"yeu_cau_{ind_id_str}"] = row_data['yeu_cau']
        self.custom_indicators["ctkt_mua_sam"][f"tieu_chi_{ind_id_str}"] = tieu_chi_raw
        self.custom_indicators["ctkt_mua_sam"][f"crit_type_{ind_id_str}"] = row_data['crit_type']
       
        values = [
            wrap_text(row_data['chi_tieu'], 40),
            wrap_text(row_data['yeu_cau'], 30),
            "",
            "",
            tieu_chi_display,
            row_data['crit_type'],
            "Xóa"
        ]
       
        item = self.ctkt_mua_sam_tree.insert("", "end", values=values, tags=("data_row", "custom_single"))
        self.ctkt_mua_sam_indicator_map[item] = ind_id_str
        self.custom_rows_ctkt_ms.append(item)
       
        line_count = 1
        if tieu_chi_display:
            line_count = tieu_chi_display.count('\r\n') + 1
       
        height = line_count * 60
        self.max_row_heights["ctkt_mua_sam"] = max(self.max_row_heights["ctkt_mua_sam"], height)
       
        style = ttk.Style()
        style.configure("Custom.Treeview", rowheight=self.max_row_heights["ctkt_mua_sam"])
        style.configure("DataRow.Treeview", rowheight=self.max_row_heights["ctkt_mua_sam"])
        
    # Cập nhật hàm on_click_action để xử lý xóa cho tab dmkt và ctkt_mua_sam
    def on_click_action(self, event):
        """
        GIẢI THÍCH: Xử lý sự kiện click vào cột "Hành động" - ĐỒNG BỘ với project_manager.py
        - SỬA: Không cho phép xóa 2 hàng đầu tiên (tên sản phẩm, tên hãng) ở tab Hãng
        """
        if not self.current_tree or not self.current_tab:
            return

        item = self.current_tree.identify_row(event.y)
        if not item:
            return

        column = self.current_tree.identify_column(event.x)
        col_idx = int(column.replace("#", "")) - 1
        columns = self.current_tree["columns"]
        if col_idx < 0 or col_idx >= len(columns):
            return

        col_name = columns[col_idx]
        if col_name != "Hành động":
            return

        ind_id = self.current_indicator_map.get(item)
        if not ind_id:
            return

        # SỬA: Không cho phép xóa 2 hàng đầu tiên (tên sản phẩm, tên hãng) ở tab Hãng
        if isinstance(ind_id, str) and (ind_id.startswith("name_") or ind_id.startswith("product_name_")):
            messagebox.showinfo("Thông báo", "Không thể xóa hàng tên sản phẩm hoặc tên hãng.\nĐể xóa toàn bộ sản phẩm tham khảo, vui lòng sử dụng chức năng xóa sản phẩm tham khảo trong menu chính.")
            return

        # Xử lý xóa chỉ tiêu tùy chỉnh trong tab CTKT mua sắm
        if isinstance(ind_id, str) and ind_id.startswith('-'):
            if not messagebox.askyesno("Xác nhận", "Bạn có chắc muốn xóa chỉ tiêu tùy chỉnh này?"):
                return
            
            self.custom_rows_ctkt_ms = [row for row in self.custom_rows_ctkt_ms if row != item]
            
            for key in list(self.custom_indicators.get(self.current_tab, {}).keys()):
                if key.endswith(ind_id):
                    del self.custom_indicators[self.current_tab][key]
            
            self.load_ctkt_mua_sam_tab()
            messagebox.showinfo("Thành công", "Đã xóa chỉ tiêu tùy chỉnh")

        # Xử lý tab dmkt
        elif self.current_tab == "dmkt":
            values = self.current_tree.item(item)['values']
            
            # SỬA: Vì bỏ phân cấp, không còn header "- ", chỉ xóa hàng đơn lẻ
            if messagebox.askyesno("Xác nhận", "Bạn có chắc muốn xóa chỉ tiêu này?"):
                self.origin_deleted.setdefault("dmkt", set()).add(int(ind_id))
                self.deleted_indicators.setdefault("dmkt", set()).add(int(ind_id))
                self.deleted_indicators.setdefault("ctkt_bo", set()).add(int(ind_id))
                self.current_tree.delete(item)
                if item in self.current_indicator_map:
                    del self.current_indicator_map[item]
                self.load_single_tab("ctkt_bo")

        # Xử lý tab ctkt_mua_sam
        elif self.current_tab == "ctkt_mua_sam":
            tags = self.current_tree.item(item)['tags']
            
            if "yeu_cau_khac" in tags:
                messagebox.showinfo("Thông báo", "Không thể xóa hàng Yêu cầu khác")
                return
            
            if "custom_single" in tags or "custom" in tags:
                if messagebox.askyesno("Xác nhận", "Bạn có chắc muốn xóa chỉ tiêu này?"):
                    for key in list(self.custom_indicators.get(self.current_tab, {})):
                        if key.endswith(str(ind_id)):
                            del self.custom_indicators[self.current_tab][key]
                    self.current_tree.delete(item)
                    if item in self.current_indicator_map:
                        del self.current_indicator_map[item]
                    if item in self.custom_rows_ctkt_ms:
                        self.custom_rows_ctkt_ms.remove(item)
            else:
                if "group_header" in tags:
                    if messagebox.askyesno("Xác nhận", "Bạn có chắc muốn xóa toàn bộ nhóm chỉ tiêu này?"):
                        children = self.current_tree.get_children()
                        my_index = list(children).index(item)
                        group_items = [item]
                        
                        for next_item in children[my_index + 1:]:
                            next_tags = self.current_tree.item(next_item)['tags']
                            if "group_header" in next_tags or "yeu_cau_khac" in next_tags or "custom" in next_tags or "custom_single" in next_tags:
                                break
                            group_items.append(next_item)
                        
                        for g_item in group_items:
                            if g_item in self.current_tree.get_children():
                                g_ind_id = self.current_indicator_map.get(g_item)
                                if g_ind_id and not str(g_ind_id).startswith('-'):
                                    self.origin_deleted.setdefault(self.current_tab, set()).add(int(g_ind_id))
                                    self.deleted_indicators.setdefault(self.current_tab, set()).add(int(g_ind_id))
                                    self.current_tree.delete(g_item)
                                    if g_item in self.current_indicator_map:
                                        del self.current_indicator_map[g_item]
                else:
                    if messagebox.askyesno("Xác nhận", "Bạn có chắc muốn xóa chỉ tiêu này?"):
                        self.origin_deleted.setdefault(self.current_tab, set()).add(int(ind_id))
                        self.deleted_indicators.setdefault(self.current_tab, set()).add(int(ind_id))
                        self.current_tree.delete(item)
                        if item in self.current_indicator_map:
                            del self.current_indicator_map[item]

        # Các tab khác: three_brands, bom, ctkt_bo
        else:
            if "yeu_cau_khac" in self.current_tree.item(item)['tags']:
                messagebox.showinfo("Thông báo", "Không thể xóa hàng Yêu cầu khác")
                return
            
            if "custom" in self.current_tree.item(item)['tags']:
                if messagebox.askyesno("Xác nhận", "Bạn có chắc muốn xóa chỉ tiêu này?"):
                    for key in list(self.custom_indicators.get(self.current_tab, {})):
                        if key.endswith(str(ind_id)):
                            del self.custom_indicators[self.current_tab][key]
                    self.current_tree.delete(item)
                    if item in self.current_indicator_map:
                        del self.current_indicator_map[item]
                    if item in self.custom_rows_ctkt_ms:
                        self.custom_rows_ctkt_ms.remove(item)
            else:
                if messagebox.askyesno("Xác nhận", "Bạn có chắc muốn xóa chỉ tiêu này?"):
                    ind_id_int = int(ind_id)
                    
                    # SỬA: Ép kiểu cha_code thành string để tránh TypeError khi nối chuỗi
                    cha_code = str(self.current_tree.item(item)['values'][0]).strip()  # Lấy Mã chỉ tiêu từ cột đầu tiên
                    
                    conn = sqlite3.connect(DB_NAME)
                    c = conn.cursor()
                    
                    # Kiểm tra nếu là mục cha (có con trực tiếp)
                    has_con = c.execute(
                        "SELECT EXISTS(SELECT 1 FROM indicators WHERE indicator_code LIKE ? AND type_id = (SELECT type_id FROM indicators WHERE id = ?))",
                        (cha_code + '.%', ind_id_int)
                    ).fetchone()[0]
                    
                    # Danh sách id cần xóa: bắt đầu bằng id của mục đang xóa
                    ind_ids_to_delete = [ind_id_int]
                    
                    # Nếu là cha, thêm tất cả id con trực tiếp vào danh sách xóa
                    if has_con:
                        con_rows = c.execute(
                            "SELECT id FROM indicators WHERE indicator_code LIKE ? AND type_id = (SELECT type_id FROM indicators WHERE id = ?)",
                            (cha_code + '.%', ind_id_int)
                        ).fetchall()
                        ind_ids_to_delete += [row[0] for row in con_rows]
                    
                    conn.close()
                    
                    # Tìm tất cả item trong treeview cần xóa (dựa trên id)
                    items_to_delete = []
                    for tree_item in self.current_tree.get_children():
                        tree_ind_id = self.current_indicator_map.get(tree_item)
                        # SỬA MỚI: Chỉ int() nếu tree_ind_id là số (kiểm tra lstrip('-').isdigit() để hỗ trợ id âm nếu có)
                        if tree_ind_id and str(tree_ind_id).lstrip('-').isdigit() and int(tree_ind_id) in ind_ids_to_delete:
                            items_to_delete.append(tree_item)
                    
                    # Cascade map cho các tab phụ thuộc
                    cascade_map = {
                        "three_brands": ["bom", "dmkt", "ctkt_bo", "ctkt_mua_sam"],
                        "bom": ["dmkt", "ctkt_bo"],
                        "dmkt": ["ctkt_bo"],
                        "ctkt_mua_sam": [],
                        "ctkt_bo": []
                    }
                    
                    # Thêm tất cả id vào origin_deleted và deleted_indicators của tab hiện tại
                    for del_id in ind_ids_to_delete:
                        self.origin_deleted.setdefault(self.current_tab, set()).add(del_id)
                        self.deleted_indicators.setdefault(self.current_tab, set()).add(del_id)
                    
                    # Cascade cho tab phụ thuộc
                    dependent_tabs = cascade_map.get(self.current_tab, [])
                    for dep_tab in dependent_tabs:
                        for del_id in ind_ids_to_delete:
                            self.deleted_indicators.setdefault(dep_tab, set()).add(del_id)
                    
                    # Xóa tất cả item tìm được từ treeview
                    for del_item in items_to_delete:
                        if del_item in self.current_tree.get_children():
                            self.current_tree.delete(del_item)
                            if del_item in self.current_indicator_map:
                                del self.current_indicator_map[del_item]
                    
                    # Reload tab hiện tại và các tab phụ thuộc
                    self.load_single_tab(self.current_tab)
                    for dep_tab in dependent_tabs:
                        self.load_single_tab(dep_tab)
                    
                    messagebox.showinfo("Thành công", f"Đã xóa chỉ tiêu khỏi tab '{self.current_tab}' và {len(dependent_tabs)} tab phụ thuộc")
    
    def check_and_mark_row(self, tree, item, ind_id, so_sanh, tieu_chi, tab_name):
        """
        GIẢI THÍCH: Kiểm tra điều kiện và đánh dấu row
        - Bôi xanh nếu giá trị rỗng (chỉ cho tab CTKT mua sắm khi có tiêu chí)
        - Không kiểm tra so sánh với sản phẩm tham khảo nữa
        """
        tags = tree.item(item)['tags']
        # SỬA MỚI: Nếu là parent thì không bôi xanh
        if "parent" in tags:
            tree.item(item, tags=("data_row",) + tuple(t for t in tags if t != "blue"))
            return
      
        if not so_sanh or not str(so_sanh).strip():
            # GIẢI THÍCH: Nếu giá trị rỗng, bôi xanh
            tree.item(item, tags=("data_row", "blue"))
            return
      
        # GIẢI THÍCH: Đối với tab CTKT mua sắm, kiểm tra cả tiêu chí
        if tab_name == "ctkt_mua_sam":
            if not tieu_chi or not str(tieu_chi).strip():
                tree.item(item, tags=("data_row", "blue"))
            else:
                tree.item(item, tags=("data_row",))
            return
      
        # GIẢI THÍCH: Các tab khác không bôi đỏ
        tree.item(item, tags=("data_row",))

    def calculate_extreme_value(self, ind_id, tree, reference_columns, danh_gia):
        """
        GIẢI THÍCH: Tính giá trị extreme (min/max) từ các sản phẩm tham khảo dựa trên danh_gia
        Trả về chuỗi rỗng nếu không có giá trị hợp lệ hoặc có non-numeric
        """
        if not reference_columns or str(ind_id).startswith('-'):
            return ""
       
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        values = []
       
        for man_id in reference_columns.values():
            value = c.execute("SELECT specification_value FROM product_specifications WHERE manufacturer_id=? AND indicator_id=?",
                            (man_id, ind_id)).fetchone()
            if value and value[0]:
                try:
                    num_value = float(value[0].strip())
                    values.append(num_value)
                except ValueError:
                    conn.close()
                    return "" # Nếu có non-numeric, trả rỗng
       
        conn.close()
       
        if not values:
            return ""
       
        if danh_gia in ["<=", "=", "<"]:
            return str(max(values))
        elif danh_gia in [">=", ">"]:
            return str(min(values))
        else: # "not"
            return ""
    def get_prev_tab_value(self, ind_id, current_tab):
        """
        GIẢI THÍCH: Lấy giá trị từ tab trước đó để kiểm tra ràng buộc
        """
        prev_tabs = {
            "bom": "three_brands",
            "dmkt": "bom",
            "ctkt_bo": "dmkt",
            "ctkt_mua_sam": "three_brands"
        }
        prev_tab = prev_tabs.get(current_tab)
        if prev_tab:
            return self.custom_indicators.get(prev_tab, {}).get(f"so_sanh_{ind_id}", "")
        return None
   
    def get_min_max_from_references(self, ind_id):
        """
        GIẢI THÍCH: Tính giá trị min và max từ các sản phẩm tham khảo cho một indicator cụ thể
        - Duyệt qua tất cả reference products trong self.reference_columns
        - Lấy specification_value từ bảng product_specifications
        - Chỉ lấy các giá trị số hợp lệ (bỏ qua rỗng và non-numeric)
        - Trả về (min, max) hoặc (None, None) nếu không có giá trị hợp lệ
        """
        if not self.reference_columns or str(ind_id).startswith('-'):
            return None, None
       
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        numeric_values = []
       
        try:
            for man_id in self.reference_columns.values():
                value = c.execute("SELECT specification_value FROM product_specifications WHERE manufacturer_id=? AND indicator_id=?",
                                (man_id, ind_id)).fetchone()
                if value and value[0]:
                    try:
                        num_value = float(value[0].strip())
                        numeric_values.append(num_value)
                    except ValueError:
                        continue # Bỏ qua giá trị không phải số
            conn.close()
           
            if numeric_values:
                return min(numeric_values), max(numeric_values)
            else:
                return None, None
               
        except Exception as e:
            conn.close()
            print(f"Lỗi khi tính min/max: {str(e)}")
            return None, None