# detail_project.py
import sqlite3
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from collections import defaultdict
from database import DB_NAME
from add_product import AddProduct
from detail_project_product import DetailProjectProduct
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os
from utils import wrap_text
import re

class DetailProject:
    def __init__(self, root, project_id, callback):
        self.root = root
        self.project_id = project_id
        self.callback = callback
        self.root.title(f"Chi tiết dự án")
        self.root.geometry("1200x600")
        
        # Tùy chỉnh style cho Treeview
        style = ttk.Style()
        style.configure("Treeview", rowheight=30)
        style.configure("Treeview.Heading", font=("Arial", 10, "bold"))
        
        # Load thông tin dự án
        self.load_project_info()
        
        # Treeview frame
        tree_frame = ttk.LabelFrame(root, text="Danh sách thiết bị", padding=5)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Scrollbar
        v_scrollbar = ttk.Scrollbar(tree_frame, orient="vertical")
        v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        h_scrollbar = ttk.Scrollbar(tree_frame, orient="horizontal")
        h_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)
        
        # Treeview với các cột (bỏ cột "Chi tiêu")
        self.tree = ttk.Treeview(
            tree_frame, 
            columns=("STT", "Danh sách thiết bị", "Chủng loại", "Đơn vị", "Số lượng", "Ghi chú", "Hành động"), 
            show="headings", 
            yscrollcommand=v_scrollbar.set,
            xscrollcommand=h_scrollbar.set
        )
        v_scrollbar.config(command=self.tree.yview)
        h_scrollbar.config(command=self.tree.xview)
        
        # Cấu hình các cột
        self.tree.heading("STT", text="STT")
        self.tree.heading("Danh sách thiết bị", text="Danh sách thiết bị")
        self.tree.heading("Chủng loại", text="Chủng loại")
        self.tree.heading("Đơn vị", text="Đơn vị")
        self.tree.heading("Số lượng", text="Số lượng")
        self.tree.heading("Ghi chú", text="Ghi chú")
        self.tree.heading("Hành động", text="Hành động")
        
        self.tree.column("STT", width=50, anchor="center")
        self.tree.column("Danh sách thiết bị", width=250, anchor="w")
        self.tree.column("Chủng loại", width=180, anchor="center")
        self.tree.column("Đơn vị", width=80, anchor="center")
        self.tree.column("Số lượng", width=100, anchor="center")
        self.tree.column("Ghi chú", width=200, anchor="w")
        self.tree.column("Hành động", width=100, anchor="center")
        
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # Bind double-click và click hành động
        self.tree.bind("<Double-1>", self.on_product_selected)
        self.tree.bind("<Button-1>", self.on_action_click)
        
        # Button frame
        button_frame = tk.Frame(root)
        button_frame.pack(fill=tk.X, pady=10, padx=10)
        
        tk.Button(button_frame, text="Thêm sản phẩm", command=self.add_product_to_project, 
                  font=("Arial", 12)).pack(side=tk.LEFT, padx=5)
        tk.Button(button_frame, text="Export files", command=self.export_files, 
                  font=("Arial", 12)).pack(side=tk.LEFT, padx=5)
        tk.Button(button_frame, text="Quay lại", command=root.destroy, 
                  font=("Arial", 12)).pack(side=tk.RIGHT, padx=5)
        
        # Load sản phẩm
        self.load_products()

        # Định nghĩa các fill và border
        self.yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        self.blue_fill = PatternFill(start_color="00A6FF", end_color="00A6FF", fill_type="solid")
        self.thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                  top=Side(style='thin'), bottom=Side(style='thin'))
        # Lưu kiểu fill của product-row template (sẽ được set khi xuất sản phẩm đầu)
        self.template_product_fill = None
    
    def load_project_info(self):
        """Hiển thị thông tin dự án."""
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        project = c.execute("SELECT name, ma_du_an, ghi_chu FROM projects WHERE id=?", 
                            (self.project_id,)).fetchone()
        if project:
            self.project_name = project[0]
            info_frame = tk.Frame(self.root, bg="#f0f0f0", pady=10)
            info_frame.pack(fill=tk.X, padx=10)
            
            info_label = tk.Label(
                info_frame, 
                text=f"Dự án: {project[0]}", 
                font=("Arial", 14, "bold"),
                bg="#f0f0f0"
            )
            info_label.pack()
        conn.close()
    
    def load_products(self):
        """
        Tải danh sách sản phẩm trong dự án.
        
        Lấy thông tin:
        - Tên sản phẩm, unit, quantity, note từ bảng products
        - Loại sản phẩm từ bảng product_types thông qua product_type_mapping_products
        """
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        
        # Query lấy sản phẩm và loại sản phẩm
        query = '''
            SELECT 
                p.id,
                p.name,
                p.ma_san_pham,
                p.note,
                p.unit,
                p.quantity,
                GROUP_CONCAT(pt.name, ', ') as product_types
            FROM product_projects pp 
            JOIN products p ON pp.product_id = p.id 
            LEFT JOIN product_type_mapping_products ptmp ON p.id = ptmp.product_id
            LEFT JOIN product_types pt ON ptmp.type_id = pt.id
            WHERE pp.project_id=? 
            GROUP BY p.id, p.name, p.ma_san_pham, p.note, p.unit, p.quantity
            ORDER BY p.name
        '''
        
        products = c.execute(query, (self.project_id,)).fetchall()
        self.tree.delete(*self.tree.get_children())
        
        for index, (prod_id, name, ma_san_pham, note, unit, quantity, product_types) in enumerate(products, 1):
            self.tree.insert(
                "", 
                "end", 
                iid=str(prod_id), 
                values=(
                    index,
                    name or "Không có tên",
                    product_types or "",
                    unit or "Bộ",
                    quantity if quantity is not None else 1,
                    note or "",
                    "Xóa"
                )
            )
        
        conn.close()
    
    def on_product_selected(self, event):
        """Double-click: Mở giao diện DetailProjectProduct."""
        column = self.tree.identify_column(event.x)
        if column == "#7":  # Cột Hành động (đã thay đổi thứ tự cột)
            return
        
        selected = self.tree.selection()
        if not selected:
            return
        prod_id = int(selected[0])
        
        detail_win = tk.Toplevel(self.root)
        DetailProjectProduct(detail_win, self.project_id, prod_id, self.refresh_products)
    
    def on_action_click(self, event):
        """Click cột Hành động: Xóa sản phẩm khỏi dự án."""
        item = self.tree.identify_row(event.y)
        column = self.tree.identify_column(event.x)
        
        if column == "#7" and item:  # Cột Hành động (đã thay đổi thứ tự cột)
            values = self.tree.item(item)['values']
            product_name = values[1] if len(values) > 1 else "sản phẩm này"
            
            if messagebox.askyesno("Xác nhận", f"Bạn có chắc muốn xóa '{product_name}' khỏi dự án?"):
                prod_id = int(item)
                
                conn = sqlite3.connect(DB_NAME)
                c = conn.cursor()
                c.execute("DELETE FROM product_projects WHERE project_id=? AND product_id=?", 
                          (self.project_id, prod_id))
                conn.commit()
                conn.close()
                
                messagebox.showinfo("Thành công", "Đã xóa sản phẩm khỏi dự án!")
                self.load_products()
                self.callback()
    
    def add_product_to_project(self):
        """Thêm sản phẩm vào dự án."""
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        
        # Lấy danh sách sản phẩm chưa có trong dự án
        query = '''
            SELECT 
                p.id, 
                p.ma_san_pham, 
                p.name,
                GROUP_CONCAT(pt.name, ', ') as product_types
            FROM products p
            LEFT JOIN product_type_mapping_products ptmp ON p.id = ptmp.product_id
            LEFT JOIN product_types pt ON ptmp.type_id = pt.id
            WHERE p.id NOT IN (
                SELECT product_id 
                FROM product_projects 
                WHERE project_id=?
            )
            GROUP BY p.id, p.ma_san_pham, p.name
            ORDER BY p.name
        '''
        
        products = c.execute(query, (self.project_id,)).fetchall()
        conn.close()
        
        if not products:
            response = messagebox.askyesno("Thông báo", "Không có sản phẩm mới để thêm!\nBạn có muốn thêm mới sản phẩm không?")
            if response:  # Nếu người dùng chọn "Yes"
                new_win = tk.Toplevel(self.root)
                AddProduct(new_win, product_id=None, parent=self)
                
                def refresh_after_add():
                    """Callback để refresh danh sách sản phẩm sau khi thêm."""
                    self.load_products()
                    self.callback()
                    new_win.destroy()
                
                new_win.protocol("WM_DELETE_WINDOW", refresh_after_add)
            return
        
        # Tạo dialog chọn sản phẩm
        dialog = tk.Toplevel(self.root)
        dialog.title("Thêm sản phẩm vào dự án")
        dialog.geometry("720x540")
        dialog.transient(self.root)
        dialog.grab_set()
        
        frame = tk.Frame(dialog)
        frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Thêm ô tìm kiếm
        search_frame = tk.Frame(frame)
        search_frame.pack(fill=tk.X, pady=(0, 10))
        
        tk.Label(search_frame, text="Tìm kiếm:", font=("Arial", 11)).pack(side=tk.LEFT)
        search_entry = tk.Text(search_frame, height=3, width=50, font=("Arial", 10))
        search_entry.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        
        tk.Label(frame, text="Chọn sản phẩm để thêm:", font=("Arial", 12, "bold")).pack(pady=(0, 10))
        
        # Treeview để hiển thị sản phẩm
        tree_frame = tk.Frame(frame)
        tree_frame.pack(fill=tk.BOTH, expand=True)
        
        scrollbar = ttk.Scrollbar(tree_frame, orient="vertical")
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        product_tree = ttk.Treeview(
            tree_frame,
            columns=("Mã SP", "Tên sản phẩm", "Chủng loại"),
            show="headings",
            yscrollcommand=scrollbar.set,
            selectmode="browse"
        )
        scrollbar.config(command=product_tree.yview)
        
        product_tree.heading("Mã SP", text="Mã SP")
        product_tree.heading("Tên sản phẩm", text="Tên sản phẩm")
        product_tree.heading("Chủng loại", text="Chủng loại")
        
        product_tree.column("Mã SP", width=100, anchor="center")
        product_tree.column("Tên sản phẩm", width=250, anchor="w")
        product_tree.column("Chủng loại", width=200, anchor="w")
        
        product_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # Thêm dữ liệu vào tree
        def update_product_tree(search_text=""):
            product_tree.delete(*product_tree.get_children())
            for prod_id, ma_san_pham, name, product_types in products:
                if (search_text.lower() in (ma_san_pham or "").lower() or 
                    search_text.lower() in (name or "").lower()):
                    product_tree.insert(
                        "",
                        "end",
                        iid=str(prod_id),
                        values=(ma_san_pham or "N/A", name or "Không có tên", product_types or "")
                    )
        
        update_product_tree()
        
        # Binding sự kiện nhập liệu
        search_entry.bind("<KeyRelease>", lambda event: update_product_tree(search_entry.get("1.0", tk.END).strip()))
        
        def add_selected():
            """Thêm sản phẩm đã chọn vào dự án."""
            selected = product_tree.selection()
            if not selected:
                messagebox.showwarning("Cảnh báo", "Vui lòng chọn một sản phẩm!")
                return
            
            prod_id = int(selected[0])
            
            conn = sqlite3.connect(DB_NAME)
            c = conn.cursor()
            try:
                c.execute("INSERT INTO product_projects (project_id, product_id) VALUES (?, ?)", 
                        (self.project_id, prod_id))
                conn.commit()
                messagebox.showinfo("Thành công", "Đã thêm sản phẩm vào dự án!")
                dialog.destroy()
                self.load_products()
                self.callback()
            except sqlite3.IntegrityError:
                messagebox.showerror("Lỗi", "Sản phẩm đã tồn tại trong dự án!")
            finally:
                conn.close()
        
        def add_new_product():
            """Mở giao diện thêm sản phẩm mới từ add_product.py."""
            new_win = tk.Toplevel(dialog)
            AddProduct(new_win, product_id=None, parent=self)
            
            def refresh_after_add():
                conn = sqlite3.connect(DB_NAME)
                c = conn.cursor()
                nonlocal products
                products = c.execute(query, (self.project_id,)).fetchall()
                conn.close()
                update_product_tree(search_entry.get("1.0", tk.END).strip())
                new_win.destroy()
        
            new_win.protocol("WM_DELETE_WINDOW", refresh_after_add)
        
        button_frame = tk.Frame(frame)
        button_frame.pack(pady=(10, 0))
        
        tk.Button(button_frame, text="Thêm mới sản phẩm", command=add_new_product, font=("Arial", 11), width=18).pack(side=tk.LEFT, padx=5)
        tk.Button(button_frame, text="Thêm", command=add_selected, font=("Arial", 11), width=10).pack(side=tk.LEFT, padx=5)
        tk.Button(button_frame, text="Hủy", command=dialog.destroy, font=("Arial", 11), width=10).pack(side=tk.LEFT, padx=5)

    def refresh_products(self):
        """Callback để refresh danh sách sản phẩm."""
        self.load_products()
        self.callback()

    def is_numeric_value(self, value):
        try:
            float(value.strip())
            return True
        except (ValueError, AttributeError):
            return False

    def load_reference_products(self, product_id):
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        references = c.execute('''SELECT m.id, m.name, m.product_name 
                                FROM reference_products rp 
                                JOIN manufacturers m ON rp.manufacturer_id = m.id 
                                WHERE rp.product_id=? ORDER BY rp.sort_order''', (product_id,)).fetchall()
        conn.close()
        return [(r[0], r[1], r[2]) for r in references]

    def get_type_id(self, product_type):
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        type_id = c.execute("SELECT id FROM product_types WHERE name=?", (product_type,)).fetchone()
        conn.close()
        return type_id[0] if type_id else None

    def load_hidden_indicators(self, product_id):
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        hidden_records = c.execute("SELECT tab_name, indicator_id FROM product_hidden_indicators WHERE product_id=?", 
                                (product_id,)).fetchall()
        conn.close()
        
        origin_deleted = defaultdict(set)
        deleted_indicators = defaultdict(set)
        cascade_map = {
            "three_brands": ["bom", "dmkt", "ctkt_bo", "ctkt_mua_sam"],
            "bom": ["dmkt", "ctkt_bo"],
            "dmkt": ["ctkt_bo"],
            "ctkt_mua_sam": [],
            "ctkt_bo": []
        }
        
        for tab_name, ind_id in hidden_records:
            if tab_name in cascade_map:
                origin_deleted[tab_name].add(ind_id)
                deleted_indicators[tab_name].add(ind_id)
                for dep_tab in cascade_map[tab_name]:
                    deleted_indicators[dep_tab].add(ind_id)
        
        return deleted_indicators, origin_deleted

    def load_custom_indicators(self, product_id):
        """
        GIẢI THÍCH: Load các giá trị custom từ database
        - Trả về dictionary custom_indicators chứa các giá trị tùy chỉnh
        - Trả về list custom_rows_ctkt_ms chứa các hàng "Yêu cầu khác" trong CTKT mua sắm
        """
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        all_custom = c.execute("SELECT tab_name, indicator_id, custom_value FROM product_custom_indicators WHERE product_id=?", 
                                (product_id,)).fetchall()
        conn.close()
        
        custom_indicators = {tab: {} for tab in ["three_brands", "bom", "dmkt", "ctkt_bo", "ctkt_mua_sam"]}
        
        for tab_name, indicator_id_composite, custom_value in all_custom:
            try:
                if isinstance(indicator_id_composite, str) and '_' in str(indicator_id_composite):
                    parts = str(indicator_id_composite).rsplit('_', 1)
                    if len(parts) == 2:
                        ind_id_str, prefix = parts
                        if custom_value.startswith(f"{prefix}_"):
                            actual_value = custom_value[len(f"{prefix}_"):]
                            key = f"{prefix}_{ind_id_str}"
                            custom_indicators[tab_name][key] = actual_value
            except Exception as e:
                print(f"Lỗi khi xử lý {indicator_id_composite}: {str(e)}")
                continue
        
        # QUAN TRỌNG: Tạo danh sách custom_rows_ctkt_ms từ custom_indicators
        custom_rows_ctkt_ms = []
        for key in custom_indicators.get("ctkt_mua_sam", {}).keys():
            if key.startswith("chi_tieu_"):
                # Lấy custom_id từ key (ví dụ: "chi_tieu_custom_123" -> "custom_123")
                parts = key.split("_")
                if len(parts) >= 3:
                    custom_id = "_".join(parts[2:])  # Lấy phần sau "chi_tieu_"
                    
                    # Kiểm tra xem đã có custom_id này chưa
                    if not any(row["id"] == custom_id for row in custom_rows_ctkt_ms):
                        row_data = {
                            "id": custom_id,
                            "chi_tieu": custom_indicators["ctkt_mua_sam"].get(f"chi_tieu_{custom_id}", ""),
                            "yeu_cau": custom_indicators["ctkt_mua_sam"].get(f"yeu_cau_{custom_id}", ""),
                            "so_sanh": custom_indicators["ctkt_mua_sam"].get(f"so_sanh_{custom_id}", ""),
                            "don_vi": custom_indicators["ctkt_mua_sam"].get(f"don_vi_{custom_id}", ""),
                            "tieu_chi": custom_indicators["ctkt_mua_sam"].get(f"tieu_chi_{custom_id}", ""),
                            "crit_type": custom_indicators["ctkt_mua_sam"].get(f"crit_type_{custom_id}", "CTCB")
                        }
                        custom_rows_ctkt_ms.append(row_data)
        
        return custom_indicators, custom_rows_ctkt_ms

    def load_indicators(self, type_id, tab_name, product_id):
        """
        GIẢI THÍCH: Load indicators với ORDER BY để "Chủng loại" lên đầu
        - SỬA: Thêm CASE WHEN để sắp xếp "Chủng loại" lên đầu tiên
        """
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        deleted_indicators, _ = self.load_hidden_indicators(product_id)
        hidden_ids = deleted_indicators[tab_name]
        placeholders = ','.join('?' * len(hidden_ids)) if hidden_ids else '0'
        
        # SỬA: ORDER BY với CASE WHEN để "Chủng loại" lên đầu
        query = f"""
            SELECT id, requirement, indicator, unit, value 
            FROM indicators 
            WHERE type_id = ? AND id NOT IN ({placeholders}) 
        """
        
        indicators = c.execute(query, (type_id,) + tuple(hidden_ids)).fetchall()
        conn.close()
        return indicators

    def calculate_extreme_value(self, ind_id, man_ids, danh_gia):
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        values = []
        for man_id in man_ids:
            value = c.execute("SELECT specification_value FROM product_specifications WHERE manufacturer_id=? AND indicator_id=?", 
                            (man_id, ind_id)).fetchone()
            if value and value[0]:
                try:
                    num_value = float(value[0].strip())
                    values.append(num_value)
                except ValueError:
                    conn.close()
                    return ""  # Non-numeric -> rỗng
        conn.close()
        if not values:
            return ""
        if danh_gia in ["<=", "=", "<"]:
            return str(max(values))
        elif danh_gia in [">=", ">"]:
            return str(min(values))
        else:
            return ""

    # Sửa load_three_brands_data: Thêm danh_gia vào row, dùng calculate_extreme_value
    def load_three_brands_data(self, product_id, product_type, reference_products, deleted_indicators, custom_indicators):
        """
        GIẢI THÍCH: Load tab 3 hãng với xử lý an toàn cho indicator=None
        - Đã có kiểm tra indc, chỉ cần đảm bảo ORDER BY đúng
        """
        type_id = self.get_type_id(product_type)
        indicators = self.load_indicators(type_id, "three_brands", product_id)
        data_rows = []
        man_ids = [mid for mid, _, _ in reference_products]
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        for index, ind in enumerate(indicators, 1):
            ind_id, req, indc, unit, value = ind
            row = [index, wrap_text(req, 50), wrap_text(indc if indc else "", 50)]  # SỬA: Đã xử lý None
            danh_gia = custom_indicators.get("three_brands", {}).get(f"danh_gia_{ind_id}", "not")
            so_sanh = custom_indicators.get("three_brands", {}).get(f"so_sanh_{ind_id}", "")
            if not so_sanh:
                so_sanh = self.calculate_extreme_value(ind_id, man_ids, danh_gia)
                if so_sanh:
                    custom_indicators.setdefault("three_brands", {})[f"so_sanh_{ind_id}"] = so_sanh
            row.append(danh_gia)
            row.append(wrap_text(so_sanh, 20))
            row.append(wrap_text(value or "", 20))
            row.append(wrap_text(unit or "", 20))
            crit_type = custom_indicators.get("three_brands", {}).get(f"crit_type_{ind_id}", "CTCB")
            row.append(crit_type)
            for man_id, _, _ in reference_products:
                val = c.execute("SELECT specification_value FROM product_specifications WHERE manufacturer_id=? AND indicator_id=?", 
                                (man_id, ind_id)).fetchone()
                row.append(wrap_text(val[0] if val else "", 50))
                ref_key = f"ref_value_{man_id}_{ind_id}"
                ref_value = custom_indicators.get("three_brands", {}).get(ref_key, "")
                row.append(wrap_text(ref_value, 20))
            row.append(ind_id)
            data_rows.append(row)
        conn.close()
        return data_rows

    # Sửa tương tự cho load_bom_data
    def load_bom_data(self, product_id, product_type, reference_products, deleted_indicators, custom_indicators):
        """
        GIẢI THÍCH: Load tab BOM với xử lý an toàn cho indicator=None
        - Đã có kiểm tra indc, giữ nguyên logic
        """
        type_id = self.get_type_id(product_type)
        indicators = self.load_indicators(type_id, "bom", product_id)
        data_rows = []
        man_ids = [mid for mid, _, _ in reference_products]
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        for index, ind in enumerate(indicators, 1):
            ind_id, req, indc, unit, _ = ind
            row = [index, wrap_text(req, 50), wrap_text(indc if indc else "", 50)]  # SỬA: Đã xử lý None
            danh_gia = custom_indicators.get("three_brands", {}).get(f"danh_gia_{ind_id}", "not")
            so_sanh = custom_indicators.get("bom", {}).get(f"so_sanh_{ind_id}", "")
            if not so_sanh:
                so_sanh = self.calculate_extreme_value(ind_id, man_ids, danh_gia)
                if so_sanh:
                    custom_indicators.setdefault("bom", {})[f"so_sanh_{ind_id}"] = so_sanh
            row.append(danh_gia)
            row.append(wrap_text(so_sanh, 20))
            row.append(wrap_text(unit or "", 20))
            crit_type = custom_indicators.get("bom", {}).get(f"crit_type_{ind_id}", "CTCB")
            row.append(crit_type)
            for man_id, _, _ in reference_products:
                val = c.execute("SELECT specification_value FROM product_specifications WHERE manufacturer_id=? AND indicator_id=?", 
                                (man_id, ind_id)).fetchone()
                row.append(wrap_text(val[0] if val else "", 50))
                ref_key = f"ref_value_{man_id}_{ind_id}"
                ref_value = custom_indicators.get("three_brands", {}).get(ref_key, "")
                row.append(wrap_text(ref_value, 20))
            row.append(ind_id)
            data_rows.append(row)
        conn.close()
        return data_rows

    def load_dmkt_data(self, product_id, product_type, deleted_indicators, custom_indicators):
        type_id = self.get_type_id(product_type)
        indicators = self.load_indicators(type_id, "dmkt", product_id)
        data_rows = []
        groups = defaultdict(list)
        for ind in indicators:
            groups[ind[1]].append(ind)
        tt = 1
        for req, ind_list in groups.items():
            has_indicator = any(ind[2] and ind[2].strip() for ind in ind_list)
            is_single_no_indicator = (len(ind_list) == 1 and (not ind_list[0][2] or not ind_list[0][2].strip()))

            if is_single_no_indicator:
                ind_id, req, indc, unit, _ = ind_list[0]
                unit = unit or ""  # thêm dòng này cho chắc

                # === SỬA TẠI ĐÂY: ép None → "" ===
                so_sanh = (custom_indicators.get("dmkt", {}).get(f"so_sanh_{ind_id}", "") or "")
                if not so_sanh:
                    bom_value = (custom_indicators.get("bom", {}).get(f"so_sanh_{ind_id}", "") or "")
                    if not bom_value:
                        hang_value = (custom_indicators.get("three_brands", {}).get(f"so_sanh_{ind_id}", "") or "")
                        if not hang_value:
                            danh_gia = (custom_indicators.get("three_brands", {}).get(f"danh_gia_{ind_id}", "not") or "not")
                            man_ids = [mid for mid, _, _ in self.load_reference_products(product_id)]
                            hang_value = (self.calculate_extreme_value(ind_id, man_ids, danh_gia) or "")
                        bom_value = hang_value or ""
                    so_sanh = bom_value or ""

                danh_gia = (custom_indicators.get("three_brands", {}).get(f"danh_gia_{ind_id}", "not") or "not")
                # === Kết thúc sửa ===

                compare_symbols = {"<=": "≤", ">=": "≥", "=": "=", "<": "<", ">": ">"}
                compare_symbol = compare_symbols.get(danh_gia, "") if danh_gia != "not" else ""
                is_numeric = self.is_numeric_value(so_sanh)
                display_value = f"- {req} {compare_symbol} {so_sanh} {unit}".strip() if is_numeric and compare_symbol else f"- {req} {so_sanh} {unit}".strip()

                row = [tt, display_value, wrap_text(so_sanh, 20), wrap_text(unit or "", 20), "Xóa", ind_id, danh_gia]
                data_rows.append(row)
                tt += 1
            else:
                row = [tt, f"- {req}", "", "", "Xóa", None]
                data_rows.append(row)
                tt += 1
                sub_tt = 1
                for sub_ind in ind_list:
                    if not sub_ind[2] or not sub_ind[2].strip():
                        continue
                    ind_id, req, indc, unit, _ = sub_ind
                    unit = unit or ""

                    # === SỬA TƯƠNG TỰ CHO SUB ===
                    so_sanh = (custom_indicators.get("dmkt", {}).get(f"so_sanh_{ind_id}", "") or "")
                    if not so_sanh:
                        bom_value = (custom_indicators.get("bom", {}).get(f"so_sanh_{ind_id}", "") or "")
                        if not bom_value:
                            hang_value = (custom_indicators.get("three_brands", {}).get(f"so_sanh_{ind_id}", "") or "")
                            if not hang_value:
                                danh_gia = (custom_indicators.get("three_brands", {}).get(f"danh_gia_{ind_id}", "not") or "not")
                                man_ids = [mid for mid, _, _ in self.load_reference_products(product_id)]
                                hang_value = (self.calculate_extreme_value(ind_id, man_ids, danh_gia) or "")
                            bom_value = hang_value or ""
                        so_sanh = bom_value or ""

                    danh_gia = (custom_indicators.get("three_brands", {}).get(f"danh_gia_{ind_id}", "not") or "not")
                    # === Kết thúc sửa sub ===

                    row = [f"{tt-1}.{sub_tt}", f"+ {indc}", wrap_text(so_sanh, 20), wrap_text(unit or "", 20), "Xóa", ind_id, danh_gia]
                    data_rows.append(row)
                    sub_tt += 1
        return data_rows

    def load_ctkt_bo_data(self, product_id, product_type, deleted_indicators, custom_indicators):
        """
        GIẢI THÍCH: Load tab CTKT bộ với xử lý an toàn cho indicator=None
        - SỬA: Thêm kiểm tra None trước khi strip()
        """
        type_id = self.get_type_id(product_type)
        indicators = self.load_indicators(type_id, "ctkt_bo", product_id)
        data_rows = []
        groups = defaultdict(list)
        for ind in indicators:
            groups[ind[1]].append(ind)
        tt = 1
        for req, ind_list in groups.items():
            if not ind_list:
                continue
            
            # SỬA: Kiểm tra None an toàn
            has_indicator = any(ind[2] and ind[2].strip() for ind in ind_list)
            is_single_no_indicator = (len(ind_list) == 1 and (not ind_list[0][2] or not ind_list[0][2].strip()))
            
            if is_single_no_indicator:
                ind_id, req, indc, unit, _ = ind_list[0]
                gia_tri = custom_indicators.get("ctkt_bo", {}).get(f"gia_tri_{ind_id}", "")
                if not gia_tri:
                    dmkt_value = custom_indicators.get("dmkt", {}).get(f"so_sanh_{ind_id}", "")
                    if not dmkt_value:
                        bom_value = custom_indicators.get("bom", {}).get(f"so_sanh_{ind_id}", "")
                        if not bom_value:
                            hang_value = custom_indicators.get("three_brands", {}).get(f"so_sanh_{ind_id}", "")
                            if not hang_value:
                                danh_gia = custom_indicators.get("three_brands", {}).get(f"danh_gia_{ind_id}", "not")
                                man_ids = [mid for mid, _, _ in self.load_reference_products(product_id)]
                                hang_value = self.calculate_extreme_value(ind_id, man_ids, danh_gia)
                            bom_value = hang_value
                        dmkt_value = bom_value
                    gia_tri = dmkt_value
                danh_gia = custom_indicators.get("three_brands", {}).get(f"danh_gia_{ind_id}", "not")
                display_indc = req
                try:
                    num_value = float(gia_tri)
                    if num_value.is_integer():
                        if danh_gia == ">=":
                            display_indc += ", không nhỏ hơn"
                        elif danh_gia == "<=":
                            display_indc += ", không lớn hơn"
                        elif danh_gia == "=":
                            display_indc += ", bằng"
                        elif danh_gia == ">":
                            display_indc += ", lớn hơn"
                        elif danh_gia == "<":
                            display_indc += ", nhỏ hơn"
                except (ValueError, TypeError):
                    pass
                row = [tt, wrap_text(display_indc, 50), wrap_text(unit or "-", 20), wrap_text(gia_tri, 50), "Xóa", ind_id]
                data_rows.append(row)
                tt += 1
            else:
                row = [tt, wrap_text(req, 50), "", "", "Xóa", None]
                data_rows.append(row)
                tt += 1
                sub_tt = 1
                for sub_ind in ind_list:
                    # SỬA: Kiểm tra None an toàn
                    if not sub_ind[2] or not sub_ind[2].strip():
                        continue
                    ind_id, req, indc, unit, _ = sub_ind
                    gia_tri = custom_indicators.get("ctkt_bo", {}).get(f"gia_tri_{ind_id}", "")
                    if not gia_tri:
                        dmkt_value = custom_indicators.get("dmkt", {}).get(f"so_sanh_{ind_id}", "")
                        if not dmkt_value:
                            bom_value = custom_indicators.get("bom", {}).get(f"so_sanh_{ind_id}", "")
                            if not bom_value:
                                hang_value = custom_indicators.get("three_brands", {}).get(f"so_sanh_{ind_id}", "")
                                if not hang_value:
                                    danh_gia = custom_indicators.get("three_brands", {}).get(f"danh_gia_{ind_id}", "not")
                                    man_ids = [mid for mid, _, _ in self.load_reference_products(product_id)]
                                    hang_value = self.calculate_extreme_value(ind_id, man_ids, danh_gia)
                                bom_value = hang_value
                            dmkt_value = bom_value
                        gia_tri = dmkt_value
                    danh_gia = custom_indicators.get("three_brands", {}).get(f"danh_gia_{ind_id}", "not")
                    display_indc = indc
                    try:
                        num_value = float(gia_tri)
                        if num_value.is_integer():
                            if danh_gia == ">=":
                                display_indc += ", không nhỏ hơn"
                            elif danh_gia == "<=":
                                display_indc += ", không lớn hơn"
                            elif danh_gia == "=":
                                display_indc += ", bằng"
                            elif danh_gia == ">":
                                display_indc += ", lớn hơn"
                            elif danh_gia == "<":
                                display_indc += ", nhỏ hơn"
                    except (ValueError, TypeError):
                        pass
                    row = [f"{tt-1}.{sub_tt}", wrap_text(display_indc, 50), wrap_text(unit or "-", 20), wrap_text(gia_tri, 50), "Xóa", ind_id]
                    data_rows.append(row)
                    sub_tt += 1
        return data_rows

    # Sửa load_ctkt_mua_sam_data: Sử dụng danh_gia cho tiêu chí đánh giá
    def load_ctkt_mua_sam_data(self, product_id, product_type, deleted_indicators, custom_indicators, custom_rows_ctkt_ms):
        type_id = self.get_type_id(product_type)
        indicators = self.load_indicators(type_id, "ctkt_mua_sam", product_id)
        data_rows = []
        groups = defaultdict(list)
        reference_products = self.load_reference_products(product_id)
        man_ids = [mid for mid, _, _ in reference_products]
        for ind in indicators:
            groups[ind[1]].append(ind)
        stt = 1
        for req, ind_list in groups.items():
            first_ind = ind_list[0]
            ind_id = first_ind[0]
            unit = first_ind[3] or ""
            danh_gia = custom_indicators.get("three_brands", {}).get(f"danh_gia_{ind_id}", "not")
            so_sanh = custom_indicators.get("ctkt_mua_sam", {}).get(f"so_sanh_{ind_id}", "")
            if not so_sanh:
                hang_value = custom_indicators.get("three_brands", {}).get(f"so_sanh_{ind_id}", "")
                if not hang_value:
                    hang_value = self.calculate_extreme_value(ind_id, man_ids, danh_gia)
                so_sanh = hang_value
            yeu_cau_str = first_ind[2] or ""
            dat_line, khong_dat_line = self.get_dat_khong_dat_lines(yeu_cau_str, so_sanh, unit, danh_gia)
            tieu_chi_display = f"{dat_line}\r\n{khong_dat_line}"
            crit_type = custom_indicators.get("ctkt_mua_sam", {}).get(f"crit_type_{ind_id}", "CTCB")
            row = [stt, wrap_text(req, 40), wrap_text(yeu_cau_str, 30), wrap_text(so_sanh, 20), 
                wrap_text(unit, 15), tieu_chi_display, crit_type, "Xóa", ind_id]
            data_rows.append(row)
            stt += 1
            for sub_ind in ind_list[1:]:
                ind_id = sub_ind[0]
                unit = sub_ind[3] or ""
                danh_gia = custom_indicators.get("three_brands", {}).get(f"danh_gia_{ind_id}", "not")
                so_sanh = custom_indicators.get("ctkt_mua_sam", {}).get(f"so_sanh_{ind_id}", "")
                if not so_sanh:
                    hang_value = custom_indicators.get("three_brands", {}).get(f"so_sanh_{ind_id}", "")
                    if not hang_value:
                        hang_value = self.calculate_extreme_value(ind_id, man_ids, danh_gia)
                    so_sanh = hang_value
                yeu_cau_str = sub_ind[2] or ""
                dat_line, khong_dat_line = self.get_dat_khong_dat_lines(yeu_cau_str, so_sanh, unit, danh_gia)
                tieu_chi_display = f"{dat_line}\r\n{khong_dat_line}"
                crit_type = custom_indicators.get("ctkt_mua_sam", {}).get(f"crit_type_{ind_id}", "CTCB")
                row = [stt, "", wrap_text(yeu_cau_str, 30), wrap_text(so_sanh, 20), 
                    wrap_text(unit, 15), tieu_chi_display, crit_type, "Xóa", ind_id]
                data_rows.append(row)
                stt += 1
        data_rows.append([stt, "Yêu cầu khác", "", "", "", "", "", "", None])
        for custom_row in custom_rows_ctkt_ms:
            custom_id = custom_row["id"]
            chi_tieu = custom_row["chi_tieu"]
            yeu_cau = custom_row["yeu_cau"]
            so_sanh = custom_row["so_sanh"]
            don_vi = custom_row["don_vi"]
            tieu_chi = custom_row["tieu_chi"].replace('\n', '\r\n')
            crit_type = custom_row["crit_type"]
            row = [stt + 1, wrap_text(chi_tieu, 40), wrap_text(yeu_cau, 30), wrap_text(so_sanh, 20), 
                wrap_text(don_vi, 15), tieu_chi, crit_type, "Xóa", custom_id]
            data_rows.append(row)
            stt += 1
        return data_rows

    # Thêm hàm get_dat_khong_dat_lines (từ code tham khảo)
    def get_dat_khong_dat_lines(self, yeu_cau_str, so_sanh, unit, danh_gia):
        unit_str = f" {unit}" if unit else ""
        is_numeric = re.match(r'^-?\d+(\.\d+)?$', str(so_sanh).strip())
        if is_numeric and yeu_cau_str:
            if danh_gia == "<=":
                dat_line = f"- Đạt: {yeu_cau_str} ≤ {so_sanh}{unit_str}"
                khong_dat_line = f"- Không đạt: {yeu_cau_str} > {so_sanh}{unit_str}"
            elif danh_gia == ">=":
                dat_line = f"- Đạt: {yeu_cau_str} ≥ {so_sanh}{unit_str}"
                khong_dat_line = f"- Không đạt: {yeu_cau_str} < {so_sanh}{unit_str}"
            elif danh_gia == "=":
                dat_line = f"- Đạt: {yeu_cau_str} = {so_sanh}{unit_str}"
                khong_dat_line = f"- Không đạt: {yeu_cau_str} ≠ {so_sanh}{unit_str}"
            elif danh_gia == "<":
                dat_line = f"- Đạt: {yeu_cau_str} < {so_sanh}{unit_str}"
                khong_dat_line = f"- Không đạt: {yeu_cau_str} ≥ {so_sanh}{unit_str}"
            elif danh_gia == ">":
                dat_line = f"- Đạt: {yeu_cau_str} > {so_sanh}{unit_str}"
                khong_dat_line = f"- Không đạt: {yeu_cau_str} ≤ {so_sanh}{unit_str}"
            else:
                dat_line = f"- Đạt: hàng hóa nhà thầu chào đáp ứng yêu cầu kỹ thuật"
                khong_dat_line = f"- Không đạt: hàng hóa nhà thầu chào không đáp ứng yêu cầu kỹ thuật"
        else:
            dat_line = f"- Đạt: hàng hóa nhà thầu chào đáp ứng yêu cầu kỹ thuật"
            khong_dat_line = f"- Không đạt: hàng hóa nhà thầu chào không đáp ứng yêu cầu kỹ thuật"
        return dat_line, khong_dat_line
    
    def export_hang_bom(self, ws, current_row, product_name, reference_products, data_rows, tab_name, is_first, product_index, max_num_products, custom_indicators):
        if not ws or not product_name or not data_rows or not isinstance(data_rows, list):
            return current_row

        num_products = len(reference_products) if reference_products and isinstance(reference_products, list) else 0
        
        def int_to_roman(num):
            vals = [
                (1000, "M"), (900, "CM"), (500, "D"), (400, "CD"),
                (100, "C"), (90, "XC"), (50, "L"), (40, "XL"),
                (10, "X"), (9, "IX"), (5, "V"), (4, "IV"), (1, "I")
            ]
            res = ""
            n = num
            for v, r in vals:
                while n >= v:
                    res += r
                    n -= v
            return res

        product_row = current_row
        manufacturer_row = current_row + 1
        detail_row = current_row + 2

        try:
            if is_first:
                # Lưu template fills
                self.template_product_fill = ws.cell(row=4, column=2).fill.copy() if ws.cell(row=4, column=2).fill else PatternFill(fill_type=None)
                self.template_manufacturer_fill = ws.cell(row=5, column=2).fill.copy() if ws.cell(row=5, column=2).fill else PatternFill(fill_type=None)
                self.template_detail_fill = ws.cell(row=6, column=2).fill.copy() if ws.cell(row=6, column=2).fill else PatternFill(fill_type=None)
                
                # XÓA MERGE CŨ Ở HÀNG 1, 2, 3 (nếu có từ template)
                merged_ranges_to_unmerge = []
                for merged_range in ws.merged_cells.ranges:
                    if merged_range.min_row <= 3 and merged_range.max_row <= 3 and merged_range.min_col >= 5:
                        merged_ranges_to_unmerge.append(str(merged_range))
                
                for merged_range in merged_ranges_to_unmerge:
                    try:
                        ws.unmerge_cells(merged_range)
                    except:
                        pass
                
                if max_num_products > 0:
                    # HÀNG 1: Merge "Sản phẩm đáp ứng" từ E1 đến cột cuối
                    row1 = 1
                    start_col = 5  # Cột E
                    end_col = 4 + max_num_products * 2  # Mỗi sản phẩm có 2 cột
                    start_col_letter = get_column_letter(start_col)
                    end_col_letter = get_column_letter(end_col)
                    
                    if end_col > start_col:
                        ws.merge_cells(f'{start_col_letter}{row1}:{end_col_letter}{row1}')
                    
                    cell_sp = ws.cell(row=row1, column=start_col)
                    cell_sp.value = "Sản phẩm đáp ứng"
                    cell_sp.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell_sp.font = Font(name="Times New Roman", size=12, bold=True)
                    if ws.cell(row=1, column=5).fill and ws.cell(row=1, column=5).fill.fill_type:
                        cell_sp.fill = ws.cell(row=1, column=5).fill.copy()
                    cell_sp.border = self.thin_border
                    
                    # THÊM BORDER CHO TẤT CẢ CÁC Ô TRONG HÀNG 1
                    for c in range(1, end_col + 1):
                        ws.cell(row=1, column=c).border = self.thin_border
                
                # HÀNG 2 và 3: Thiết lập header cho từng sản phẩm tham khảo
                for i in range(max_num_products):
                    col = 5 + i * 2  # Cột chỉ tiêu
                    ref_col = col + 1  # Cột tham chiếu
                    col_letter = get_column_letter(col)
                    ref_col_letter = get_column_letter(ref_col)
                    
                    # HÀNG 2: Merge "Tham khảo X" qua 2 cột
                    ws.merge_cells(f'{col_letter}2:{ref_col_letter}2')
                    cell_thamkhao = ws.cell(row=2, column=col)
                    cell_thamkhao.value = f"Tham khảo {i+1}"
                    cell_thamkhao.font = Font(name="Times New Roman", size=12, bold=True)
                    cell_thamkhao.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    if ws.cell(row=2, column=5).fill and ws.cell(row=2, column=5).fill.fill_type:
                        cell_thamkhao.fill = ws.cell(row=2, column=5).fill.copy()
                    cell_thamkhao.border = self.thin_border
                    ws.cell(row=2, column=ref_col).border = self.thin_border
                    
                    # HÀNG 3: "Chỉ tiêu kỹ thuật" và "Tham chiếu"
                    cell_chitieu = ws.cell(row=3, column=col)
                    cell_chitieu.value = "Chỉ tiêu kỹ thuật"
                    cell_chitieu.font = Font(name="Times New Roman", size=12, bold=True)
                    cell_chitieu.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    if ws.cell(row=3, column=5).fill and ws.cell(row=3, column=5).fill.fill_type:
                        cell_chitieu.fill = ws.cell(row=3, column=5).fill.copy()
                    cell_chitieu.border = self.thin_border
                    
                    cell_thamchieu = ws.cell(row=3, column=ref_col)
                    cell_thamchieu.value = "Tham chiếu"
                    cell_thamchieu.font = Font(name="Times New Roman", size=12, bold=True)
                    cell_thamchieu.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    if ws.cell(row=3, column=6).fill and ws.cell(row=3, column=6).fill.fill_type:
                        cell_thamchieu.fill = ws.cell(row=3, column=6).fill.copy()
                    cell_thamchieu.border = self.thin_border
                
                # THÊM BORDER CHO CÁC Ô CỐ ĐỊNH (cột A-D) TRONG HÀNG 2, 3
                for row in [2, 3]:
                    for c in range(1, 5):
                        cell = ws.cell(row=row, column=c)
                        cell.border = self.thin_border
                        
            else:
                # SẢN PHẨM THỨ 2 TRỞ ĐI: Copy từ template
                max_col_to_copy = 4 + num_products * 2
                
                for src_r, tgt_r in [(4, product_row), (5, manufacturer_row), (6, detail_row)]:
                    for c in range(1, max_col_to_copy + 1):
                        source_cell = ws.cell(row=src_r, column=c)
                        target_cell = ws.cell(row=tgt_r, column=c)
                        target_cell.value = source_cell.value if source_cell.value is not None else ""
                        if source_cell.has_style:
                            target_cell.font = source_cell.font.copy() if source_cell.font else Font(name="Times New Roman", size=12)
                            target_cell.border = source_cell.border.copy() if source_cell.border else self.thin_border
                            target_cell.alignment = source_cell.alignment.copy() if source_cell.alignment else Alignment(wrap_text=True)
                        else:
                            target_cell.font = Font(name="Times New Roman", size=12)
                            target_cell.border = self.thin_border
                            target_cell.alignment = Alignment(wrap_text=True)
                
                # Để trống các cột thừa VÀ THÊM BORDER
                for extra_c in range(max_col_to_copy + 1, 5 + max_num_products * 2):
                    for tgt_r in [product_row, manufacturer_row, detail_row]:
                        cell = ws.cell(row=tgt_r, column=extra_c)
                        cell.value = ""
                        cell.font = Font(name="Times New Roman", size=12)
                        cell.border = self.thin_border
                        cell.alignment = Alignment(wrap_text=True)
                
                # Áp dụng fill
                fill_product = self.template_product_fill or PatternFill(fill_type=None)
                for c in range(1, max_col_to_copy + 1):
                    if c >= 5 and (c - 5) % 2 == 1:
                        continue
                    if fill_product and fill_product.fill_type:
                        ws.cell(row=product_row, column=c).fill = fill_product

                fill_manufacturer = self.template_manufacturer_fill or PatternFill(fill_type=None)
                for c in range(1, max_col_to_copy + 1):
                    if fill_manufacturer and fill_manufacturer.fill_type:
                        ws.cell(row=manufacturer_row, column=c).fill = fill_manufacturer
                
                fill_detail = self.template_detail_fill or PatternFill(fill_type=None)
                for c in range(1, max_col_to_copy + 1):
                    if fill_detail and fill_detail.fill_type:
                        ws.cell(row=detail_row, column=c).fill = fill_detail

            # Ghi số La Mã
            cell = ws.cell(row=product_row, column=1)
            cell.value = int_to_roman(product_index)
            cell.font = Font(name="Times New Roman", size=12, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.thin_border

            # Ghi tên sản phẩm
            cell = ws.cell(row=product_row, column=2)
            cell.value = product_name if product_name else "Unnamed Product"
            cell.font = Font(name="Times New Roman", size=12, bold=True)
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            cell.border = self.thin_border
            
            # THÊM BORDER CHO CỘT 3, 4 TRONG HÀNG PRODUCT
            for c in [3, 4]:
                ws.cell(row=product_row, column=c).border = self.thin_border
                ws.cell(row=manufacturer_row, column=c).border = self.thin_border
                ws.cell(row=detail_row, column=c).border = self.thin_border

            # Ghi thông tin sản phẩm tham khảo
            for i, (man_id, name, prod_name) in enumerate(reference_products or []):
                col = 5 + i * 2  # Cột chỉ tiêu
                ref_col = col + 1  # Cột tham chiếu
                ref_col_letter = get_column_letter(ref_col)
                
                # Ghi tên sản phẩm tham khảo
                cell_prod = ws.cell(row=product_row, column=col)
                cell_prod.value = prod_name or 'Không có tên sản phẩm'
                cell_prod.font = Font(name="Times New Roman", size=12, bold=True)
                cell_prod.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell_prod.border = self.thin_border
                
                # Ghi giá trị tham chiếu cho tên sản phẩm
                ref_key = f"ref_value_{man_id}_product_name"
                ref_value = custom_indicators.get("three_brands", {}).get(ref_key, "")
                cell_ref = ws.cell(row=product_row, column=ref_col)
                cell_ref.value = ref_value if ref_value is not None else ""
                cell_ref.font = Font(name="Times New Roman", size=12)
                cell_ref.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell_ref.border = self.thin_border
                
                # Ghi tên hãng
                cell_manu = ws.cell(row=manufacturer_row, column=col)
                cell_manu.value = name or 'Không có tên hãng'
                cell_manu.font = Font(name="Times New Roman", size=12, bold=True)
                cell_manu.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell_manu.border = self.thin_border
                
                # ĐẢM BẢO Ô MANUFACTURER Ở CỘT THAM CHIẾU CÓ BORDER
                ws.cell(row=manufacturer_row, column=ref_col).border = self.thin_border
                
                # MERGE cột tham chiếu từ product_row đến manufacturer_row
                try:
                    ws.merge_cells(f'{ref_col_letter}{product_row}:{ref_col_letter}{manufacturer_row}')
                except:
                    pass

            # Ghi "Chỉ tiêu kỹ thuật chi tiết"
            cell = ws.cell(row=detail_row, column=2)
            cell.value = "Chỉ tiêu kỹ thuật chi tiết"
            cell.font = Font(name="Times New Roman", size=12, bold=True)
            cell.alignment = Alignment(horizontal='left', vertical='top')
            cell.border = self.thin_border

            # Ghi dữ liệu chi tiết
            data_start = detail_row + 1
            max_col_all = 4 + max_num_products * 2  # Cột lớn nhất cần border
            
            for idx, row_data in enumerate(data_rows):
                if not row_data or len(row_data) < 7:
                    continue
                r = data_start + idx
                
                # Cột 1: STT
                cell = ws.cell(row=r, column=1)
                cell.value = row_data[0] if row_data[0] is not None else ""
                cell.font = Font(name="Times New Roman", size=12)
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                cell.border = self.thin_border
                
                # Cột 2: Yêu cầu
                cell = ws.cell(row=r, column=2)
                cell.value = str(row_data[1]).replace('\n', ' ') if row_data[1] is not None else ""
                cell.font = Font(name="Times New Roman", size=12)
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                cell.border = self.thin_border
                
                # Cột 3: Chỉ tiêu (gộp)
                chi_tieu = str(row_data[2]).replace('\n', ' ') if row_data[2] is not None else ""
                danh_gia = str(row_data[3]) if row_data[3] is not None else "not"
                so_sanh = str(row_data[4]).replace('\n', ' ') if row_data[4] is not None else ""
                
                if tab_name == "three_brands":
                    don_vi = str(row_data[6]).replace('\n', ' ') if len(row_data) > 6 and row_data[6] is not None else ""
                else:
                    don_vi = str(row_data[5]).replace('\n', ' ') if len(row_data) > 5 and row_data[5] is not None else ""
                
                compare_symbols = {"<=": "≤", ">=": "≥", "=": "=", "<": "<", ">": ">"}
                compare_symbol = compare_symbols.get(danh_gia, "") if danh_gia != "not" else ""
                is_numeric = self.is_numeric_value(so_sanh)
                chi_tieu_gop = f"{chi_tieu} {compare_symbol} {so_sanh} {don_vi}".strip() if is_numeric and compare_symbol else f"{chi_tieu} {so_sanh} {don_vi}".strip()
                
                cell = ws.cell(row=r, column=3)
                cell.value = chi_tieu_gop if chi_tieu_gop else ""
                cell.font = Font(name="Times New Roman", size=12)
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                cell.border = self.thin_border
                
                # Cột 4: Loại chỉ tiêu
                crit_type_idx = 7 if tab_name == "three_brands" else 6
                cell = ws.cell(row=r, column=4)
                cell.value = str(row_data[crit_type_idx]).replace('\n', ' ') if len(row_data) > crit_type_idx and row_data[crit_type_idx] is not None else ""
                cell.font = Font(name="Times New Roman", size=12)
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                cell.border = self.thin_border
                
                # Các cột sản phẩm tham khảo (có dữ liệu)
                ref_start = 8 if tab_name == "three_brands" else 7
                ref_col_offset = 0
                for val in row_data[ref_start:ref_start + num_products * 2] if len(row_data) > ref_start else []:
                    col = 5 + ref_col_offset
                    cell = ws.cell(row=r, column=col)
                    cell.value = str(val).replace('\n', ' ') if val is not None else ""
                    cell.font = Font(name="Times New Roman", size=12)
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    cell.border = self.thin_border
                    ref_col_offset += 1
                
                # THÊM BORDER CHO TẤT CẢ CÁC CỘT TRỐNG (từ num_products đến max_num_products)
                for extra_col in range(5 + num_products * 2, max_col_all + 1):
                    cell = ws.cell(row=r, column=extra_col)
                    cell.value = ""
                    cell.font = Font(name="Times New Roman", size=12)
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    cell.border = self.thin_border
                
                # Kiểm tra và bôi màu xanh
                ind_id = row_data[-1] if row_data and len(row_data) > 0 else None
                if ind_id and self.should_mark_blue_for_export(ind_id, so_sanh, tab_name):
                    max_col_with_data = 4 + num_products * 2
                    for c in range(1, max_col_with_data + 1):
                        cell = ws.cell(row=r, column=c)
                        cell.fill = self.blue_fill
                
                # Tính chiều cao hàng
                max_lines = max([str(v).count('\n') + 1 for v in row_data if isinstance(v, str) and v is not None] or [1])
                ws.row_dimensions[r].height = 15 * max_lines if max_lines > 0 else 15

            # *** BÔI MÀU XANH LÁ CHO HÀNG TÊN SẢN PHẨM THAM KHẢO ***
            # Áp dụng cho TẤT CẢ các sản phẩm (bao gồm cả sản phẩm đầu tiên)
            for i, (man_id, name, prod_name) in enumerate(reference_products or []):
                col = 5 + i * 2  # Cột chỉ tiêu (tên sản phẩm)
                cell_prod = ws.cell(row=product_row, column=col)
                # Bôi màu xanh lá cho ô tên sản phẩm tham khảo
                cell_prod.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")

            # Merge cột B cho các hàng có cùng giá trị
            data_end = data_start + len(data_rows) - 1
            if data_rows and data_start <= data_end:
                current_r = data_start
                while current_r <= data_end:
                    start_merge = current_r
                    current_value = ws[f'B{current_r}'].value if ws[f'B{current_r}'].value is not None else ""
                    current_r += 1
                    while current_r <= data_end and ws[f'B{current_r}'].value == current_value:
                        current_r += 1
                    if current_r - start_merge > 1 and current_value:
                        try:
                            ws.merge_cells(f'B{start_merge}:B{current_r-1}')
                            ws[f'B{start_merge}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                        except:
                            pass

            # ĐẢM BẢO TẤT CẢ CÁC Ô TRONG VÙNG SỬ DỤNG ĐỀU CÓ BORDER
            # Từ hàng 1 đến data_end, từ cột 1 đến max_col_all
            for r in range(1, data_end + 1):
                for c in range(1, max_col_all + 1):
                    cell = ws.cell(row=r, column=c)
                    if not cell.border or cell.border.left.style is None:
                        cell.border = self.thin_border

            return data_start + len(data_rows)
            
        except Exception as e:
            print(f"Lỗi trong export_hang_bom: {str(e)}")
            import traceback
            traceback.print_exc()
            return current_row

    def should_mark_blue_for_export(self, ind_id, current_value, tab_name, is_header=False):
        if tab_name == "dmkt" and is_header:
            if current_value:  # Nếu header có current_value (single no sub), kiểm tra rỗng
                return not str(current_value).strip()
            else:  # Header không có value (có sub), không tô blue
                return False
        if not ind_id:
            return False
        return not str(current_value).strip()

    def export_files(self):
        folder = filedialog.askdirectory(title="Chọn thư mục lưu files")
        if not folder:
            return

        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        products_query = '''
            SELECT 
                p.id,
                p.name,
                GROUP_CONCAT(pt.name, ', ') as product_types
            FROM product_projects pp 
            JOIN products p ON pp.product_id = p.id 
            LEFT JOIN product_type_mapping_products ptmp ON p.id = ptmp.product_id
            LEFT JOIN product_types pt ON ptmp.type_id = pt.id
            WHERE pp.project_id=? 
            GROUP BY p.id, p.name
            ORDER BY p.name
        '''
        products_list = c.execute(products_query, (self.project_id,)).fetchall()
        conn.close()

        if not products_list:
            messagebox.showerror("Lỗi", "Không có sản phẩm nào trong dự án để xuất file")
            return

        max_num_products = max(len(self.load_reference_products(pid)) for pid, _, _ in products_list) if products_list else 0

        project_name = self.project_name.strip() if self.project_name else "Unnamed_Project"
        safe_project_name = re.sub(r'[<>:"/\\|?*]', '', project_name)
        safe_project_name = safe_project_name.strip()[:50] + "_"

        script_dir = os.path.dirname(os.path.abspath(__file__))
        template_path = os.path.join(script_dir, "mau.xlsx")
        if not os.path.exists(template_path):
            messagebox.showerror("Lỗi", "File template 'mau.xlsx' không tồn tại trong thư mục script!")
            return

        # Xuất file 3 hãng
        try:
            wb_3hang = openpyxl.load_workbook(template_path)
            ws_3hang = wb_3hang.active
            is_first = True
            current_row_3hang = 4
            product_counter = 1
            for product in products_list:
                product_id, product_name, product_types = product
                reference_products = self.load_reference_products(product_id) or []
                deleted_indicators, _ = self.load_hidden_indicators(product_id)
                custom_indicators, _ = self.load_custom_indicators(product_id)
                data_rows = self.load_three_brands_data(product_id, product_types or "", reference_products, deleted_indicators, custom_indicators) or []
                current_row_3hang = self.export_hang_bom(ws_3hang, current_row_3hang, product_name, reference_products, data_rows, "three_brands", is_first, product_counter, max_num_products, custom_indicators)
                is_first = False
                product_counter += 1
            output_path_3hang = os.path.join(folder, f"{safe_project_name}_3 hãng.xlsx")
            wb_3hang.save(output_path_3hang)
        except Exception as e:
            messagebox.showerror("Lỗi", f"Lỗi khi xuất file 3 hãng: {str(e)}")

        # Xuất file BOM
        try:
            wb_bom = openpyxl.load_workbook(template_path)
            ws_bom = wb_bom.active
            is_first = True
            current_row_bom = 4
            product_counter = 1
            for product in products_list:
                product_id, product_name, product_types = product
                reference_products = self.load_reference_products(product_id) or []
                deleted_indicators, _ = self.load_hidden_indicators(product_id)
                custom_indicators, _ = self.load_custom_indicators(product_id)
                data_rows = self.load_bom_data(product_id, product_types or "", reference_products, deleted_indicators, custom_indicators) or []
                current_row_bom = self.export_hang_bom(ws_bom, current_row_bom, product_name, reference_products, data_rows, "bom", is_first, product_counter, max_num_products, custom_indicators)
                is_first = False
                product_counter += 1
            output_path_bom = os.path.join(folder, f"{safe_project_name}_BOM.xlsx")
            wb_bom.save(output_path_bom)
        except Exception as e:
            messagebox.showerror("Lỗi", f"Lỗi khi xuất file BOM: {str(e)}")

        # ==========================================
        # ========== EXPORT FILE DMKT ==============
        # ==========================================

        wb_dmkt = openpyxl.Workbook()
        ws_dmkt = wb_dmkt.active

        # Màu header (giống CTKT mua sắm)
        header_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")

        # Màu cho dòng tên sản phẩm
        product_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")

        # ===== Tạo header =====
        ws_dmkt['A1'] = "TT"
        ws_dmkt['B1'] = "Yêu cầu kỹ thuật"

        for col in ['A1', 'B1']:
            cell = ws_dmkt[col]
            cell.font = Font(name="Times New Roman", size=12, bold=True)
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            cell.border = self.thin_border
            cell.fill = header_fill

        current_row_dmkt = 2
        product_stt = 1
        roman_numerals = ['I','II','III','IV','V','VI','VII','VIII','IX','X']

        # ===== Ghi từng sản phẩm =====
        for product in products_list:
            product_id, product_name, product_types = product

            deleted_indicators, _ = self.load_hidden_indicators(product_id)
            custom_indicators, _ = self.load_custom_indicators(product_id)

            data_rows = self.load_dmkt_data(product_id, product_types or "", deleted_indicators, custom_indicators)

            # Tạo STT dạng La Mã
            stt = roman_numerals[product_stt - 1] if product_stt <= 10 else str(product_stt)

            # ===== Dòng tên sản phẩm (tô xanh lá) =====
            ws_dmkt[f"A{current_row_dmkt}"] = stt
            ws_dmkt[f"B{current_row_dmkt}"] = product_name

            for col in ['A', 'B']:
                cell = ws_dmkt[f"{col}{current_row_dmkt}"]
                cell.font = Font(name="Times New Roman", size=12, bold=True)
                cell.alignment = Alignment(wrap_text=True, vertical='center')
                cell.border = self.thin_border
                cell.fill = product_fill

            current_row_dmkt += 1

            # ===== Ghi từng chỉ tiêu =====
            for row_data in data_rows:
                ws_dmkt[f"A{current_row_dmkt}"] = row_data[0]

                if str(row_data[1]).startswith("- "):
                    ws_dmkt[f"B{current_row_dmkt}"] = row_data[1]
                    so_sanh = str(row_data[2]).replace('\n', ' ') if len(row_data) > 2 else ""
                    ind_id = row_data[5] if len(row_data) > 5 else None

                    if self.should_mark_blue_for_export(ind_id, so_sanh, "dmkt"):
                        ws_dmkt[f"A{current_row_dmkt}"].fill = self.blue_fill
                        ws_dmkt[f"B{current_row_dmkt}"].fill = self.blue_fill

                else:
                    so_sanh = str(row_data[2]).replace('\n', ' ')
                    don_vi = str(row_data[3]).replace('\n', ' ')
                    danh_gia = str(row_data[6]) if len(row_data) > 6 else "not"

                    compare_symbols = {"<=": "≤", ">=": "≥", "=": "=", "<": "<", ">": ">"}
                    compare_symbol = compare_symbols.get(danh_gia, "") if danh_gia != "not" else ""

                    is_numeric = self.is_numeric_value(so_sanh)

                    yeu_cau = (
                        f"+ {row_data[1][2:]} {compare_symbol} {so_sanh} {don_vi}".strip()
                        if is_numeric and compare_symbol
                        else f"+ {row_data[1][2:]} {so_sanh} {don_vi}".strip()
                    )

                    ws_dmkt[f"B{current_row_dmkt}"] = yeu_cau

                    ind_id = row_data[5] if len(row_data) > 5 else None
                    if self.should_mark_blue_for_export(ind_id, so_sanh, "dmkt"):
                        ws_dmkt[f"A{current_row_dmkt}"].fill = self.blue_fill
                        ws_dmkt[f"B{current_row_dmkt}"].fill = self.blue_fill

                # Giao diện chung
                for col in ['A', 'B']:
                    cell = ws_dmkt[f"{col}{current_row_dmkt}"]
                    cell.font = Font(name="Times New Roman", size=12)
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    cell.border = self.thin_border

                # Auto height
                lines = ws_dmkt[f"B{current_row_dmkt}"].value.count('\n') + 1 if ws_dmkt[f"B{current_row_dmkt}"].value else 1
                ws_dmkt.row_dimensions[current_row_dmkt].height = 15 * lines

                current_row_dmkt += 1

            product_stt += 1

        ws_dmkt.column_dimensions['A'].width = 8
        ws_dmkt.column_dimensions['B'].width = 50

        wb_dmkt.save(os.path.join(folder, f"{safe_project_name}_DMKT.xlsx"))

        # =================================================
        # =============== EXPORT CTKT BỘ ==================
        # =================================================

        wb_ctkt_bo = openpyxl.Workbook()
        ws_ctkt_bo = wb_ctkt_bo.active

        # Header fill + product fill giống CTKT Mua Sắm
        header_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
        product_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")

        headers = ["TT", "Chỉ tiêu", "Đơn vị", "Giá trị"]
        for col, header in zip(['A1','B1','C1','D1'], headers):
            ws_ctkt_bo[col] = header
            ws_ctkt_bo[col].font = Font(name="Times New Roman", size=12, bold=True)
            ws_ctkt_bo[col].alignment = Alignment(wrap_text=True, vertical="center")
            ws_ctkt_bo[col].border = self.thin_border
            ws_ctkt_bo[col].fill = header_fill

        current_row_ctkt_bo = 2
        product_stt = 1

        for product in products_list:
            product_id, product_name, product_types = product
            deleted_indicators, _ = self.load_hidden_indicators(product_id)
            custom_indicators, _ = self.load_custom_indicators(product_id)

            data_rows = self.load_ctkt_bo_data(product_id, product_types, deleted_indicators, custom_indicators)

            stt = roman_numerals[product_stt - 1] if product_stt <= 10 else str(product_stt)

            # ===== DÒNG TÊN SẢN PHẨM TÔ MÀU XANH LÁ =====
            ws_ctkt_bo[f"A{current_row_ctkt_bo}"] = stt
            ws_ctkt_bo[f"B{current_row_ctkt_bo}"] = product_name

            for col in range(1, 5):
                cell = ws_ctkt_bo.cell(row=current_row_ctkt_bo, column=col)
                cell.font = Font(name="Times New Roman", size=12, bold=True)
                cell.alignment = Alignment(wrap_text=True, vertical="center")
                cell.border = self.thin_border
                cell.fill = product_fill

            current_row_ctkt_bo += 1

            # ===== Ghi dữ liệu từng chỉ tiêu =====
            for row_data in data_rows:
                max_lines = 1

                for i, val in enumerate(row_data[:4]):
                    cell = ws_ctkt_bo.cell(row=current_row_ctkt_bo, column=i + 1)

                    text = str(val).replace('\n', ' ') if val else ""
                    cell.value = text
                    cell.font = Font(name="Times New Roman", size=12)
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    cell.border = self.thin_border

                    max_lines = max(max_lines, text.count('\n') + 1)

                ws_ctkt_bo.row_dimensions[current_row_ctkt_bo].height = 15 * max_lines

                ind_id = row_data[-1]
                gia_tri = str(row_data[3]).replace('\n', ' ') if row_data[3] else ""

                if self.should_mark_blue_for_export(ind_id, gia_tri, "ctkt_bo"):
                    for col in range(1, 5):
                        ws_ctkt_bo.cell(row=current_row_ctkt_bo, column=col).fill = self.blue_fill

                current_row_ctkt_bo += 1

            product_stt += 1

        ws_ctkt_bo.column_dimensions['A'].width = 8
        ws_ctkt_bo.column_dimensions['B'].width = 30
        ws_ctkt_bo.column_dimensions['C'].width = 15
        ws_ctkt_bo.column_dimensions['D'].width = 20

        wb_ctkt_bo.save(os.path.join(folder, f"{safe_project_name}_Ctkt bộ.xlsx"))

        # Xuất file CTKT mua sắm
        wb_ctkt_ms = openpyxl.Workbook()
        ws_ctkt_ms = wb_ctkt_ms.active

        # Thiết lập header với màu nền xanh nhạt
        headers = ["TT", "Yêu cầu kỹ thuật", "Tiêu chí đánh giá", "Loại chỉ tiêu\n(CTCB/\nKCB)"]
        header_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")

        for col, header in zip(['A1','B1','C1','D1'], headers):
            ws_ctkt_ms[col] = header
            ws_ctkt_ms[col].font = Font(name="Times New Roman", size=12, bold=True)
            ws_ctkt_ms[col].alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
            ws_ctkt_ms[col].border = self.thin_border
            ws_ctkt_ms[col].fill = header_fill

        current_row_ctkt_ms = 2
        product_stt = 1
        for product in products_list:
            product_id, product_name, product_types = product
            deleted_indicators, _ = self.load_hidden_indicators(product_id)
            custom_indicators, custom_rows_ctkt_ms = self.load_custom_indicators(product_id)
            data_rows = self.load_ctkt_mua_sam_data(product_id, product_types, deleted_indicators, custom_indicators, custom_rows_ctkt_ms)
            
            # Header sản phẩm với màu xanh lá
            stt = roman_numerals[product_stt - 1] if product_stt <= 10 else str(product_stt)
            product_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
            
            ws_ctkt_ms[f'A{current_row_ctkt_ms}'] = stt
            ws_ctkt_ms[f'A{current_row_ctkt_ms}'].font = Font(name="Times New Roman", size=12, bold=True)
            ws_ctkt_ms[f'A{current_row_ctkt_ms}'].fill = product_fill
            
            ws_ctkt_ms[f'B{current_row_ctkt_ms}'] = product_name
            ws_ctkt_ms[f'B{current_row_ctkt_ms}'].font = Font(name="Times New Roman", size=12, bold=True)
            ws_ctkt_ms[f'B{current_row_ctkt_ms}'].fill = product_fill
            
            for col in range(1, 5):
                cell = ws_ctkt_ms.cell(row=current_row_ctkt_ms, column=col)
                cell.border = self.thin_border
                if col > 2:
                    cell.fill = product_fill
            current_row_ctkt_ms += 1
            
            # Dòng "Chỉ tiêu kỹ thuật chi tiết"
            ws_ctkt_ms[f'B{current_row_ctkt_ms}'] = "Chỉ tiêu kỹ thuật chi tiết"
            ws_ctkt_ms[f'B{current_row_ctkt_ms}'].font = Font(name="Times New Roman", size=12, bold=True)
            for col in range(1, 5):
                cell = ws_ctkt_ms.cell(row=current_row_ctkt_ms, column=col)
                cell.border = self.thin_border
            current_row_ctkt_ms += 1
            
            req_number = 0
            sub_stt = 0
            is_yeu_cau_khac_mode = False
            last_chi_tieu = None
            
            for row_data in data_rows:
                chi_tieu_display = str(row_data[1]).replace('\n', ' ') if row_data[1] is not None else ""
                yeu_cau_display = str(row_data[2]).replace('\n', ' ') if row_data[2] is not None else ""
                so_sanh = str(row_data[3]).replace('\n', ' ') if row_data[3] is not None else ""
                don_vi = str(row_data[4]).replace('\n', ' ') if row_data[4] is not None else ""
                tieu_chi = str(row_data[5]).replace('\r\n', '\n') if row_data[5] is not None else ""
                loai_chi_tieu = str(row_data[6]).replace('\n', ' ') if row_data[6] is not None else ""
                ind_id = row_data[-1]
                
                # Xử lý "Yêu cầu khác"
                if chi_tieu_display == "Yêu cầu khác":
                    req_number += 1
                    stt_value = str(req_number)
                    ws_ctkt_ms[f'A{current_row_ctkt_ms}'] = stt_value
                    ws_ctkt_ms[f'A{current_row_ctkt_ms}'].font = Font(name="Times New Roman", size=12, bold=False)
                    ws_ctkt_ms[f'B{current_row_ctkt_ms}'] = "Yêu cầu khác"
                    ws_ctkt_ms[f'B{current_row_ctkt_ms}'].font = Font(name="Times New Roman", size=12, bold=False)
                    is_yeu_cau_khac_mode = True
                    sub_stt = 0
                    last_chi_tieu = "Yêu cầu khác"
                    for col in range(1, 5):
                        cell = ws_ctkt_ms.cell(row=current_row_ctkt_ms, column=col)
                        cell.alignment = Alignment(wrap_text=True, vertical="top")
                        cell.border = self.thin_border
                    current_row_ctkt_ms += 1
                    continue
                
                # Kiểm tra nếu là yêu cầu kỹ thuật mới (chi_tieu_display khác rỗng và khác last_chi_tieu)
                if chi_tieu_display and chi_tieu_display != last_chi_tieu:
                    req_number += 1
                    sub_stt = 0
                    last_chi_tieu = chi_tieu_display
                    
                    # Ghi dòng yêu cầu kỹ thuật
                    ws_ctkt_ms[f'A{current_row_ctkt_ms}'] = str(req_number)
                    ws_ctkt_ms[f'A{current_row_ctkt_ms}'].font = Font(name="Times New Roman", size=12, bold=False)
                    ws_ctkt_ms[f'B{current_row_ctkt_ms}'] = chi_tieu_display
                    ws_ctkt_ms[f'B{current_row_ctkt_ms}'].font = Font(name="Times New Roman", size=12, bold=False)
                    
                    for col in range(1, 5):
                        cell = ws_ctkt_ms.cell(row=current_row_ctkt_ms, column=col)
                        cell.alignment = Alignment(wrap_text=True, vertical="top")
                        cell.border = self.thin_border
                    current_row_ctkt_ms += 1
                
                # Ghi dòng chi tiết (chỉ tiêu con hoặc yêu cầu không có chỉ tiêu con)
                sub_stt += 1
                stt_value = f"{req_number}.{sub_stt}"
                
                # Xây dựng cột yêu cầu kỹ thuật
                is_numeric = self.is_numeric_value(so_sanh)
                
                # Lấy danh_gia từ custom_indicators nếu có
                danh_gia = "not"
                if ind_id and not is_yeu_cau_khac_mode:
                    danh_gia = custom_indicators.get("three_brands", {}).get(f"danh_gia_{ind_id}", "not")
                
                # Xây dựng yêu cầu kỹ thuật với ký hiệu so sánh đúng
                if is_numeric and danh_gia != "not":
                    compare_symbols = {"<=": "≤", ">=": "≥", "=": "=", "<": "<", ">": ">"}
                    compare_symbol = compare_symbols.get(danh_gia, "")
                    yeu_cau_export = f"{yeu_cau_display} {compare_symbol} {so_sanh} {don_vi}".strip()
                else:
                    yeu_cau_export = f"{yeu_cau_display} {so_sanh} {don_vi}".strip()
                
                ws_ctkt_ms[f'A{current_row_ctkt_ms}'] = stt_value
                ws_ctkt_ms[f'B{current_row_ctkt_ms}'] = yeu_cau_export if yeu_cau_export else ""
                ws_ctkt_ms[f'C{current_row_ctkt_ms}'] = tieu_chi
                
                # Chỉ điền loại chỉ tiêu nếu là dòng duy nhất (không có sub)
                # Kiểm tra: nếu chi_tieu_display không rỗng và là dòng đầu tiên của nhóm
                ws_ctkt_ms[f'D{current_row_ctkt_ms}'] = loai_chi_tieu if loai_chi_tieu else ""
                
                for col in range(1, 5):
                    cell = ws_ctkt_ms.cell(row=current_row_ctkt_ms, column=col)
                    cell.font = Font(name="Times New Roman", size=12)
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                    cell.border = self.thin_border
                
                # Bôi màu xanh nếu thiếu dữ liệu
                if not is_yeu_cau_khac_mode:
                    if self.should_mark_blue_for_export(ind_id, so_sanh, "ctkt_mua_sam"):
                        for col in range(1, 5):
                            ws_ctkt_ms.cell(row=current_row_ctkt_ms, column=col).fill = self.blue_fill
                
                current_row_ctkt_ms += 1
            
            product_stt += 1

        # Thiết lập độ rộng cột
        ws_ctkt_ms.column_dimensions['A'].width = 8
        ws_ctkt_ms.column_dimensions['B'].width = 50
        ws_ctkt_ms.column_dimensions['C'].width = 60
        ws_ctkt_ms.column_dimensions['D'].width = 15

        wb_ctkt_ms.save(os.path.join(folder, f"{safe_project_name}_Ctkt mua sắm.xlsx"))

        messagebox.showinfo("Thành công", f"Đã xuất 5 files Excel")